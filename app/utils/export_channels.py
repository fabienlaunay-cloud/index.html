"""Multichannel exports — turn SynqIO listings into PIM/e-commerce-ready files.

SynqIO's generated content is channel-agnostic (title, bullets, description,
brand, EAN, price, images…). This module opens the outputs beyond Amazon:

- ``pim``            → universal XLSX, one column per data point (any PIM / ERP)
- ``shopify``        → Shopify product-import CSV (products + image rows)
- ``woocommerce``    → WooCommerce built-in product importer CSV
- ``prestashop``     → PrestaShop product-import CSV (";" separated, FR default)
- ``akeneo``         → Akeneo PIM CSV (";" separated, locale-scoped attributes)
- ``google-merchant``→ Google Merchant Center feed (TSV)

Every builder takes the same inputs: the listings plus the per-SKU generated
image URLs ({sku: {img_id: url}}) as returned by _get_image_urls_for_skus.
Parent variation rows (is_parent=True) carry no sellable content of their own
and are skipped; children export as standalone products with their variation
value kept in the name/attributes.
"""
import csv
import io
import re
from typing import List, Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from app.models import AmazonListing
from app.utils.export import _IMAGE_IDS_ORDERED, _strip_html

# Currency by marketplace — default EUR (all supported marketplaces are EU)
_CURRENCY = {"amazon_uk": "GBP", "amazon_se": "SEK", "amazon_pl": "PLN"}

# Akeneo locale code by marketplace
_LOCALE = {
    "amazon_fr": "fr_FR", "amazon_de": "de_DE", "amazon_it": "it_IT",
    "amazon_es": "es_ES", "amazon_uk": "en_GB", "amazon_nl": "nl_NL",
    "amazon_se": "sv_SE", "amazon_pl": "pl_PL", "amazon_be": "fr_BE",
}


def _currency(listing: AmazonListing) -> str:
    return _CURRENCY.get(listing.marketplace.value, "EUR")


def _locale(listing: AmazonListing) -> str:
    return _LOCALE.get(listing.marketplace.value, "fr_FR")


def _images_ordered(sku: str, image_urls: Optional[dict]) -> list:
    """Return this SKU's generated image URLs in canonical slot order."""
    imgs = (image_urls or {}).get(sku) or {}
    return [imgs[i] for i in _IMAGE_IDS_ORDERED if i in imgs]


def _sellable(listings: List[AmazonListing]) -> List[AmazonListing]:
    return [l for l in listings if not l.is_parent]


def _slug(text: str) -> str:
    s = re.sub(r"[^a-z0-9]+", "-", (text or "").lower())
    return s.strip("-")[:70]


def _bullets_html(listing: AmazonListing) -> str:
    """Bullets as an HTML <ul> — reused in Shopify/Woo/Presta descriptions."""
    items = "".join(f"<li>{b}</li>" for b in listing.bullet_points if b)
    return f"<ul>{items}</ul>" if items else ""


def _full_html_description(listing: AmazonListing) -> str:
    """Description + bullet list, ready for platforms with a single HTML body."""
    parts = []
    if listing.description:
        parts.append(listing.description)
    ul = _bullets_html(listing)
    if ul:
        parts.append(ul)
    return "<br><br>".join(parts)


def _short_description(listing: AmazonListing) -> str:
    return getattr(listing, "item_highlights", "") or (listing.bullet_points[0] if listing.bullet_points else "")


def _oneline(text: str) -> str:
    """Feed-safe single-line plain text (no tabs / newlines / HTML)."""
    no_html = re.sub(r"<[^>]+>", " ", text or "")  # tags → space so words don't merge
    return re.sub(r"[\s\t\r\n]+", " ", no_html).strip()


# ── 1. Universal PIM XLSX ────────────────────────────────────────────────────

def to_pim_xlsx(listings: List[AmazonListing], image_urls: Optional[dict] = None) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Produits"

    headers = [
        "sku", "ean", "brand", "category", "title", "short_description",
        "bullet_1", "bullet_2", "bullet_3", "bullet_4", "bullet_5",
        "description_html", "description_text", "keywords",
        "price", "currency", "color", "size", "material", "weight_kg",
        "language", "marketplace_origin", "seo_score",
        "parent_sku", "variation_theme", "variation_value",
    ] + [f"image_url_{i+1}" for i in range(len(_IMAGE_IDS_ORDERED))]

    ws.append(headers)
    purple = PatternFill("solid", fgColor="764BA2")
    white = Font(color="FFFFFF", bold=True, size=10)
    for col in range(1, len(headers) + 1):
        c = ws.cell(1, col)
        c.font = white
        c.fill = purple
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"

    for l in _sellable(listings):
        imgs = _images_ordered(l.sku, image_urls)
        imgs += [""] * (len(_IMAGE_IDS_ORDERED) - len(imgs))
        ws.append([
            l.sku, l.ean or "", l.brand, l.category, l.title, _short_description(l),
            *( (l.bullet_points + [""] * 5)[:5] ),
            l.description, _strip_html(l.description), l.backend_keywords,
            l.price if l.price is not None else "", _currency(l),
            l.color or "", l.size or "", l.material or "",
            l.weight_kg if l.weight_kg is not None else "",
            _locale(l).split("_")[0], l.marketplace.value, l.seo_score or "",
            l.parent_sku or "", l.variation_theme or "", l.variation_value or "",
            *imgs,
        ])

    # Reasonable column widths for the text-heavy columns
    wide = {"E": 60, "F": 45, "G": 40, "H": 40, "I": 40, "J": 40, "K": 40, "L": 70, "M": 70, "N": 40}
    for col, w in wide.items():
        ws.column_dimensions[col].width = w

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ── 2. Shopify product CSV ───────────────────────────────────────────────────

def to_shopify_csv(listings: List[AmazonListing], image_urls: Optional[dict] = None) -> bytes:
    out = io.StringIO()
    fieldnames = [
        "Handle", "Title", "Body (HTML)", "Vendor", "Type", "Tags", "Published",
        "Option1 Name", "Option1 Value",
        "Variant SKU", "Variant Grams", "Variant Inventory Policy",
        "Variant Fulfillment Service", "Variant Price", "Variant Requires Shipping",
        "Variant Taxable", "Variant Barcode",
        "Image Src", "Image Position", "Image Alt Text",
        "SEO Title", "SEO Description", "Status",
    ]
    w = csv.DictWriter(out, fieldnames=fieldnames)
    w.writeheader()

    for l in _sellable(listings):
        handle = _slug(f"{l.title} {l.sku}") or _slug(l.sku)
        imgs = _images_ordered(l.sku, image_urls)
        keywords = ", ".join(k for k in (l.backend_keywords or "").split() if k)[:255]
        w.writerow({
            "Handle": handle,
            "Title": l.title,
            "Body (HTML)": _full_html_description(l),
            "Vendor": l.brand,
            "Type": l.category,
            "Tags": keywords,
            "Published": "TRUE",
            "Option1 Name": "Title" if not l.variation_value else (l.variation_theme or "Variante"),
            "Option1 Value": l.variation_value or "Default Title",
            "Variant SKU": l.sku,
            "Variant Grams": int(l.weight_kg * 1000) if l.weight_kg else "",
            "Variant Inventory Policy": "deny",
            "Variant Fulfillment Service": "manual",
            "Variant Price": l.price if l.price is not None else "",
            "Variant Requires Shipping": "TRUE",
            "Variant Taxable": "TRUE",
            "Variant Barcode": l.ean or "",
            "Image Src": imgs[0] if imgs else "",
            "Image Position": 1 if imgs else "",
            "Image Alt Text": l.title if imgs else "",
            "SEO Title": l.title[:70],
            "SEO Description": _oneline(_short_description(l) or l.description)[:320],
            "Status": "active",
        })
        # Additional images: one row per image, only Handle + Image fields
        for pos, url in enumerate(imgs[1:], start=2):
            w.writerow({"Handle": handle, "Image Src": url, "Image Position": pos,
                        "Image Alt Text": l.title})

    return out.getvalue().encode("utf-8-sig")


# ── 3. WooCommerce product CSV ───────────────────────────────────────────────

def to_woocommerce_csv(listings: List[AmazonListing], image_urls: Optional[dict] = None) -> bytes:
    out = io.StringIO()
    fieldnames = [
        "Type", "SKU", "GTIN, UPC, EAN, or ISBN", "Name", "Published",
        "Is featured?", "Visibility in catalog", "Short description", "Description",
        "Regular price", "Categories", "Tags", "Weight (kg)", "Images",
        "Attribute 1 name", "Attribute 1 value(s)",
        "Attribute 2 name", "Attribute 2 value(s)",
        "Meta: brand",
    ]
    w = csv.DictWriter(out, fieldnames=fieldnames)
    w.writeheader()

    for l in _sellable(listings):
        imgs = _images_ordered(l.sku, image_urls)
        keywords = ", ".join((l.backend_keywords or "").split())
        attr1 = ("Marque", l.brand) if l.brand else ("", "")
        attr2 = ("", "")
        if l.color:
            attr2 = ("Couleur", l.color)
        elif l.material:
            attr2 = ("Matière", l.material)
        w.writerow({
            "Type": "simple",
            "SKU": l.sku,
            "GTIN, UPC, EAN, or ISBN": l.ean or "",
            "Name": l.title,
            "Published": 1,
            "Is featured?": 0,
            "Visibility in catalog": "visible",
            "Short description": _short_description(l),
            "Description": _full_html_description(l),
            "Regular price": l.price if l.price is not None else "",
            "Categories": l.category,
            "Tags": keywords,
            "Weight (kg)": l.weight_kg if l.weight_kg is not None else "",
            "Images": ", ".join(imgs),
            "Attribute 1 name": attr1[0], "Attribute 1 value(s)": attr1[1],
            "Attribute 2 name": attr2[0], "Attribute 2 value(s)": attr2[1],
            "Meta: brand": l.brand,
        })

    return out.getvalue().encode("utf-8-sig")


# ── 4. PrestaShop product CSV (";" separated — PrestaShop default) ───────────

def to_prestashop_csv(listings: List[AmazonListing], image_urls: Optional[dict] = None) -> bytes:
    out = io.StringIO()
    fieldnames = [
        "Reference #", "Active (0/1)", "Name", "Categories (x,y,z...)",
        "Price tax included", "Brand", "EAN13", "Summary", "Description",
        "Tags (x,y,z...)", "Weight", "Meta title", "Meta description",
        "Image URLs (x,y,z...)",
    ]
    w = csv.DictWriter(out, fieldnames=fieldnames, delimiter=";")
    w.writeheader()

    for l in _sellable(listings):
        imgs = _images_ordered(l.sku, image_urls)
        keywords = ",".join((l.backend_keywords or "").split())
        w.writerow({
            "Reference #": l.sku,
            "Active (0/1)": 1,
            "Name": l.title,
            "Categories (x,y,z...)": l.category,
            "Price tax included": l.price if l.price is not None else "",
            "Brand": l.brand,
            "EAN13": l.ean or "",
            "Summary": _short_description(l),
            "Description": _full_html_description(l),
            "Tags (x,y,z...)": keywords,
            "Weight": l.weight_kg if l.weight_kg is not None else "",
            "Meta title": l.title[:70],
            "Meta description": _oneline(_short_description(l) or l.description)[:160],
            "Image URLs (x,y,z...)": ",".join(imgs),
        })

    return out.getvalue().encode("utf-8-sig")


# ── 5. Akeneo PIM CSV (";" separated, locale-scoped attributes) ──────────────

def to_akeneo_csv(listings: List[AmazonListing], image_urls: Optional[dict] = None) -> bytes:
    sellable = _sellable(listings)
    # One locale per export keeps columns consistent; use the first listing's
    locale = _locale(sellable[0]) if sellable else "fr_FR"
    currency = _currency(sellable[0]) if sellable else "EUR"

    out = io.StringIO()
    fieldnames = [
        "sku", "enabled", "family", "categories",
        f"name-{locale}", f"short_description-{locale}", f"description-{locale}",
        f"bullet_1-{locale}", f"bullet_2-{locale}", f"bullet_3-{locale}",
        f"bullet_4-{locale}", f"bullet_5-{locale}",
        f"keywords-{locale}",
        f"price-{currency}", "brand", "ean", "color", "size", "material",
        "weight", "image_urls",
    ]
    w = csv.DictWriter(out, fieldnames=fieldnames, delimiter=";")
    w.writeheader()

    for l in sellable:
        imgs = _images_ordered(l.sku, image_urls)
        bullets = (l.bullet_points + [""] * 5)[:5]
        w.writerow({
            "sku": l.sku,
            "enabled": 1,
            "family": _slug(l.category).replace("-", "_") or "default",
            "categories": _slug(l.category).replace("-", "_"),
            f"name-{locale}": l.title,
            f"short_description-{locale}": _short_description(l),
            f"description-{locale}": l.description,
            f"bullet_1-{locale}": bullets[0], f"bullet_2-{locale}": bullets[1],
            f"bullet_3-{locale}": bullets[2], f"bullet_4-{locale}": bullets[3],
            f"bullet_5-{locale}": bullets[4],
            f"keywords-{locale}": l.backend_keywords,
            f"price-{currency}": l.price if l.price is not None else "",
            "brand": l.brand, "ean": l.ean or "",
            "color": l.color or "", "size": l.size or "", "material": l.material or "",
            "weight": l.weight_kg if l.weight_kg is not None else "",
            "image_urls": ",".join(imgs),
        })

    return out.getvalue().encode("utf-8-sig")


# ── 6. Google Merchant Center feed (TSV) ─────────────────────────────────────

def to_google_merchant_tsv(listings: List[AmazonListing], image_urls: Optional[dict] = None) -> bytes:
    out = io.StringIO()
    headers = [
        "id", "title", "description", "link", "image_link", "additional_image_link",
        "availability", "price", "brand", "gtin", "condition",
        "item_group_id", "color", "size", "material", "shipping_weight",
    ]
    out.write("\t".join(headers) + "\n")

    for l in _sellable(listings):
        imgs = _images_ordered(l.sku, image_urls)
        price = f"{l.price:.2f} {_currency(l)}" if l.price is not None else ""
        row = [
            l.sku,
            _oneline(l.title)[:150],
            _oneline(f"{l.description} {' '.join(l.bullet_points)}")[:5000],
            "",  # link: the merchant's own product URL — to fill on their site
            imgs[0] if imgs else "",
            ",".join(imgs[1:]) if len(imgs) > 1 else "",
            "in stock",
            price,
            _oneline(l.brand),
            l.ean or "",
            "new",
            l.parent_sku or "",
            _oneline(l.color or ""),
            _oneline(l.size or ""),
            _oneline(l.material or ""),
            f"{l.weight_kg} kg" if l.weight_kg else "",
        ]
        out.write("\t".join(row) + "\n")

    return out.getvalue().encode("utf-8")


# ── 7. Meta (Facebook / Instagram Shopping) catalogue CSV ────────────────────
# Same product-data spec family as Google; Meta Commerce Manager ingests CSV.

def to_meta_csv(listings: List[AmazonListing], image_urls: Optional[dict] = None) -> bytes:
    out = io.StringIO()
    w = csv.DictWriter(out, fieldnames=[
        "id", "title", "description", "availability", "condition", "price",
        "link", "image_link", "additional_image_link", "brand", "gtin",
        "item_group_id", "color", "size", "material",
    ])
    w.writeheader()
    for l in _sellable(listings):
        imgs = _images_ordered(l.sku, image_urls)
        w.writerow({
            "id": l.sku,
            "title": _oneline(l.title)[:200],
            "description": _oneline(f"{l.description} {' '.join(l.bullet_points)}")[:9999],
            "availability": "in stock",
            "condition": "new",
            "price": f"{l.price:.2f} {_currency(l)}" if l.price is not None else "",
            "link": "",  # merchant's own product URL — to fill on their site
            "image_link": imgs[0] if imgs else "",
            "additional_image_link": ",".join(imgs[1:11]) if len(imgs) > 1 else "",
            "brand": _oneline(l.brand),
            "gtin": l.ean or "",
            "item_group_id": l.parent_sku or "",
            "color": _oneline(l.color or ""),
            "size": _oneline(l.size or ""),
            "material": _oneline(l.material or ""),
        })
    return out.getvalue().encode("utf-8")


# ── 8. eBay (Seller Hub / File Exchange CSV) ─────────────────────────────────

def to_ebay_csv(listings: List[AmazonListing], image_urls: Optional[dict] = None) -> bytes:
    out = io.StringIO()
    w = csv.DictWriter(out, fieldnames=[
        "Action", "CustomLabel", "Title", "Subtitle", "Description",
        "PicURL", "Quantity", "StartPrice", "ConditionID",
        "Brand", "EAN", "C:Couleur", "C:Taille", "C:Matière",
    ])
    w.writeheader()
    for l in _sellable(listings):
        imgs = _images_ordered(l.sku, image_urls)
        w.writerow({
            "Action": "Add",
            "CustomLabel": l.sku,                      # eBay SKU
            "Title": _oneline(l.title)[:80],           # eBay caps titles at 80 chars
            "Subtitle": _oneline(_short_description(l))[:55],
            "Description": _full_html_description(l),
            "PicURL": "|".join(imgs[:12]),
            "Quantity": 1,
            "StartPrice": l.price if l.price is not None else "",
            "ConditionID": 1000,                        # 1000 = New
            "Brand": _oneline(l.brand),
            "EAN": l.ean or "",
            "C:Couleur": _oneline(l.color or ""),
            "C:Taille": _oneline(l.size or ""),
            "C:Matière": _oneline(l.material or ""),
        })
    return out.getvalue().encode("utf-8-sig")


# ── 9. Cdiscount (Marketplace CSV) ───────────────────────────────────────────

def to_cdiscount_csv(listings: List[AmazonListing], image_urls: Optional[dict] = None) -> bytes:
    out = io.StringIO()
    w = csv.DictWriter(out, fieldnames=[
        "Sku vendeur", "EAN", "Marque", "Nature du produit", "Catégorie",
        "Libellé produit", "Description courte", "Description longue",
        "Image 1", "Image 2", "Image 3", "Image 4",
        "Prix", "Quantité", "État", "Couleur", "Taille",
    ], delimiter=";")
    w.writeheader()
    for l in _sellable(listings):
        imgs = _images_ordered(l.sku, image_urls) + [""] * 4
        w.writerow({
            "Sku vendeur": l.sku,
            "EAN": l.ean or "",
            "Marque": _oneline(l.brand),
            "Nature du produit": _oneline(l.category or "Standard"),
            "Catégorie": _oneline(l.category or ""),
            "Libellé produit": _oneline(l.title)[:132],   # Cdiscount cap
            "Description courte": _oneline(_short_description(l))[:420],
            "Description longue": _oneline(f"{l.description} {' '.join(l.bullet_points)}")[:5000],
            "Image 1": imgs[0], "Image 2": imgs[1], "Image 3": imgs[2], "Image 4": imgs[3],
            "Prix": l.price if l.price is not None else "",
            "Quantité": 1,
            "État": "Neuf",
            "Couleur": _oneline(l.color or ""),
            "Taille": _oneline(l.size or ""),
        })
    return out.getvalue().encode("utf-8-sig")


# ── 10. TikTok Shop (batch upload CSV) ───────────────────────────────────────

def to_tiktok_csv(listings: List[AmazonListing], image_urls: Optional[dict] = None) -> bytes:
    out = io.StringIO()
    w = csv.DictWriter(out, fieldnames=[
        "seller_sku", "product_name", "product_description", "category",
        "brand", "barcode", "price", "quantity",
        "main_image_url", "other_images", "color", "size",
    ])
    w.writeheader()
    for l in _sellable(listings):
        imgs = _images_ordered(l.sku, image_urls)
        w.writerow({
            "seller_sku": l.sku,
            "product_name": _oneline(l.title)[:255],
            "product_description": _full_html_description(l)[:10000],
            "category": _oneline(l.category or ""),
            "brand": _oneline(l.brand),
            "barcode": l.ean or "",
            "price": l.price if l.price is not None else "",
            "quantity": 1,
            "main_image_url": imgs[0] if imgs else "",
            "other_images": ",".join(imgs[1:9]) if len(imgs) > 1 else "",
            "color": _oneline(l.color or ""),
            "size": _oneline(l.size or ""),
        })
    return out.getvalue().encode("utf-8")


# ── Registry used by the API endpoint ────────────────────────────────────────

CHANNELS = {
    "pim": {
        "builder": to_pim_xlsx,
        "filename": "synqio_produits_pim.xlsx",
        "media_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "label": "PIM / Universel (XLSX)",
    },
    "shopify": {
        "builder": to_shopify_csv,
        "filename": "shopify_produits.csv",
        "media_type": "text/csv",
        "label": "Shopify (CSV)",
    },
    "woocommerce": {
        "builder": to_woocommerce_csv,
        "filename": "woocommerce_produits.csv",
        "media_type": "text/csv",
        "label": "WooCommerce (CSV)",
    },
    "prestashop": {
        "builder": to_prestashop_csv,
        "filename": "prestashop_produits.csv",
        "media_type": "text/csv",
        "label": "PrestaShop (CSV)",
    },
    "akeneo": {
        "builder": to_akeneo_csv,
        "filename": "akeneo_produits.csv",
        "media_type": "text/csv",
        "label": "Akeneo PIM (CSV)",
    },
    "google-merchant": {
        "builder": to_google_merchant_tsv,
        "filename": "google_merchant_feed.tsv",
        "media_type": "text/tab-separated-values",
        "label": "Google Shopping (TSV)",
    },
    "meta": {
        "builder": to_meta_csv,
        "filename": "meta_catalogue.csv",
        "media_type": "text/csv",
        "label": "Meta — Facebook & Instagram Shopping (CSV)",
    },
    "ebay": {
        "builder": to_ebay_csv,
        "filename": "ebay_produits.csv",
        "media_type": "text/csv",
        "label": "eBay (CSV File Exchange)",
    },
    "cdiscount": {
        "builder": to_cdiscount_csv,
        "filename": "cdiscount_produits.csv",
        "media_type": "text/csv",
        "label": "Cdiscount (CSV)",
    },
    "tiktok": {
        "builder": to_tiktok_csv,
        "filename": "tiktok_shop_produits.csv",
        "media_type": "text/csv",
        "label": "TikTok Shop (CSV)",
    },
}
