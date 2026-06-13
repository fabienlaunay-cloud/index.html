"""
Fills an Amazon inventory template (.xlsm/.xlsx) with listing data.

Supports all Amazon template types:
  - Legacy category templates (fptcustom): attributeRow=3, dataRow=4
    e.g. old ANIMAL_COLLAR.xlsm — has item_sku, item_name, bullet_point1…
  - Modern "Listing" templates (feedType=256): labelRow=4, attributeRow=5, dataRow=7
    e.g. ANIMAL_COLLAR_2.xlsm — qualified names like
    item_name[marketplace_id=A13V1IB3VIYZZH][language_tag=fr_FR]#1.value
  - Listing Loader / Price & Quantity: attributeRow=5, dataRow=7

Row 1 metadata contains `settings=…attributeRow=N&dataRow=M&ptds=<b64>` which we
parse to locate the correct rows, worksheet and product type automatically.
"""

import base64
import io
import re
import urllib.parse
from typing import List, Optional

import openpyxl

from app.models import AmazonListing

_HTML_RE = re.compile(r'<[^>]+>')
_QUALIFIER_RE = re.compile(r'\[[^\]]*\]')

# Generated image slots, in Amazon column order (main + other1..6)
_IMAGE_IDS_ORDERED = [
    "hero", "lifestyle_1", "infographic", "detail",
    "dimensions", "packaging", "lifestyle_2",
]


def _strip_html(text: str) -> str:
    return _HTML_RE.sub('', text or '').strip()


def _normalize_attr(name: str) -> str:
    """Strip [marketplace_id=…]/[language_tag=…] qualifiers so modern template
    attribute names match stable keys:
    item_name[marketplace_id=X][language_tag=fr_FR]#1.value → item_name#1.value
    """
    return _QUALIFIER_RE.sub('', name or '').strip()


def _parse_template_settings(wb: openpyxl.Workbook) -> tuple[dict, Optional[openpyxl.worksheet.worksheet.Worksheet]]:
    """
    Read row 1 of the first sheet that contains 'settings=' metadata.
    Returns (dict with attributeRow/dataRow/productType, worksheet found).
    """
    defaults = {"attributeRow": 3, "dataRow": 4, "productType": ""}
    for ws in wb.worksheets:
        for col in range(1, min(10, ws.max_column + 1)):
            val = str(ws.cell(1, col).value or '')
            if 'settings=' in val or 'attributeRow=' in val:
                qs = val.split('settings=', 1)[-1] if 'settings=' in val else val
                params = dict(urllib.parse.parse_qsl(qs))
                result = {**defaults}
                if 'attributeRow' in params:
                    result['attributeRow'] = int(params['attributeRow'])
                if 'dataRow' in params:
                    result['dataRow'] = int(params['dataRow'])
                # Modern templates carry the product type base64-encoded in ptds
                if params.get('ptds'):
                    try:
                        decoded = base64.b64decode(params['ptds']).decode('utf-8').strip()
                        # may be a comma-separated list — take the first
                        result['productType'] = decoded.split(',')[0].strip()
                    except Exception:
                        pass
                return result, ws
    return defaults, None


def _find_template_worksheet(wb: openpyxl.Workbook, attr_row: int) -> Optional[openpyxl.worksheet.worksheet.Worksheet]:
    """Return the worksheet whose attributeRow contains recognisable column names."""
    markers = {'item_sku', 'contribution_sku', 'external_product_id', 'externally_assigned', 'record_action'}
    for ws in wb.worksheets:
        for col in range(1, min(ws.max_column + 1, 20)):
            val = str(ws.cell(attr_row, col).value or '').lower()
            if any(m in val for m in markers):
                return ws
    for ws in wb.worksheets:
        if ws.cell(attr_row, 1).value:
            return ws
    return wb.active


def _build_col_index(ws, attr_row: int) -> dict[str, int]:
    """Map attribute_name → column number. Both the raw name and its
    qualifier-stripped normalization are indexed (first occurrence wins).
    Uses ws[attr_row] to iterate ALL cells (avoids max_column undercount)."""
    index = {}
    for cell in ws[attr_row]:
        val = str(cell.value or '').strip()
        if not val:
            continue
        col = cell.column
        if val not in index:
            index[val] = col
        norm = _normalize_attr(val)
        if norm and norm not in index:
            index[norm] = col
    return index


# ── Mapping ───────────────────────────────────────────────────────────────────

def _bullet(l: AmazonListing, i: int) -> str:
    return l.bullet_points[i] if len(l.bullet_points) > i else ""


# Exact-match rules — keys are either legacy names or normalized modern names.
_EXACT_MAP = {
    # Legacy category template style (fptcustom)
    "item_sku":                               lambda l: l.sku,
    "brand_name":                             lambda l: l.brand,
    "manufacturer":                           lambda l: l.brand,
    "part_number":                            lambda l: l.sku,
    "item_name":                              lambda l: l.title,
    "item_highlights":                        lambda l: getattr(l, "item_highlights", "") or "",
    "product_description":                    lambda l: _strip_html(l.description),
    "bullet_point1":                          lambda l: _bullet(l, 0),
    "bullet_point2":                          lambda l: _bullet(l, 1),
    "bullet_point3":                          lambda l: _bullet(l, 2),
    "bullet_point4":                          lambda l: _bullet(l, 3),
    "bullet_point5":                          lambda l: _bullet(l, 4),
    "generic_keywords":                       lambda l: l.backend_keywords,
    "external_product_id":                    lambda l: l.ean or "",
    "external_product_id_type":               lambda l: "EAN" if l.ean else "",
    "update_delete":                          lambda l: "Update",
    "standard_price":                         lambda l: str(l.price) if l.price else "",
    "quantity":                               lambda l: "1",
    "color_name":                             lambda l: l.color or (l.variation_value or "") or "",
    "material_type":                          lambda l: l.material or "",
    "item_weight":                            lambda l: str(l.weight_kg) if l.weight_kg else "",
    "item_weight_unit_of_measure":            lambda l: "KG" if l.weight_kg else "",
    "parent_child":                           lambda l: "Parent" if l.is_parent else ("Child" if l.parent_sku else ""),
    "parent_sku":                             lambda l: l.parent_sku or "",
    "relationship_type":                      lambda l: "Variation" if l.parent_sku else "",
    "variation_theme":                        lambda l: l.variation_theme or "",

    # Modern "Listing" template style (normalized names, feedType=256)
    "contribution_sku#1.value":               lambda l: l.sku,
    "::record_action":                        lambda l: "full_update",
    "item_name#1.value":                      lambda l: l.title,
    # "Point fort de l'article" — the new Item Highlights field
    "title_differentiation#1.value":          lambda l: getattr(l, "item_highlights", "") or "",
    "brand#1.value":                          lambda l: l.brand,
    "manufacturer#1.value":                   lambda l: l.brand,
    "model_name#1.value":                     lambda l: l.title[:50] if l.title else "",
    "model_number#1.value":                   lambda l: l.sku,
    "part_number#1.value":                    lambda l: l.sku,
    "product_description#1.value":            lambda l: _strip_html(l.description),
    "bullet_point#1.value":                   lambda l: _bullet(l, 0),
    "bullet_point#2.value":                   lambda l: _bullet(l, 1),
    "bullet_point#3.value":                   lambda l: _bullet(l, 2),
    "bullet_point#4.value":                   lambda l: _bullet(l, 3),
    "bullet_point#5.value":                   lambda l: _bullet(l, 4),
    "generic_keyword#1.value":                lambda l: l.backend_keywords,
    "color#1.value":                          lambda l: l.color or (l.variation_value or "") or "",
    "dominant_color#1.value":                 lambda l: l.color or (l.variation_value or "") or "",
    "size#1.value":                           lambda l: getattr(l, "size", None) or "",
    "size_name":                              lambda l: getattr(l, "size", None) or "",
    "material#1.value":                       lambda l: l.material or "",
    "amzn1.volt.ca.product_id_type":          lambda l: "EAN" if l.ean else "",
    "amzn1.volt.ca.product_id_value":         lambda l: l.ean or "",
    "externally_assigned_product_identifier#1.type":  lambda l: "EAN" if l.ean else "",
    "externally_assigned_product_identifier#1.value": lambda l: l.ean or "",
    # condition_type intentionally omitted — let the template's example row provide
    # the locale-correct value (e.g. "Neuf" for FR) instead of hardcoding "new"
    "fulfillment_availability#1.quantity":    lambda l: "1",
    "item_package_quantity#1.value":          lambda l: "1",
    "number_of_items#1.value":                lambda l: "1",
    # Variations (modern)
    "parentage_level#1.value":                lambda l: "Parent" if l.is_parent else ("Enfant" if l.parent_sku else ""),
    "child_parent_sku_relationship#1.parent_sku": lambda l: l.parent_sku or "",
    "child_parent_sku_relationship#1.child_relationship_type": lambda l: "variation" if l.parent_sku else "",
    "variation_theme#1.name":                 lambda l: _resolve_variation_theme(l.variation_theme or "") if l.is_parent else "",
    # Logistics / compliance smart defaults (overridable via compliance_data)
    "unit_count#1.value":                     lambda l: "" if l.is_parent else "1",
    "unit_count#1.type.value":               lambda l: "" if l.is_parent else "unité",
    "merchant_shipping_group#1.value":        lambda l: "Modèle par défaut Amazon",
    "supplier_declared_dg_hz_regulation#1.value": lambda l: "Non applicable",
    # "Piles nécessaires ?" / "Piles fournies ?" are mandatory in most categories.
    # Default to "Non" so the export never trips the required-attribute check; the
    # user can override per-export via the compliance form (applied in step 7f).
    "batteries_required#1.value":             lambda l: "Non",
    "batteries_included#1.value":             lambda l: "Non",
    # "Description de la garantie" is conditionally required. Default to the EU
    # statutory minimum (2-year legal guarantee of conformity) so the export
    # passes; the compliance form overrides it per-export (step 7f).
    "warranty_description#1.value":           lambda l: "Garantie légale de conformité de 2 ans",
}

# Image slot → normalized modern column / legacy column
_IMAGE_COLUMN_KEYS = [
    ("main_product_image_locator#1.media_location", "main_image_url"),
    ("other_product_image_locator_1#1.media_location", "other_image_url1"),
    ("other_product_image_locator_2#1.media_location", "other_image_url2"),
    ("other_product_image_locator_3#1.media_location", "other_image_url3"),
    ("other_product_image_locator_4#1.media_location", "other_image_url4"),
    ("other_product_image_locator_5#1.media_location", "other_image_url5"),
    ("other_product_image_locator_6#1.media_location", "other_image_url6"),
]

# Pattern-based rules for columns whose names vary by marketplace ID
_PATTERN_RULES = [
    # Price: any column containing 'our_price' and 'value_with_tax' (modern)
    (lambda attr: 'our_price' in attr and 'value_with_tax' in attr and 'b2b' not in attr.lower(),
     lambda l: str(l.price) if l.price else ""),
]

# Map SynqIO internal theme names → valid Amazon FR variation theme values
_THEME_TO_AMAZON = {
    "ColorName":         "COULEUR",
    "SizeClass":         "TAILLE",
    "ColorName-SizeClass": "COULEUR/TAILLE",
    "SizeClass-ColorName": "COULEUR/TAILLE",
    "Color":             "COULEUR",
    "Size":              "TAILLE",
    "Color-Size":        "COULEUR/TAILLE",
}


def _resolve_variation_theme(theme: str) -> str:
    """Convert a SynqIO internal theme name ('SizeClass', 'ColorName'…)
    to the valid Amazon FR value ('TAILLE', 'COULEUR'…).
    If the value is already in Amazon format (all-caps / contains slash)
    it is returned as-is."""
    if not theme:
        return ""
    if theme in _THEME_TO_AMAZON:
        return _THEME_TO_AMAZON[theme]
    # Already an Amazon-format value (COULEUR, COULEUR/TAILLE, ANIMAL_COLLAR_TYPE/…)
    return theme


def _derive_parent_sku(skus: list) -> str:
    """Derive a parent SKU from variation child SKUs by extracting the
    common prefix+suffix wrapper (e.g. ['C-COL-ROSE-TU','C-COL-BLEU-TU']
    → 'C-COL-TU'). Falls back to first SKU + '-PARENT'."""
    if len(skus) < 2:
        return f"{skus[0]}-PARENT" if skus else "PARENT"
    # Common prefix
    prefix = skus[0]
    for s in skus[1:]:
        while prefix and not s.startswith(prefix):
            prefix = prefix[:-1]
    prefix = prefix.rstrip("-_.")
    # Common suffix
    suffix = skus[0]
    for s in skus[1:]:
        while suffix and not s.endswith(suffix):
            suffix = suffix[1:]
    suffix = suffix.lstrip("-_.")
    if prefix and suffix and prefix != suffix:
        candidate = f"{prefix}-{suffix}"
    elif prefix:
        candidate = prefix
    else:
        candidate = f"{skus[0]}-PARENT"
    # Ensure it differs from all children
    return candidate if candidate not in skus else f"{candidate}-P"


def _make_parent_listing(children: list, parent_sku: str, variation_theme: str = "") -> AmazonListing:
    """Build a minimal parent listing from variation children."""
    ref = children[0]
    return AmazonListing(
        sku=parent_sku,
        title=ref.title,
        item_highlights=ref.item_highlights,
        bullet_points=ref.bullet_points,
        description=ref.description,
        backend_keywords=ref.backend_keywords,
        brand=ref.brand,
        category=ref.category,
        price=None,
        ean=None,
        color=None,
        material=ref.material,
        variation_theme=variation_theme,
        variation_value=None,
        is_parent=True,
        parent_sku=None,
        children=[],
    )


def get_product_type_from_bytes(template_bytes: bytes) -> str:
    """Extract the Amazon product type code from template bytes (e.g. 'ANIMAL_COLLAR')."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(template_bytes), keep_vba=True, data_only=True)
        settings, _ = _parse_template_settings(wb)
        return (settings.get('productType') or '').upper()
    except Exception:
        return ''


def fill_amazon_template(template_bytes: bytes, listings: List[AmazonListing],
                         image_urls: Optional[dict] = None,
                         compliance_data: Optional[dict] = None) -> bytes:
    """
    Read an Amazon category template (.xlsm/.xlsx), detect its structure,
    fill it with listing data, and return the result as bytes.
    image_urls: optional {sku: {slot_id: public_url}} to fill image columns.

    Compliance/logistics columns (packaging dimensions, country of origin,
    warranty, battery info, dangerous-goods regulation, safety instructions…)
    are inherited from the example row that Amazon ships inside every template,
    so clients never have to re-enter them for each variation.
    """
    wb = openpyxl.load_workbook(io.BytesIO(template_bytes), keep_vba=True, data_only=True)

    # 1. Parse settings — also returns the exact worksheet that holds them
    settings, settings_ws = _parse_template_settings(wb)
    attr_row = settings['attributeRow']
    data_row = settings['dataRow']
    product_type = settings.get('productType', '')

    # 2. Use the settings worksheet if found; fall back to heuristic search
    ws = settings_ws or _find_template_worksheet(wb, attr_row)
    if ws is None:
        raise ValueError("Impossible de trouver la feuille template dans ce fichier Amazon.")

    # 3. Build column index (raw + normalized names)
    col_index = _build_col_index(ws, attr_row)

    # 4. Product type detection
    fpt_col = col_index.get("feed_product_type")
    if fpt_col and not product_type:
        for r in range(attr_row + 1, attr_row + 5):
            v = ws.cell(r, fpt_col).value
            if v:
                product_type = str(v)
                break
        if not product_type:
            for col in range(1, min(6, ws.max_column + 1)):
                val = str(ws.cell(1, col).value or '')
                if 'TemplateSignature=' in val:
                    sig = val.split('TemplateSignature=')[1].split('&')[0]
                    try:
                        product_type = base64.b64decode(sig).decode('utf-8').strip()
                    except Exception:
                        pass
                    break
    pt_col = col_index.get("product_type#1.value") or fpt_col

    # 5. (No example-row inheritance: the ABC123 row in Amazon templates contains
    #    demo/fictitious values that fail validation for real products.  Compliance
    #    and logistics fields — packaging dimensions, dangerous-goods classification,
    #    battery info, REP material, B2B pricing tiers, etc. — must be entered by
    #    the user in Excel after export, or provided via the extra{} dict on each
    #    listing when the caller supplies real values.)
    example_defaults: dict[int, object] = {}

    # 6. Identify "managed" columns — content fields our mapping knows about.
    #    Compliance columns (NOT in this set) inherit from the example row.
    managed_cols: set[int] = set()
    for attr in _EXACT_MAP:
        c = col_index.get(attr)
        if c:
            managed_cols.add(c)
    for modern_key, legacy_key in _IMAGE_COLUMN_KEYS:
        c = col_index.get(modern_key) or col_index.get(legacy_key)
        if c:
            managed_cols.add(c)
    for pattern_fn, _ in _PATTERN_RULES:
        for attr, c in col_index.items():
            if pattern_fn(attr):
                managed_cols.add(c)
    if pt_col:
        managed_cols.add(pt_col)

    # 6b. Auto-generate the variation family when children are exported without
    #     an explicit parent. Triggers when (a) listings carry nested
    #     declination children (sizes per color → expanded to one row each),
    #     (b) all listings declare a variation_theme, OR (c) their SKUs share a
    #     non-trivial common prefix AND suffix (≥2 chars each), a reliable
    #     signal for variation siblings.
    has_parent = any(l.is_parent for l in listings)
    has_parent_sku = any(l.parent_sku for l in listings)
    has_nested_children = any(getattr(l, "children", None) for l in listings)
    # Maps expanded declination SKU → base listing SKU (for image lookup)
    base_sku_of: dict[str, str] = {}
    if not has_parent and not has_parent_sku and (len(listings) > 1 or has_nested_children):
        skus = [l.sku for l in listings]
        themes = [l.variation_theme for l in listings if l.variation_theme]
        all_have_theme = len(themes) == len(listings)
        # SKU pattern on the BASE listings (before declination expansion):
        # compute common prefix and suffix
        _pfx = skus[0]
        for s in skus[1:]:
            while _pfx and not s.startswith(_pfx):
                _pfx = _pfx[:-1]
        _pfx = _pfx.rstrip("-_.")
        _sfx = skus[0]
        for s in skus[1:]:
            while _sfx and not s.endswith(_sfx):
                _sfx = _sfx[1:]
        _sfx = _sfx.lstrip("-_.")
        sku_pattern_clear = len(skus) > 1 and len(_pfx) >= 2 and len(_sfx) >= 2 and _pfx != _sfx
        if has_nested_children or all_have_theme or sku_pattern_clear:
            parent_sku = _derive_parent_sku(skus)
            # Per-base-listing color, extracted from the SKU middle segment
            # (the part that differs between color variants) when not set.
            color_by_sku: dict[str, str] = {}
            if sku_pattern_clear:
                for l in listings:
                    mid_start = len(_pfx) + 1
                    mid_end = len(l.sku) - len(_sfx) - 1
                    if mid_end > mid_start:
                        color_by_sku[l.sku] = l.sku[mid_start:mid_end]
            # Build the child rows: one per nested declination (size), or the
            # base listing itself when it has no declinations.
            child_rows = []
            for l in listings:
                base_color = l.color or color_by_sku.get(l.sku) or None
                kids = getattr(l, "children", None) or []
                if kids:
                    for c in kids:
                        base_sku_of[c.sku] = l.sku
                        child_rows.append(l.model_copy(update={
                            "sku": c.sku,
                            "price": c.price if c.price is not None else l.price,
                            "ean": c.ean or None,
                            "color": c.color or base_color,
                            "size": c.size or None,
                            "variation_value": c.variation_value or None,
                            "parent_sku": parent_sku,
                            "children": [],
                        }))
                else:
                    updates: dict = {"parent_sku": parent_sku}
                    if base_color and not l.color:
                        updates["color"] = base_color
                        if not l.variation_value:
                            updates["variation_value"] = base_color
                    child_rows.append(l.model_copy(update=updates))
            # Variation theme: explicit (resolved to Amazon FR vocab) wins;
            # otherwise inferred from what actually varies between children.
            explicit_theme = _resolve_variation_theme(themes[0]) if all_have_theme and themes else ""
            if explicit_theme:
                parent_theme = explicit_theme
            else:
                n_colors = len({r.color for r in child_rows if r.color})
                n_sizes = len({getattr(r, "size", None) for r in child_rows if getattr(r, "size", None)})
                if n_sizes > 1 and n_colors > 1:
                    parent_theme = "COULEUR/TAILLE"
                elif n_sizes > 1:
                    parent_theme = "TAILLE"
                else:
                    parent_theme = "COULEUR"
            parent_listing = _make_parent_listing(list(listings), parent_sku, variation_theme=parent_theme)
            listings = [parent_listing] + child_rows

    # 7. Write listing data starting at data_row
    for i, listing in enumerate(listings):
        row_num = data_row + i

        # 7a. Inherit compliance/logistics from example row (unmanaged cols only)
        for col, val in example_defaults.items():
            if col not in managed_cols:
                ws.cell(row_num, col).value = val

        # 7b. product_type
        if product_type and pt_col:
            ws.cell(row_num, pt_col).value = product_type

        # 7c. Exact-match attributes (legacy + normalized modern)
        #     Write even empty string to clear any inherited example value.
        for attr, getter in _EXACT_MAP.items():
            col = col_index.get(attr)
            if col:
                value = getter(listing)
                ws.cell(row_num, col).value = value if value else None

        # 7d. Images — expanded declination rows fall back to their base
        #     listing's images (same color photos shared across sizes)
        sku_imgs = (image_urls or {}).get(listing.sku) \
            or (image_urls or {}).get(base_sku_of.get(listing.sku, ""), {})
        if sku_imgs:
            for slot_id, (modern_key, legacy_key) in zip(_IMAGE_IDS_ORDERED, _IMAGE_COLUMN_KEYS):
                url = sku_imgs.get(slot_id)
                if not url:
                    continue
                col = col_index.get(modern_key) or col_index.get(legacy_key)
                if col:
                    ws.cell(row_num, col).value = url

        # 7e. Pattern-match attributes (e.g. price column with marketplace ID)
        for (pattern_fn, value_fn) in _PATTERN_RULES:
            for attr, col in col_index.items():
                if pattern_fn(attr):
                    value = value_fn(listing)
                    if value and not ws.cell(row_num, col).value:
                        ws.cell(row_num, col).value = value
                    break  # fill only first matching price column

        # 7f. Per-export compliance values (country_of_origin, packaging dims, batteries, etc.)
        #     Applied after _EXACT_MAP so user-provided values take precedence.
        #     variation_theme is parent-only — never write it to child rows from compliance_data.
        if compliance_data:
            _parent_only_attrs = {"variation_theme#1.name"}
            _gtin_exempt = compliance_data.get("_gtin_exempt")
            for attr, value in compliance_data.items():
                if attr.startswith("_"):
                    continue  # internal flags, not column keys
                if attr in _parent_only_attrs and listing.parent_sku:
                    continue
                col = col_index.get(attr)
                if col and value not in (None, ""):
                    ws.cell(row_num, col).value = value
            # GTIN exemption: clear all EAN/identifier columns
            if _gtin_exempt:
                _ean_attrs = [
                    "externally_assigned_product_identifier#1.type",
                    "externally_assigned_product_identifier#1.value",
                    "external_product_id", "external_product_id_type",
                    "amzn1.volt.ca.product_id_type", "amzn1.volt.ca.product_id_value",
                ]
                for attr in _ean_attrs:
                    col = col_index.get(attr)
                    if col:
                        ws.cell(row_num, col).value = None

    # 8. Save preserving macros if the source was .xlsm
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
