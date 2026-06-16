import csv
import json
import io
from typing import List, Optional
from app.models import AmazonListing
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Ordered image IDs matching Amazon flat file slot columns
_IMAGE_IDS_ORDERED = [
    "hero",        # main-image-url
    "lifestyle_1", # other-image-url1
    "infographic", # other-image-url2
    "detail",      # other-image-url3
    "dimensions",  # other-image-url4
    "packaging",   # other-image-url5
    "lifestyle_2", # other-image-url6
]


def to_csv_bytes(listings: List[AmazonListing]) -> bytes:
    output = io.StringIO()
    if not listings:
        return b""
    fieldnames = [
        "sku", "marketplace", "title", "item_highlights",
        "bullet_point_1", "bullet_point_2",
        "bullet_point_3", "bullet_point_4", "bullet_point_5",
        "description", "backend_keywords", "brand", "category",
        "price", "ean", "seo_score",
    ]
    writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()
    for listing in listings:
        bullets = listing.bullet_points + [""] * 5
        writer.writerow({
            "sku": listing.sku,
            "marketplace": listing.marketplace.value,
            "title": listing.title,
            "item_highlights": getattr(listing, "item_highlights", "") or "",
            "bullet_point_1": bullets[0],
            "bullet_point_2": bullets[1],
            "bullet_point_3": bullets[2],
            "bullet_point_4": bullets[3],
            "bullet_point_5": bullets[4],
            "description": listing.description,
            "backend_keywords": listing.backend_keywords,
            "brand": listing.brand,
            "category": listing.category,
            "price": listing.price or "",
            "ean": listing.ean or "",
            "seo_score": listing.seo_score or "",
        })
    return output.getvalue().encode("utf-8-sig")


def to_json_bytes(listings: List[AmazonListing]) -> bytes:
    return json.dumps(
        [l.model_dump() for l in listings],
        ensure_ascii=False,
        indent=2,
        default=str,
    ).encode("utf-8")


import re as _re

def _strip_html(text: str) -> str:
    return _re.sub(r'<[^>]+>', '', text or '').strip()


def to_amazon_flat_file_bytes(listings: List[AmazonListing]) -> bytes:
    """Amazon Listing Loader flat file (tab-separated, UTF-8-BOM).

    Compatible with Seller Central > Inventory > Add Products via Spreadsheet.
    Uses the standard Listing Loader template columns accepted across categories.
    """
    output = io.StringIO()
    # fptcustom UPLOAD field IDs use underscores; hyphenated forms are rejected
    # on upload (90012 'sku absent' / 90061 'invalid header').
    headers = [
        "item_sku",
        "update_delete",
        "item_name",
        "item_highlights",
        "brand_name",
        "manufacturer",
        "item_type",
        "product_description",
        "bullet_point1",
        "bullet_point2",
        "bullet_point3",
        "bullet_point4",
        "bullet_point5",
        "generic_keywords",
        "standard_price",
        "quantity",
        "external_product_id",
        "external_product_id_type",
        "condition_type",
    ]
    # Row 1 : TemplateType MUST come before column names (Amazon error 90009 otherwise)
    output.write("\t".join(["TemplateType=fptcustom", "Version=2021.1201"] + [""] * (len(headers) - 2)) + "\n")
    # Row 2 : column names
    output.write("\t".join(headers) + "\n")

    for listing in listings:
        bullets = listing.bullet_points + [""] * 5
        row = [
            listing.sku,
            "a",                                        # add / partial-update
            listing.title,
            getattr(listing, "item_highlights", "") or "",
            listing.brand,
            listing.brand,
            listing.category.lower().replace(" ", "_") or "home",
            _strip_html(listing.description),           # no HTML in flat files
            bullets[0], bullets[1], bullets[2], bullets[3], bullets[4],
            listing.backend_keywords,
            str(listing.price) if listing.price else "",
            "1",                                        # quantity placeholder
            listing.ean or "",
            "EAN" if listing.ean else "",
            "New",
        ]
        output.write("\t".join(row) + "\n")

    return output.getvalue().encode("utf-8-sig")


def _style_header_row(ws, n_cols: int):
    """Bold + purple header, gray metadata row."""
    purple = PatternFill("solid", fgColor="764BA2")
    gray   = PatternFill("solid", fgColor="E5E7EB")
    white  = Font(color="FFFFFF", bold=True, size=10)
    dark   = Font(color="374151", size=9)
    for col in range(1, n_cols + 1):
        h = ws.cell(1, col)
        h.font = white;  h.fill = purple
        h.alignment = Alignment(horizontal="center", vertical="center")
        m = ws.cell(2, col)
        m.font = dark;   m.fill = gray
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 16


def to_amazon_flat_file_xlsx(listings: List[AmazonListing], image_urls: Optional[dict] = None) -> bytes:
    """Amazon flat file for NEW products — proper .xlsx, uploadable to Seller Central."""
    # fptcustom UPLOAD field IDs use underscores (item_sku/item_name/…). The
    # hyphenated forms are the Category-Listings-Report download style and are
    # rejected on upload (90012 'sku absent' / 90061 'invalid header').
    headers = [
        "item_sku", "update_delete", "item_name", "brand_name", "manufacturer",
        "item_type", "product_description",
        "bullet_point1", "bullet_point2", "bullet_point3", "bullet_point4", "bullet_point5",
        "generic_keywords", "standard_price", "quantity",
        "external_product_id", "external_product_id_type", "condition_type",
        "main_image_url",
        "other_image_url1", "other_image_url2", "other_image_url3",
        "other_image_url4", "other_image_url5", "other_image_url6",
    ]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Nouveaux Produits"
    ws.append(["TemplateType=fptcustom", "Version=2021.1201"] + [""] * (len(headers) - 2))
    ws.append(headers)
    for listing in listings:
        bullets = listing.bullet_points + [""] * 5
        sku_imgs = (image_urls or {}).get(listing.sku, {})
        img_row = [sku_imgs.get(_IMAGE_IDS_ORDERED[0], "")]
        img_row += [sku_imgs.get(iid, "") for iid in _IMAGE_IDS_ORDERED[1:]]
        ws.append([
            listing.sku, "a", listing.title, listing.brand, listing.brand,
            listing.category.lower().replace(" ", "_") or "home",
            _strip_html(listing.description),
            bullets[0], bullets[1], bullets[2], bullets[3], bullets[4],
            listing.backend_keywords,
            str(listing.price) if listing.price else "", "1",
            listing.ean or "", "EAN" if listing.ean else "", "New",
        ] + img_row)
    _style_header_row(ws, len(headers))
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = max(12, len(h) + 2)
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["G"].width = 50
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def to_listing_loader_xlsx(listings: List[AmazonListing]) -> bytes:
    """Amazon Listing Loader for EXISTING products — proper .xlsx, uploadable to Seller Central."""
    headers = [
        "sku", "product-id", "product-id-type", "add-delete", "price", "quantity",
        "condition-type", "item-name", "brand-name", "item-description",
        "bullet-point1", "bullet-point2", "bullet-point3", "bullet-point4", "bullet-point5",
        "generic-keywords",
    ]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Produits Existants"
    ws.append(["TemplateType=Offer", "Version=2021.1201"] + [""] * (len(headers) - 2))
    ws.append(headers)
    for listing in listings:
        bullets = listing.bullet_points + [""] * 5
        ws.append([
            listing.sku, listing.ean or "", "4" if listing.ean else "", "a",
            str(listing.price) if listing.price else "", "1", "New",
            listing.title, listing.brand, _strip_html(listing.description),
            bullets[0], bullets[1], bullets[2], bullets[3], bullets[4],
            listing.backend_keywords,
        ])
    _style_header_row(ws, len(headers))
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = max(12, len(h) + 2)
    ws.column_dimensions["H"].width = 40
    ws.column_dimensions["J"].width = 50
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def to_variation_flat_file_xlsx(listings: List[AmazonListing], image_urls: Optional[dict] = None) -> bytes:
    """
    Amazon Variation flat file for NEW products with parent-child relationships.
    Handles both regular listings and variation groups in a single sheet.

    Parent rows: full content, no price/EAN, parent-child="parent"
    Child rows: minimal content (price, EAN, color, size), parent-child="child"
    Standalone rows: regular flat file format
    """
    headers = [
        "sku", "update-delete", "parent-child", "parent-sku",
        "relationship-type", "variation-theme",
        "item-name", "brand_name", "manufacturer", "item_type",
        "product-description",
        "bullet-point1", "bullet-point2", "bullet-point3", "bullet-point4", "bullet-point5",
        "generic-keywords", "standard-price", "quantity",
        "external-product-id", "external-product-id-type", "condition-type",
        "color-name", "size-name",
        "main-image-url",
        "other-image-url1", "other-image-url2", "other-image-url3",
        "other-image-url4", "other-image-url5", "other-image-url6",
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Variations"
    ws.append(["TemplateType=fptcustom", "Version=2021.1201"] + [""] * (len(headers) - 2))
    ws.append(headers)

    parents = {l.sku: l for l in listings if l.is_parent}

    def _img_row(sku: str) -> list:
        sku_imgs = (image_urls or {}).get(sku, {})
        row = [sku_imgs.get(_IMAGE_IDS_ORDERED[0], "")]
        row += [sku_imgs.get(iid, "") for iid in _IMAGE_IDS_ORDERED[1:]]
        return row

    for listing in listings:
        bullets = listing.bullet_points + [""] * 5
        item_type = listing.category.lower().replace(" ", "_") or "home"

        if listing.is_parent:
            # Parent row — includes images
            ws.append([
                listing.sku, "a", "parent", "", "", listing.variation_theme or "",
                listing.title, listing.brand, listing.brand, item_type,
                _strip_html(listing.description),
                bullets[0], bullets[1], bullets[2], bullets[3], bullets[4],
                listing.backend_keywords, "", "", "", "", "",
                "", "",
            ] + _img_row(listing.sku))
            # Child rows from children list (no images on child rows)
            for child in listing.children:
                ws.append([
                    child.sku, "a", "child", listing.sku,
                    "Variation", listing.variation_theme or "",
                    listing.title, listing.brand, listing.brand, item_type,
                    "", "", "", "", "", "", "",
                    str(child.price) if child.price else "", "1",
                    child.ean or "", "EAN" if child.ean else "", "New",
                    child.color or "", child.size or "",
                ] + [""] * 7)
        elif listing.parent_sku:
            # Orphan child — write as standalone
            if listing.parent_sku not in parents:
                ws.append([
                    listing.sku, "a", "child", listing.parent_sku,
                    "Variation", listing.variation_theme or "",
                    listing.title, listing.brand, listing.brand, item_type,
                    _strip_html(listing.description),
                    bullets[0], bullets[1], bullets[2], bullets[3], bullets[4],
                    listing.backend_keywords,
                    str(listing.price) if listing.price else "", "1",
                    listing.ean or "", "EAN" if listing.ean else "", "New",
                    listing.color or "", "",
                ] + _img_row(listing.sku))
        else:
            # Standalone (non-variation) listing
            ws.append([
                listing.sku, "a", "", "", "", "",
                listing.title, listing.brand, listing.brand, item_type,
                _strip_html(listing.description),
                bullets[0], bullets[1], bullets[2], bullets[3], bullets[4],
                listing.backend_keywords,
                str(listing.price) if listing.price else "", "1",
                listing.ean or "", "EAN" if listing.ean else "", "New",
                listing.color or "", "",
            ] + _img_row(listing.sku))

    _style_header_row(ws, len(headers))
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = max(10, len(h) + 2)
    ws.column_dimensions["G"].width = 40
    ws.column_dimensions["K"].width = 50

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def to_listing_loader_bytes(listings: List[AmazonListing]) -> bytes:
    """Amazon Listing Loader — for products already in Amazon's catalog.

    Use when the product has an existing ASIN or EAN already indexed on Amazon.
    Updates offer + content (title, bullets, description, keywords) on the existing page.
    Compatible with: Seller Central > Inventory > Add Products via Spreadsheet
                     > 'Mettre en vente des produits figurant dans le catalogue'.
    """
    output = io.StringIO()
    headers = [
        "sku",
        "product-id",
        "product-id-type",
        "add-delete",
        "price",
        "quantity",
        "condition-type",
        "item-name",
        "brand-name",
        "item-description",
        "bullet-point1",
        "bullet-point2",
        "bullet-point3",
        "bullet-point4",
        "bullet-point5",
        "generic-keywords",
    ]
    output.write("\t".join(["TemplateType=Offer", "Version=2021.1201"] + [""] * (len(headers) - 2)) + "\n")
    output.write("\t".join(headers) + "\n")

    for listing in listings:
        bullets = listing.bullet_points + [""] * 5
        product_id = listing.ean or ""
        product_id_type = "4" if listing.ean else ""   # 4=EAN, 1=ASIN, 3=UPC
        row = [
            listing.sku,
            product_id,
            product_id_type,
            "a",
            str(listing.price) if listing.price else "",
            "1",
            "New",
            listing.title,
            listing.brand,
            _strip_html(listing.description),
            bullets[0], bullets[1], bullets[2], bullets[3], bullets[4],
            listing.backend_keywords,
        ]
        output.write("\t".join(row) + "\n")

    return output.getvalue().encode("utf-8-sig")
