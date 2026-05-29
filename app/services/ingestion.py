import io
import csv
import json
from typing import List, BinaryIO, Optional
from app.models import RawProduct


COLUMN_MAP = {
    # SKU / référence
    "sku": "sku", "référence": "sku", "ref": "sku", "reference": "sku", "id": "sku",
    # Nom
    "nom": "name", "name": "name", "titre": "name", "title": "name", "libellé": "name",
    # Marque
    "marque": "brand", "brand": "brand",
    # Catégorie
    "catégorie": "category", "category": "category", "cat": "category",
    # Description
    "description": "description", "desc": "description",
    # Prix
    "prix": "price", "price": "price", "tarif": "price",
    # EAN
    "ean": "ean", "gtin": "ean", "barcode": "ean", "code barre": "ean",
    # Poids
    "poids": "weight_kg", "weight": "weight_kg", "weight_kg": "weight_kg",
    "poids kg": "weight_kg",
    # Dimensions
    "dimensions": "dimensions_cm", "dim": "dimensions_cm",
    "dimensions cm": "dimensions_cm",
    # Couleur / Matière
    "couleur": "color", "color": "color", "colour": "color",
    "matière": "material", "material": "material", "matériau": "material",
    # Images (URL externe)
    "images": "images", "image": "images", "photo": "images", "photos": "images",
    "image url": "images", "image_url": "images", "photo url": "images", "photo_url": "images",
    "url image": "images", "url photo": "images",
    # Image_Fichier — nom de fichier uploadé via ZIP (converti en /api/photos/{nom})
    "image fichier": "image_file", "image_fichier": "image_file",
    "fichier image": "image_file", "photo fichier": "image_file",
    "fichier photo": "image_file", "nom fichier": "image_file",
    # Caractéristiques
    "caractéristiques": "features", "features": "features", "points clés": "features",
    # Mots-clés produit
    "mots clés": "focus_keywords", "mots_clés": "focus_keywords", "mots-clés": "focus_keywords",
    "mots cles": "focus_keywords", "mots_cles": "focus_keywords",
    "keywords": "focus_keywords", "search terms": "focus_keywords",
    # Segment (alias catégorie)
    "segment": "category",
}


def _normalise_header(h: str) -> str:
    h = h.strip().lower().replace("_", " ")
    if h.endswith(" *"):
        h = h[:-2].rstrip()
    return h


def _map_row(row: dict, custom_mapping: dict = None) -> dict:
    mapped: dict = {}
    extra: dict = {}
    for key, value in row.items():
        norm = _normalise_header(key)
        # custom_mapping takes priority: try exact key first, then normalised key
        custom_target = None
        if custom_mapping:
            if key in custom_mapping and custom_mapping[key]:
                custom_target = custom_mapping[key]
            elif norm in custom_mapping and custom_mapping[norm]:
                custom_target = custom_mapping[norm]
        if custom_target:
            mapped[custom_target] = value
        else:
            target = COLUMN_MAP.get(norm)
            if target:
                mapped[target] = value
            else:
                extra[key] = value
    mapped.setdefault("extra", extra)
    return mapped


def _coerce(mapped: dict) -> RawProduct:
    if "price" in mapped and mapped["price"] not in (None, ""):
        try:
            mapped["price"] = float(str(mapped["price"]).replace(",", "."))
        except ValueError:
            mapped.pop("price", None)

    if "weight_kg" in mapped and mapped["weight_kg"] not in (None, ""):
        try:
            mapped["weight_kg"] = float(str(mapped["weight_kg"]).replace(",", "."))
        except ValueError:
            mapped.pop("weight_kg", None)

    for list_field in ("images", "features"):
        if list_field in mapped and isinstance(mapped[list_field], str):
            mapped[list_field] = [
                v.strip() for v in mapped[list_field].split("|") if v.strip()
            ]

    if "focus_keywords" in mapped and isinstance(mapped["focus_keywords"], str):
        mapped["focus_keywords"] = [
            k.strip() for k in mapped["focus_keywords"].replace(";", ",").split(",") if k.strip()
        ]

    # Image_Fichier → convertir en URL interne /api/photos/{nom} si pas d'Image_URL
    if mapped.get("image_file"):
        fname = str(mapped.pop("image_file")).strip()
        if fname and not mapped.get("images"):
            mapped["images"] = [f"/api/photos/{fname}"]
    else:
        mapped.pop("image_file", None)

    if not mapped.get("sku"):
        mapped["sku"] = mapped.get("name", "UNKNOWN")[:20]

    return RawProduct(**{k: v for k, v in mapped.items() if v not in (None, "", [])})


def get_headers_and_sample(filename: str, content: bytes) -> dict:
    """Parse just the headers and first 3 data rows without full RawProduct coercion.

    Returns:
        {
            "headers": [...],
            "sample_rows": [{...}, ...],
            "auto_mapped": {header: field_or_None}
        }
    """
    ext = filename.rsplit(".", 1)[-1].lower()
    headers = []
    sample_rows = []

    if ext == "csv":
        text = content.decode("utf-8-sig", errors="replace")
        sample_text = text[:2000]
        delimiter = ";" if sample_text.count(";") > sample_text.count(",") else ","
        reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
        headers = list(reader.fieldnames or [])
        for i, row in enumerate(reader):
            if i >= 3:
                break
            sample_rows.append(dict(row))

    elif ext in ("xlsx", "xls"):
        try:
            import openpyxl
        except ImportError:
            raise RuntimeError("openpyxl manquant — pip install openpyxl")
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        headers = [str(cell.value or "").strip() for cell in next(ws.iter_rows(max_row=1))]
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            if i >= 3:
                break
            row_dict = {headers[j]: (row[j] if j < len(row) else None) for j in range(len(headers))}
            if all(v is None for v in row_dict.values()):
                continue
            sample_rows.append(row_dict)

    elif ext == "json":
        data = json.loads(content)
        if isinstance(data, dict):
            data = data.get("products", data.get("items", [data]))
        if data:
            # Collect all keys across first 3 items
            seen = {}
            for item in data[:3]:
                for k in item.keys():
                    seen[k] = seen.get(k, None)
            headers = list(seen.keys())
            for item in data[:3]:
                sample_rows.append({h: item.get(h) for h in headers})
    else:
        raise ValueError(f"Format non supporté: .{ext}  (csv, xlsx, json acceptés)")

    # Determine auto_mapped: header → field or None
    auto_mapped = {}
    for h in headers:
        norm = _normalise_header(h)
        auto_mapped[h] = COLUMN_MAP.get(norm)  # None if unrecognised

    return {
        "headers": headers,
        "sample_rows": sample_rows,
        "auto_mapped": auto_mapped,
    }


def parse_csv(content: bytes, delimiter: str = None, custom_mapping: dict = None) -> List[RawProduct]:
    text = content.decode("utf-8-sig", errors="replace")
    if delimiter is None:
        sample = text[:2000]
        delimiter = ";" if sample.count(";") > sample.count(",") else ","
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    products = []
    for row in reader:
        try:
            products.append(_coerce(_map_row(dict(row), custom_mapping)))
        except Exception:
            pass
    return products


def parse_excel(content: bytes, custom_mapping: dict = None) -> List[RawProduct]:
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl manquant — pip install openpyxl")

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    headers = [str(cell.value or "").strip() for cell in next(ws.iter_rows(max_row=1))]
    products = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
        if all(v is None for v in row_dict.values()):
            continue
        try:
            products.append(_coerce(_map_row(row_dict, custom_mapping)))
        except Exception:
            pass
    return products


def parse_json(content: bytes, custom_mapping: dict = None) -> List[RawProduct]:
    data = json.loads(content)
    if isinstance(data, dict):
        data = data.get("products", data.get("items", [data]))
    products = []
    for item in data:
        try:
            products.append(_coerce(_map_row(item, custom_mapping)))
        except Exception:
            pass
    return products


def parse_file(filename: str, content: bytes, custom_mapping: dict = None) -> List[RawProduct]:
    ext = filename.rsplit(".", 1)[-1].lower()
    if ext == "csv":
        return parse_csv(content, custom_mapping=custom_mapping)
    if ext in ("xlsx", "xls"):
        return parse_excel(content, custom_mapping=custom_mapping)
    if ext == "json":
        return parse_json(content, custom_mapping=custom_mapping)
    raise ValueError(f"Format non supporté: .{ext}  (csv, xlsx, json acceptés)")
