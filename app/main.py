import os
import io
import re
import json
import time
import asyncio
import base64
import uuid
import httpx
from uuid import uuid4
from dotenv import load_dotenv
from app.logger import log
load_dotenv(override=False)  # local dev only — Railway injects vars directly

if os.getenv("SENTRY_DSN"):
    import sentry_sdk
    from sentry_sdk.integrations.fastapi import FastApiIntegration
    sentry_sdk.init(
        dsn=os.getenv("SENTRY_DSN"),
        integrations=[FastApiIntegration()],
        traces_sample_rate=0.1,
        send_default_pii=False,
    )

from collections import defaultdict
from threading import Lock as _Lock
from typing import List, Optional
from fastapi import FastAPI, UploadFile, File, HTTPException, Form, Request

# ── In-memory per-user rate limiter ──────────────────────────────────────────
# State is per-process — fine for single-instance Railway deployment.
_rl_data: dict[str, list[float]] = defaultdict(list)
_rl_lock = _Lock()

def _rate_limit(key: str, limit: int, window: int = 3600) -> bool:
    """Return True if request is allowed; False if rate-limit exceeded.
    key: e.g. f"{email}:generate". limit: max calls. window: seconds."""
    now = time.time()
    with _rl_lock:
        _rl_data[key] = [t for t in _rl_data[key] if now - t < window]
        if len(_rl_data[key]) >= limit:
            return False
        _rl_data[key].append(now)
        return True
from fastapi.responses import Response, HTMLResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from starlette.middleware.base import BaseHTTPMiddleware
from pydantic import BaseModel

from app.models import (
    GenerationRequest, GenerationResult, PublishRequest, PublishResult,
    RawProduct, Marketplace, AmazonListing,
)
from app.services.ingestion import parse_file, get_headers_and_sample
from app.services.ai_agent import generate_listings_batch
from app.services.amazon_sp import publish_listings
from app.services.auth import verify_token
from app.services.image_gen import generate_product_images, AMAZON_IMAGE_TYPES, _generate_image_dalle3
from app.services.usage import log_usage, get_user_usage, get_all_users_usage
from app.services import storage
from app.utils.export import to_csv_bytes, to_json_bytes, to_amazon_flat_file_bytes, to_listing_loader_bytes, to_amazon_flat_file_xlsx, to_listing_loader_xlsx, to_variation_flat_file_xlsx
from app.utils.variation_handler import group_by_parent, build_parent_product, expand_to_variation_listings
from app.db import (init_db, save_generation, list_generations, get_generation,
                    delete_generation, update_generation_label,
                    update_generation_listings, save_job, update_job_db,
                    load_recent_jobs, add_tracked_listing, list_tracked_listings,
                    delete_tracked_listing, add_snapshot, get_snapshots,
                    delete_snapshot, get_tracking_summary,
                    save_session, list_saved_sessions, get_saved_session, delete_saved_session,
                    save_ab_experiment, list_ab_experiments, get_ab_experiment,
                    update_ab_experiment, delete_ab_experiment,
                    save_generated_image, get_generated_images, get_generated_images_bulk,
                    save_amazon_template, list_amazon_templates, get_amazon_template,
                    get_amazon_template_by_type,
                    amazon_template_exists, delete_amazon_template)
from app.services.ab_testing import (generate_ab_variants, create_amazon_experiment,
                                      get_amazon_experiment_status, cancel_amazon_experiment)
from app.routes.auth import router as auth_router, admin_router
from app.routes.amazon_oauth import router as amazon_router
from app.routes.chat import router as chat_router
from app.routes.stripe_webhook import router as stripe_router
from app.routes.insights import router as insights_router
from app.routes.drive import router as drive_router
from app.routes.drive_oauth import router as drive_oauth_router
from app.routes.agent import router as agent_router

# Routes sans authentification
PUBLIC_PATHS = {"/", "/health", "/api/auth/login", "/api/auth/setup", "/api/auth/needs-setup", "/api/marketplaces", "/api/amazon/callback", "/api/amazon/login", "/api/drive/oauth/callback", "/api/template", "/api/auth/unsubscribe", "/api/auth/forgot-password", "/api/auth/reset-password", "/api/stripe/webhook", "/api/contact"}
# Invite paths are public (token-based auth)
# /api/v1 is the public API — authenticated per-request via API key (see routes/public_api.py)
# /api/feed/ are pull feeds — the capability token in the path is the auth (see routes/feeds.py)
PUBLIC_PREFIX_PATHS = ("/api/auth/invite/", "/api/photos/", "/api/auth/reset-password/", "/api/v1", "/api/feed/")


class AuthMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        path = request.url.path

        # Laisser passer les routes publiques, les fichiers statiques et les chemins à préfixe public
        if path in PUBLIC_PATHS or not path.startswith("/api/") or any(path.startswith(p) for p in PUBLIC_PREFIX_PATHS):
            return await call_next(request)

        auth = request.headers.get("Authorization", "")
        if not auth.startswith("Bearer "):
            return JSONResponse({"detail": "Non authentifié"}, status_code=401)

        email = verify_token(auth.split(" ", 1)[1])
        if not email:
            return JSONResponse({"detail": "Session expirée, reconnectez-vous"}, status_code=401)

        # owner_email = the authenticated (agency) account — used for auth, plan,
        # quota and workspace management. user_email = the active data scope: a
        # workspace id when one is selected AND owned by this account, else the
        # account itself (personal space). All content queries key on user_email.
        request.state.owner_email = email
        ws = request.headers.get("X-Workspace", "").strip()
        if ws.startswith("ws_"):
            from app.db import workspace_owner
            if workspace_owner(ws) == email:
                request.state.user_email = ws
                return await call_next(request)
        request.state.user_email = email
        return await call_next(request)


app = FastAPI(
    title="Marketplace AI Agent",
    description="Transforme des données produits brutes en fiches marketplaces optimisées SEO",
    version="1.0.0",
)

_ALLOWED_ORIGINS = os.getenv(
    "ALLOWED_ORIGINS",
    "https://synqio.io,https://www.synqio.io",
).split(",")

app.add_middleware(
    CORSMiddleware,
    allow_origins=_ALLOWED_ORIGINS,
    allow_methods=["GET", "POST", "PUT", "DELETE", "PATCH", "OPTIONS"],
    allow_headers=["Authorization", "Content-Type", "X-Workspace", "X-Api-Key"],
    allow_credentials=True,
)
app.add_middleware(AuthMiddleware)

from starlette.middleware.base import BaseHTTPMiddleware
from starlette.responses import Response as _Response

class SecurityHeadersMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request, call_next):
        response = await call_next(request)
        response.headers["X-Frame-Options"] = "DENY"
        response.headers["X-Content-Type-Options"] = "nosniff"
        response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
        response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains"
        return response

app.add_middleware(SecurityHeadersMiddleware)

app.include_router(auth_router)
app.include_router(admin_router)
app.include_router(amazon_router)
app.include_router(chat_router)
app.include_router(stripe_router)
app.include_router(insights_router)
app.include_router(drive_router)
app.include_router(drive_oauth_router)
app.include_router(agent_router)
from app.routes.public_api import router as public_api_router
app.include_router(public_api_router)
from app.routes.feeds import router as feeds_router
app.include_router(feeds_router)


@app.on_event("startup")
async def startup():
    _check_secrets()
    sentry_dsn = os.getenv("SENTRY_DSN")
    if sentry_dsn:
        import sentry_sdk
        from sentry_sdk.integrations.fastapi import FastApiIntegration
        from sentry_sdk.integrations.starlette import StarletteIntegration
        sentry_sdk.init(
            dsn=sentry_dsn,
            integrations=[StarletteIntegration(transaction_style="endpoint"),
                          FastApiIntegration()],
            traces_sample_rate=0.05,
            environment=os.getenv("RAILWAY_ENVIRONMENT", "development"),
            release=os.getenv("RAILWAY_GIT_COMMIT_SHA", "unknown"),
        )
        log.info("Sentry initialised", extra={"endpoint": "startup"})
    from app.db import DB_PATH
    db_abs = os.path.abspath(DB_PATH)
    log.info(f"[SynqIO] DB path: {db_abs}")
    log.info(f"[SynqIO] DB on volume: {db_abs.startswith('/app/data')}")
    log.info(f"[SynqIO] DB exists: {os.path.isfile(db_abs)}")
    init_db()
    _bootstrap_admin()
    _seed_amazon_templates()
    try:
        _jobs.update(load_recent_jobs())
    except Exception:
        pass


from fastapi.responses import JSONResponse
from fastapi.exceptions import RequestValidationError

@app.exception_handler(404)
async def not_found_handler(request: Request, exc):
    return JSONResponse(status_code=404, content={"detail": "Ressource introuvable"})

@app.exception_handler(500)
async def server_error_handler(request: Request, exc):
    log.error(f"Unhandled 500 on {request.url.path}: {exc}")
    return JSONResponse(status_code=500, content={"detail": f"Erreur serveur — {type(exc).__name__}: {str(exc)[:200]}"})

@app.exception_handler(RequestValidationError)
async def validation_error_handler(request: Request, exc: RequestValidationError):
    return JSONResponse(status_code=422, content={"detail": "Données invalides", "errors": exc.errors()})


def _seed_amazon_templates():
    """Seed DB from any .xlsm/.xlsx files in data/amazon_templates/ that aren't there yet."""
    tpl_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data", "amazon_templates")
    if not os.path.isdir(tpl_dir):
        return
    for fname in os.listdir(tpl_dir):
        if not fname.endswith((".xlsm", ".xlsx")):
            continue
        # "_"-prefixed files are special bundled templates (e.g. the generic
        # Listing Loader) — seeded into the DB too (so the pack reads them
        # reliably), but kept out of the category picker via list_amazon_templates.
        product_type = fname.rsplit(".", 1)[0].split("_v")[0].upper()
        if amazon_template_exists(product_type):
            continue
        fpath = os.path.join(tpl_dir, fname)
        with open(fpath, "rb") as f:
            data_b64 = base64.b64encode(f.read()).decode()
        label = product_type.replace("_", " ").title()
        save_amazon_template(str(uuid4()), label, product_type, fname, data_b64, "system")
        log.info(f"[Templates] Seeded: {product_type}")


def _check_secrets():
    import secrets as _secrets
    jwt = os.getenv("JWT_SECRET", "")
    insecure_default = "change-me-use-a-long-random-string-in-production"
    if not jwt or jwt == insecure_default or len(jwt) < 32:
        suggestion = _secrets.token_hex(32)
        log.warning(
            f"\n{'='*60}\n"
            f"SECURITY WARNING: JWT_SECRET is weak or missing.\n"
            f"   Set this env var in Railway:\n"
            f"   JWT_SECRET={suggestion}\n"
            f"{'='*60}\n"
        )


def _bootstrap_admin():
    """Crée ou restaure le compte admin depuis ADMIN_EMAIL / ADMIN_PASSWORD.
    Appelé au démarrage : garantit que l'admin garde is_admin=1 même après
    un redéploiement Railway qui efface la DB éphémère.
    """
    email = os.getenv("ADMIN_EMAIL", "").strip().lower()
    password = os.getenv("ADMIN_PASSWORD", "").strip()
    if not email or not password:
        print("[bootstrap_admin] ADMIN_EMAIL ou ADMIN_PASSWORD manquant — admin non créé", flush=True)
        return
    from app.db import get_db
    from app.services.auth import create_user, hash_password
    try:
        conn = get_db()
        exists = conn.execute("SELECT 1 FROM users WHERE email = ?", (email,)).fetchone()
        if not exists:
            conn.close()
            create_user(email, password, name="Admin", is_admin=True)
            print(f"[bootstrap_admin] Admin créé : {email}", flush=True)
        else:
            conn.execute(
                "UPDATE users SET is_admin=1, is_active=1, password_hash=? WHERE email=?",
                (hash_password(password), email),
            )
            conn.commit()
            conn.close()
            print(f"[bootstrap_admin] Admin mis à jour : {email}", flush=True)
    except Exception as exc:
        print(f"[bootstrap_admin] ERREUR : {exc}", flush=True)


# ── Frontend ─────────────────────────────────────────────────────────────────

@app.get("/", response_class=HTMLResponse, include_in_schema=False)
async def root():
    with open("frontend/index.html", "r", encoding="utf-8") as f:
        return HTMLResponse(content=f.read())


@app.get("/sitemap.xml", include_in_schema=False)
async def sitemap():
    with open("frontend/sitemap.xml", "r", encoding="utf-8") as f:
        return Response(content=f.read(), media_type="application/xml")


def _gather_catalog_for_scan(email: str) -> list:
    """Merge the live SP-API catalog (what is actually published — title/ean/status)
    with the richest generated listing per SKU (bullets/keywords/description). Each
    listing carries status/quantity/sales so the panel can filter (active, in stock,
    selling). The live title wins when present (it is what Amazon shows).

    Important: when a live catalog exists, generated listings only *enrich* matching
    live SKUs — a generated-only draft (e.g. an audit test) never appears as its own
    row, so the panel reflects the connected Amazon catalog, not leftover test data."""
    from app.db import get_catalog, get_catalog_summary, get_catalog_sales_map
    from app.routes.public_api import _latest_listings
    merged: dict = {}
    sales_map = {}
    try:
        sales_map = get_catalog_sales_map(email)
    except Exception:
        pass
    summary = get_catalog_summary(email) or {}
    has_live_catalog = bool(summary)
    for mkt in summary:
        for it in get_catalog(email, mkt):
            asin = (it.get("asin") or "").upper()
            s = sales_map.get((mkt, asin)) or {}
            key = (mkt, it.get("sku") or "")
            entry = {"sku": it.get("sku") or "", "marketplace": mkt,
                     "title": it.get("title") or "",
                     "asin": asin, "status": (it.get("status") or "").lower(),
                     "quantity": it.get("quantity"), "price": it.get("price"),
                     "_units": s.get("units") or 0, "_revenue": s.get("revenue") or 0,
                     "_source": "live"}
            # The catalog sync report (GET_MERCHANT_LISTINGS_ALL_DATA) does NOT carry
            # the EAN — a blank here means "unknown", not "missing on Amazon". Only
            # flag EAN when we actually have the field (e.g. a real value synced).
            if (it.get("ean") or "").strip():
                entry["ean"] = it.get("ean")
            merged[key] = entry
    for g in _latest_listings(email):
        key = (g.get("marketplace") or "", g.get("sku") or "")
        if key in merged:
            base = merged[key]
            for f in ("bullet_points", "backend_keywords", "description"):
                if f in g:
                    base[f] = g[f]
            if not base.get("title"):
                base["title"] = g.get("title") or ""
        elif not has_live_catalog:
            # No Seller Central connected yet → the generated listings ARE the catalog.
            # With a live catalog present, ignore generated-only rows (test/audit drafts).
            merged[key] = {**g, "status": "active", "_source": "generated"}
    return list(merged.values())


def _today_iso() -> str:
    import datetime as _d
    return _d.date.today().isoformat()


def _sales_overview(email: str) -> dict:
    """From the tracked listings' latest snapshots: how many sell + total revenue."""
    from app.db import list_tracked_listings, get_snapshots
    tracked = list_tracked_listings(email)
    selling, revenue = 0, 0.0
    for l in tracked:
        snaps = get_snapshots(l["id"], email)
        if not snaps:
            continue
        last = snaps[-1]
        if (last.get("units_ordered") or 0) > 0:
            selling += 1
            revenue += last.get("revenue") or 0
    return {"tracked": len(tracked), "selling": selling, "revenue": round(revenue, 2)}


@app.get("/api/compliance/scan")
async def compliance_scan(request: Request):
    """Conformity Watchdog — scan the user's catalog against Amazon rules,
    with a catalog (active/inactive/stock) + sales overview."""
    from app.services.compliance import scan_listings, detect_degradation
    from app.db import (record_health_snapshot, get_health_history,
                        get_catalog_overview, get_catalog_sales_overview,
                        get_catalog_insights)
    email = request.state.user_email
    report = scan_listings(_gather_catalog_for_scan(email))
    try:
        report["catalog_overview"] = get_catalog_overview(email)
        # Prefer the catalog-wide Sales & Traffic cross-reference; fall back to
        # the tracked-listings overview until the seller pulls catalog sales.
        cat_sales = get_catalog_sales_overview(email)
        report["sales_overview"] = cat_sales if cat_sales.get("has_data") else _sales_overview(email)
    except Exception:
        pass
    try:
        report["insights"] = get_catalog_insights(email)
    except Exception:
        pass
    # Degradation vs the previous snapshot → live alert banner (compare before
    # recording today's point). Email alerts themselves are sent by the cron.
    try:
        hist = get_health_history(email, 2)
        prev = next((h for h in reversed(hist) if h.get("date") != _today_iso()), None)
        report["alert"] = detect_degradation(report, prev)
    except Exception:
        report["alert"] = []
    try:
        record_health_snapshot(email, report)  # one point/day for the trend chart
    except Exception:
        pass
    return report


@app.get("/api/compliance/history")
async def compliance_history(request: Request):
    """Catalog health score over time (for the trend chart)."""
    from app.db import get_health_history
    return {"history": get_health_history(request.state.user_email, limit=30)}


@app.post("/api/business/sync")
async def business_sync(request: Request):
    """Auto-feed the tracked listings from the Amazon Sales & Traffic report (SP-API)."""
    from app.services.business_watchdog import sync_business_snapshots
    try:
        res = await sync_business_snapshots(request.state.user_email)
    except RuntimeError as e:
        raise HTTPException(400, str(e))
    except Exception as e:
        raise HTTPException(502, f"Synchronisation impossible : {type(e).__name__}")
    return res


class CompetitorAdd(BaseModel):
    asin: str
    label: str = ""
    my_sku: str = ""
    marketplace: str = "amazon_fr"


def _my_price_for_sku(email: str, sku: str):
    """Best-effort: your own price for a SKU, from the latest generation."""
    if not sku:
        return None
    from app.routes.public_api import _latest_listings
    for l in _latest_listings(email):
        if l.get("sku") == sku and l.get("price") is not None:
            try:
                return round(float(l["price"]), 2)
            except (TypeError, ValueError):
                return None
    return None


@app.get("/api/competitors")
async def competitors_list(request: Request):
    from app.db import list_competitors
    email = request.state.user_email
    rows = list_competitors(email)
    for c in rows:
        c["my_price"] = _my_price_for_sku(email, c.get("my_sku") or "")
        if c.get("last_price") is not None and c.get("my_price") is not None:
            c["vs_you"] = round(c["last_price"] - c["my_price"], 2)  # >0 = concurrent plus cher
    return {"competitors": rows}


@app.post("/api/competitors")
async def competitors_add(body: CompetitorAdd, request: Request):
    from app.db import list_competitors, add_competitor
    email = request.state.user_email
    asin = (body.asin or "").strip().upper()
    if not (5 <= len(asin) <= 12) or not asin.isalnum():
        raise HTTPException(400, "ASIN invalide (ex : B0XXXXXXXX)")
    if len(list_competitors(email)) >= 100:
        raise HTTPException(400, "Maximum 100 concurrents suivis")
    add_competitor(f"cmp_{uuid4().hex[:12]}", email, asin, body.marketplace,
                   (body.label or "").strip()[:80], (body.my_sku or "").strip())
    return {"ok": True}


@app.delete("/api/competitors/{cid}")
async def competitors_delete(cid: str, request: Request):
    from app.db import delete_competitor
    if not delete_competitor(cid, request.state.user_email):
        raise HTTPException(404, "Concurrent introuvable")
    return {"deleted": True}


@app.post("/api/competitors/refresh")
async def competitors_refresh(request: Request):
    """Scrape current prices for all watched competitors via Apify, store
    snapshots, and return per-competitor price + change vs last check."""
    from app.db import list_competitors, update_competitor_price
    email = request.state.user_email
    comps = list_competitors(email)
    if not comps:
        return {"refreshed": 0, "competitors": []}
    if not _rate_limit(f"{email}:cmp-refresh", limit=6, window=3600):
        raise HTTPException(429, "Limite de 6 rafraîchissements/heure — réessayez plus tard")
    # group ASINs by marketplace
    by_mkt: dict = {}
    for c in comps:
        by_mkt.setdefault(c["marketplace"], []).append(c["asin"])
    scraped: dict = {}
    for mkt, asins in by_mkt.items():
        try:
            scraped.update(await _apify_fetch_products(list(set(asins)), mkt))
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(502, f"Scraping impossible : {type(e).__name__}")

    results = []
    refreshed = 0
    for c in comps:
        data = scraped.get(c["asin"])
        prev = c.get("last_price")
        if data:
            update_competitor_price(c["id"], email, data["price"], data["title"], data["rating"])
            refreshed += 1
        my_price = _my_price_for_sku(email, c.get("my_sku") or "")
        cur = (data or {}).get("price")
        results.append({
            "id": c["id"], "asin": c["asin"], "label": c.get("label") or "",
            "title": (data or {}).get("title") or c.get("last_title") or "",
            "price": cur, "rating": (data or {}).get("rating"),
            "prev_price": prev,
            "delta": (round(cur - prev, 2) if (cur is not None and prev is not None) else None),
            "my_price": my_price,
            "vs_you": (round(cur - my_price, 2) if (cur is not None and my_price is not None) else None),
            "found": bool(data),
        })
    return {"refreshed": refreshed, "competitors": results}


@app.get("/api/business/scan")
async def business_scan(request: Request):
    """Business Watchdog — detect declining tracked listings, correlated with
    content/compliance issues on the same SKU."""
    from app.services.business_watchdog import scan_business, annotate_content_issues
    from app.services.compliance import scan_listings
    email = request.state.user_email
    report = scan_business(email)
    if report.get("alerting"):
        try:
            content = scan_listings(_gather_catalog_for_scan(email))
            bad_skus = {i.get("sku") for i in content.get("listings", []) if i.get("sku")}
            annotate_content_issues(report, bad_skus)
        except Exception:
            pass
    return report


_TITLE_ISSUE_CODES = {"title_too_long", "title_forbidden_chars", "title_all_caps", "title_superlative"}


@app.post("/api/compliance/fix-titles")
async def compliance_fix_titles(request: Request):
    """Close the loop — auto-fix non-compliant titles across all stored
    generations (≤75 chars + Item Highlights) and persist them. Only rewrites
    SynqIO-generated listings (not live Amazon catalog items)."""
    if not os.getenv("ANTHROPIC_API_KEY"):
        raise HTTPException(503, "Service IA indisponible")
    from app.services.compliance import check_listing
    from app.db import get_generation, save_generation
    from app.routes.public_api import _latest_listings
    email = request.state.user_email
    CAP = 40  # bound AI cost/time per click

    targets = []
    for l in _latest_listings(email):
        if not l.get("batch_id"):
            continue  # live-catalog-only item — cannot rewrite here
        if any(i["code"] in _TITLE_ISSUE_CODES for i in check_listing(l)):
            targets.append(l)
    total_flagged = len(targets)
    if not total_flagged:
        return {"fixed": 0, "remaining": 0, "message": "Aucun titre à corriger."}
    targets = targets[:CAP]

    sem = asyncio.Semaphore(5)

    async def _fix(l):
        async with sem:
            try:
                res = await _reformat_title_core(
                    title=l.get("title", ""), brand=l.get("brand", ""),
                    category=l.get("category", ""), bullets=l.get("bullet_points") or [],
                    item_highlights=l.get("item_highlights", ""),
                    backend_keywords=l.get("backend_keywords", ""),
                    marketplace=l.get("marketplace", "amazon_fr"))
                return (l, res)
            except Exception:
                return (l, None)

    results = await asyncio.gather(*[_fix(l) for l in targets])

    by_batch: dict = {}
    for l, res in results:
        if res:
            by_batch.setdefault(l["batch_id"], {})[l.get("sku")] = res

    fixed = 0
    for batch_id, fixes in by_batch.items():
        gen = get_generation(batch_id, email)
        if not gen:
            continue
        listings = gen.get("listings", [])
        changed = False
        for item in listings:
            r = fixes.get(item.get("sku"))
            if r:
                item["title"] = r["title"]
                if r.get("item_highlights"):
                    item["item_highlights"] = r["item_highlights"]
                if r.get("seo_score") is not None:
                    item["seo_score"] = r["seo_score"]
                changed = True
                fixed += 1
        if changed:
            save_generation(email, batch_id, gen.get("marketplace", "amazon_fr"),
                            listings, gen.get("label", ""))

    return {"fixed": fixed, "remaining": max(0, total_flagged - fixed)}


class WorkOnRequest(BaseModel):
    skus: list[str]
    enrich: bool = False


@app.post("/api/compliance/work-on")
async def compliance_work_on(req: WorkOnRequest, request: Request):
    """Load selected catalog listings into the generator to rework them. Optionally
    enrich each with the bullets/description/images pulled per-SKU from Amazon."""
    from app.db import get_catalog_rows_by_skus
    email = request.state.user_email
    skus = [s for s in (req.skus or []) if s]
    if not skus:
        raise HTTPException(400, "Aucune fiche sélectionnée")
    rows = get_catalog_rows_by_skus(email, skus)
    by_sku = {r["sku"]: r for r in rows}

    # Enrichment calls Amazon once per SKU, so it's bounded to keep the request
    # under the gateway timeout; the base data (title/price/ean) loads for ALL
    # selected fiches regardless.
    ENRICH_MAX = 60
    enriched: dict = {}
    enrich_note = ""
    if req.enrich:
        from app.services.amazon_sp import fetch_listing_details
        from app.models import Marketplace
        errs: list = []
        remaining = ENRICH_MAX
        by_mkt: dict = {}
        for r in rows:
            by_mkt.setdefault(r.get("marketplace") or "amazon_fr", []).append(r["sku"])
        for mkt_str, mkt_skus in by_mkt.items():
            if remaining <= 0:
                break
            try:
                mkt = Marketplace(mkt_str)
            except ValueError:
                mkt = Marketplace.AMAZON_FR
            try:
                res = await fetch_listing_details(email, mkt, mkt_skus, limit=remaining, errors=errs)
                enriched.update(res)
                remaining -= len(mkt_skus)
            except RuntimeError as e:
                enrich_note = str(e)
            except Exception:
                pass  # enrichment is best-effort; base data still returned
        if not enriched and not enrich_note:
            if any(e in (403, 401) for e in errs):
                enrich_note = ("Enrichissement refusé par Amazon (rôle « Product Listing » "
                               "manquant sur l'app). Les fiches sont chargées sans bullets/description.")
            elif errs:
                enrich_note = "Enrichissement indisponible pour le moment — fiches chargées sans contenu Amazon."
        elif len(skus) > ENRICH_MAX:
            enrich_note = f"Enrichissement limité aux {ENRICH_MAX} premières fiches (appel Amazon par fiche)."

    products = []
    for sku in skus:
        r = by_sku.get(sku, {"sku": sku})
        e = enriched.get(sku, {})
        products.append({
            "sku": sku,
            "name": e.get("title") or r.get("title") or sku,
            "brand": e.get("brand", "") or "",
            "ean": r.get("ean") or "",
            "price": r.get("price"),
            "description": e.get("description", "") or "",
            "features": e.get("bullet_points", []) or [],
            "images": e.get("images", []) or [],
            "asin": r.get("asin") or "",
        })
    return {"products": products, "enriched": len(enriched),
            "total": len(products), "note": enrich_note}


@app.post("/api/compliance/send-digest")
async def compliance_send_digest(request: Request):
    """Send the health digest by email. Body may carry {"to": "..."} to send it to
    a chosen address (defaults to the logged-in account) — handy to demo/preview it."""
    import re as _re
    from app.services.compliance import scan_listings
    from app.services.email import send_catalog_health, _can_send
    email = request.state.user_email  # workspace-scoped: what catalog to scan
    owner = getattr(request.state, "owner_email", None) or email
    to = owner
    try:
        body = await request.json()
        cand = (body or {}).get("to", "").strip()
        if cand:
            if not _re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", cand):
                raise HTTPException(400, "Adresse email invalide")
            to = cand
    except HTTPException:
        raise
    except Exception:
        pass  # no/invalid body → send to the account owner
    report = scan_listings(_gather_catalog_for_scan(email))
    if not report.get("total"):
        return {"sent": False, "reason": "empty_catalog", "to": to}
    if not _can_send():
        return {"sent": False, "reason": "smtp_unconfigured", "to": to}
    sent = await asyncio.get_event_loop().run_in_executor(None, send_catalog_health, to, report)
    return {"sent": sent, "to": to, "score": report.get("score"),
            "non_compliant": report.get("non_compliant")}


@app.post("/api/admin/compliance/send-all-digests")
async def compliance_send_all(request: Request):
    """Send the weekly digest to every active user. Intended to be triggered by a
    scheduler (e.g. a weekly Railway cron hitting this endpoint). Admin only."""
    from app.db import get_db
    from app.services.compliance import scan_listings
    from app.services.email import send_catalog_health
    conn = get_db()
    row = conn.execute("SELECT is_admin FROM users WHERE email = ?", (request.state.user_email,)).fetchone()
    if not (row and row["is_admin"]):
        conn.close()
        raise HTTPException(403, "Réservé aux admins")
    emails = [r["email"] for r in conn.execute("SELECT email FROM users").fetchall()]
    conn.close()
    sent = 0
    for em in emails:
        try:
            report = scan_listings(_gather_catalog_for_scan(em))
            if report.get("total", 0) and await asyncio.get_event_loop().run_in_executor(None, send_catalog_health, em, report):
                sent += 1
        except Exception:
            continue
    return {"users": len(emails), "digests_sent": sent}


class ContactRequest(BaseModel):
    name: str = ""
    first_name: str = ""
    email: str
    message: str = ""


@app.post("/api/contact")
async def api_contact(req: ContactRequest, request: Request):
    """Public contact form (marketing pages). Emails the SynqIO inbox."""
    email = (req.email or "").strip()
    if "@" not in email or "." not in email.split("@")[-1] or len(email) < 6:
        raise HTTPException(400, "Adresse email invalide")
    # Simple abuse guard: 5 submissions / hour per IP
    ip = request.client.host if request.client else "unknown"
    if not _rate_limit(f"contact:{ip}", limit=5, window=3600):
        raise HTTPException(429, "Trop de demandes — réessayez plus tard")
    full_name = " ".join(p for p in [(req.first_name or "").strip(), (req.name or "").strip()] if p) or "—"
    from app.services.email import send_contact
    sent = await asyncio.get_event_loop().run_in_executor(
        None, send_contact, full_name, email, req.message or ""
    )
    # `sent` is False only when SMTP isn't configured; the UI falls back to Calendly
    return {"ok": True, "delivered": sent}


# ── Favicon (Google Search needs a real PNG/ICO ≥48px, with correct MIME) ─────

_FAVICON_SVG = b"""<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 32 32"><rect width="32" height="32" rx="6" fill="#764BA2"/><path d="M16 4L5 9.5l11 5.5 11-5.5L16 4z" stroke="white" stroke-width="1.8" stroke-linejoin="round" fill="none"/><path d="M5 22l11 5.5 11-5.5M5 16l11 5.5 11-5.5" stroke="white" stroke-width="1.8" stroke-linejoin="round" fill="none"/></svg>"""

def _serve_static_icon(path: str, media_type: str):
    try:
        with open(path, "rb") as f:
            return Response(content=f.read(), media_type=media_type,
                            headers={"Cache-Control": "public, max-age=86400"})
    except FileNotFoundError:
        return Response(content=_FAVICON_SVG, media_type="image/svg+xml",
                        headers={"Cache-Control": "public, max-age=86400"})

@app.get("/favicon.svg", include_in_schema=False)
async def favicon_svg():
    return Response(content=_FAVICON_SVG, media_type="image/svg+xml",
                    headers={"Cache-Control": "public, max-age=86400"})

@app.get("/favicon.ico", include_in_schema=False)
async def favicon_ico():
    return _serve_static_icon("frontend/favicon.ico", "image/x-icon")

@app.get("/favicon-96.png", include_in_schema=False)
async def favicon_png96():
    return _serve_static_icon("frontend/favicon-96.png", "image/png")

@app.get("/favicon-192.png", include_in_schema=False)
async def favicon_png192():
    return _serve_static_icon("frontend/favicon-192.png", "image/png")


# ── CoBrowse config ───────────────────────────────────────────────────────────

@app.get("/api/cobrowse-config")
async def cobrowse_config(request: Request):
    from app.db import get_config
    key = os.getenv("COBROWSE_LICENSE_KEY", "").strip() or get_config("COBROWSE_LICENSE_KEY", "").strip()
    return {"license_key": key}


# ── Health ────────────────────────────────────────────────────────────────────

@app.get("/health")
async def health():
    from app.db import DB_PATH
    db_abs = os.path.abspath(DB_PATH)
    db_exists = os.path.isfile(db_abs)
    db_size = os.path.getsize(db_abs) if db_exists else 0
    user_count = 0
    try:
        from app.db import get_db
        conn = get_db()
        row = conn.execute("SELECT COUNT(*) AS cnt FROM users").fetchone()
        user_count = row["cnt"] if row else 0
        conn.close()
    except Exception:
        pass
    return {
        "status": "ok",
        "ai_ready": bool(os.getenv("ANTHROPIC_API_KEY")),
        "version": "1.0.1",
        "git_commit": os.getenv("RAILWAY_GIT_COMMIT_SHA", "unknown")[:7],
        "db_path": db_abs,
        "db_on_volume": db_abs.startswith("/app/data"),
        "db_exists": db_exists,
        "db_size_bytes": db_size,
        "user_count": user_count,
    }


# ── Template ─────────────────────────────────────────────────────────────────

@app.get("/api/template")
async def download_template():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()

    # ── Onglet 1 : Produits ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Produits"

    headers = [
        "SKU", "Nom", "Marque", "Segment", "EAN", "Prix",
        "Image_Fichier", "Poids_kg", "Dimensions_cm",
        "Caractéristiques", "Mots_cles",
        "Parent_SKU", "Variation_Theme", "Variation_Value", "Taille",
    ]
    mandatory = {"SKU", "Nom", "Marque", "Segment"}
    photo_col = {"Image_Fichier"}
    variation_col = {"Parent_SKU", "Variation_Theme", "Variation_Value", "Taille"}

    # Styles en-tête
    fill_mandatory = PatternFill("solid", fgColor="764BA2")
    fill_photo     = PatternFill("solid", fgColor="0F766E")   # vert sarcelle pour Image_Fichier
    fill_optional  = PatternFill("solid", fgColor="B39DDB")
    fill_variation = PatternFill("solid", fgColor="2563EB")   # bleu pour colonnes variation
    font_header    = Font(color="FFFFFF", bold=True, size=11)
    font_example   = Font(color="555555", italic=True, size=10)
    center         = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, h in enumerate(headers, 1):
        label = h + (" *" if h in mandatory else "")
        cell = ws.cell(row=1, column=col, value=label)
        if h in mandatory:
            cell.fill = fill_mandatory
        elif h in photo_col:
            cell.fill = fill_photo
        elif h in variation_col:
            cell.fill = fill_variation
        else:
            cell.fill = fill_optional
        cell.font = font_header
        cell.alignment = center

    # Exemple 1 — collier (standalone)
    ex1 = [
        "BC-COL-001", "Collier Étoile Dorée Chien M", "Bande de Canailles",
        "collier chien", "3760123456789", "24.90",
        "collier-etoile-doree.jpg", "0.085", "35x2 cm",
        "Pendentif étoile doré|Boucle sécurité|Réglage 5 positions|Nylon recyclé certifié",
        "collier chien tendance, collier chien pendentif, collier chien fantaisie",
        "", "", "", "",
    ]
    # Exemple 2 — laisse déclinaison couleur/taille (variation)
    ex2 = [
        "BC-LAI-012-ROUGE-S", "Laisse Léopard Chien Rouge S", "Bande de Canailles",
        "laisse chien", "3760123456790", "29.90",
        "laisse-leopard-rouge.jpg", "0.120", "120x2 cm",
        "Mousqueton inox doré|Poignée rembourrée|Motif léopard|Longueur 1.2m",
        "laisse chien design, laisse chien fantaisie, laisse chien colorée",
        "BC-LAI-012", "ColorName-SizeClass", "Rouge / S", "S",
    ]
    ex3 = [
        "BC-LAI-012-BLEU-M", "Laisse Léopard Chien Bleu M", "Bande de Canailles",
        "laisse chien", "3760123456791", "29.90",
        "laisse-leopard-bleu.jpg", "0.120", "120x2 cm",
        "Mousqueton inox doré|Poignée rembourrée|Motif léopard|Longueur 1.2m",
        "laisse chien design, laisse chien fantaisie, laisse chien colorée",
        "BC-LAI-012", "ColorName-SizeClass", "Bleu / M", "M",
    ]

    for col, v in enumerate(ex1, 1):
        cell = ws.cell(row=2, column=col, value=v)
        cell.font = font_example
    for col, v in enumerate(ex2, 1):
        cell = ws.cell(row=3, column=col, value=v)
        cell.font = font_example
    for col, v in enumerate(ex3, 1):
        cell = ws.cell(row=4, column=col, value=v)
        cell.font = font_example

    # Largeurs de colonnes
    widths = [18, 34, 20, 18, 16, 8, 30, 10, 14, 48, 48, 16, 20, 18, 10]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # ── Onglet 2 : Guide ─────────────────────────────────────────────────────
    guide = wb.create_sheet("Guide")
    guide.column_dimensions["A"].width = 20
    guide.column_dimensions["B"].width = 14
    guide.column_dimensions["C"].width = 68
    guide.column_dimensions["D"].width = 40

    guide_headers = ["Colonne", "Obligatoire ?", "Description", "Exemple"]
    for col, h in enumerate(guide_headers, 1):
        cell = guide.cell(row=1, column=col, value=h)
        cell.fill = PatternFill("solid", fgColor="1F2937")
        cell.font = Font(color="FFFFFF", bold=True)

    guide_rows = [
        ("SKU",             "✅ Oui",       "Référence unique du produit. Doit être stable — c'est la clé de matching.", "BC-COL-001"),
        ("Nom",             "✅ Oui",       "Nom brut du produit, sans optimisation SEO — l'IA s'en charge.", "Collier Étoile Dorée M"),
        ("Marque",          "✅ Oui",       "Nom exact de la marque, tel qu'il apparaît sur Amazon.", "Bande de Canailles"),
        ("Segment",         "✅ Oui",       "Famille produit. Tous les produits du même segment partagent les mêmes mots-clés si Mots_cles est vide.", "collier chien"),
        ("EAN",             "Recommandé",  "Code-barres GTIN-13. Obligatoire pour créer une fiche Amazon.", "3760123456789"),
        ("Prix",            "Recommandé",  "Prix de vente TTC en euros. Décimales avec point ou virgule.", "24.90"),
        ("Image_Fichier",   "📸 Clé images","Nom exact du fichier photo dans votre ZIP (sans chemin). "
                                           "Uploadez le ZIP via 'Uploader photos produits'. "
                                           "L'IA analyse la photo pour extraire couleurs, matières et détails — "
                                           "inutile de remplir ces champs à la main. "
                                           "Gardez les noms d'origine, aucun renommage nécessaire.", "laisse-leopard.jpg"),
        ("Poids_kg",        "Optionnel",   "Poids en kg. Aide l'IA à rédiger les bullet points techniques.", "0.085"),
        ("Dimensions_cm",   "Optionnel",   "Format L×l×h ou L×l en cm.", "35x2 cm"),
        ("Caractéristiques","Optionnel",   "Points clés du produit, séparés par |. Si vide, l'IA les déduit de la photo et du nom.", "Mousqueton inox|Poignée rembourrée|Motif léopard"),
        ("Mots_cles",       "Optionnel",   "Mots-clés SEO spécifiques à CE produit, séparés par virgule. Écrasent les mots-clés globaux de l'interface pour ce produit.", "laisse chien design, laisse fantaisie"),
        ("Parent_SKU",      "Déclinaison", "SKU du produit parent Amazon. Remplir uniquement pour les déclinaisons (plusieurs lignes du même produit avec des couleurs/tailles différentes). Toutes les lignes ayant le même Parent_SKU sont regroupées et génèrent UNE fiche parent + autant de fiches enfants.", "BC-LAI-012"),
        ("Variation_Theme", "Déclinaison", "Type de variation Amazon. Valeurs courantes : ColorName, SizeClass, ColorName-SizeClass. Doit être identique pour toutes les déclinaisons du même groupe.", "ColorName-SizeClass"),
        ("Variation_Value", "Déclinaison", "Valeur de la déclinaison pour cette ligne. Format libre, ex: 'Rouge / L', 'Bleu / M', 'Noir'. Apparaît dans la table des déclinaisons.", "Rouge / S"),
        ("Taille",          "Déclinaison", "Taille spécifique de cette déclinaison (S, M, L, XL, 30, 40…). Renseigné si la variation porte sur la taille.", "S"),
    ]

    for r, row_data in enumerate(guide_rows, 2):
        for col, val in enumerate(row_data, 1):
            cell = guide.cell(row=r, column=col, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if r % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F9FAFB")
            # Mettre en évidence la ligne Image_Fichier
            if row_data[0] == "Image_Fichier":
                cell.fill = PatternFill("solid", fgColor="CCFBF1")
                cell.font = Font(color="0F4C40", bold=(col == 1))
        guide.row_dimensions[r].height = 50 if row_data[0] == "Image_Fichier" else 36

    guide.row_dimensions[1].height = 24

    # ── Légende ──────────────────────────────────────────────────────────────
    legend_row = len(guide_rows) + 3
    guide.cell(row=legend_row,   column=1, value="* Colonnes violettes foncées = obligatoires").font = Font(bold=True, color="764BA2")
    guide.cell(row=legend_row+1, column=1, value="* Colonne verte = photo de référence (recommandée fortement)").font = Font(bold=True, color="0F766E")
    guide.cell(row=legend_row+2, column=1, value="* Colonnes violettes claires = optionnelles").font = Font(italic=True, color="9333EA")
    guide.cell(row=legend_row+3, column=1, value="* Colonnes bleues = déclinaisons (variations Amazon)").font = Font(bold=True, color="1D4ED8")
    guide.cell(row=legend_row+4, column=1,
        value="Conseil : fournissez toujours la photo via Image_Fichier + ZIP. "
              "L'IA analyse la photo et extrait automatiquement couleurs, matières et détails. "
              "Moins vous remplissez de texte, plus la photo prime.").font = Font(italic=True, color="6B7280")

    # ── Export ────────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=synqio_template.xlsx"},
    )


# ── Photo store (R2 or local disk) ────────────────────────────────────────────

from app.db import _PROJECT_ROOT

_IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".webp", ".gif"}
_IMAGE_CONTENT_TYPES = {
    "jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png",
    "webp": "image/webp", "gif": "image/gif",
}


def _save_photo(filename: str, data: bytes, content_type: str = "application/octet-stream"):
    storage.put(filename, data, content_type)


def _load_photo(filename: str) -> bytes | None:
    return storage.get(filename)


_ZIP_MAX_COMPRESSED   = 100 * 1024 * 1024   # 100 MB compressed
_ZIP_MAX_UNCOMPRESSED = 500 * 1024 * 1024   # 500 MB total decompressed
_ZIP_MAX_SINGLE_FILE  =  50 * 1024 * 1024   # 50 MB per image


@app.post("/api/upload-photos-zip")
async def upload_photos_zip(request: Request):
    import zipfile

    content = await request.body()
    if not content:
        raise HTTPException(400, "Fichier vide")
    if len(content) > _ZIP_MAX_COMPRESSED:
        raise HTTPException(413, "ZIP trop volumineux (max 100 Mo)")

    try:
        zf = zipfile.ZipFile(io.BytesIO(content))
    except Exception:
        raise HTTPException(400, "Fichier ZIP invalide")

    # Zip-bomb guard: check total decompressed size before extracting anything
    total_uncompressed = sum(i.file_size for i in zf.infolist())
    if total_uncompressed > _ZIP_MAX_UNCOMPRESSED:
        raise HTTPException(413, f"Contenu ZIP trop volumineux décompressé (max 500 Mo)")

    matched = []
    skipped = []

    for entry in zf.namelist():
        # Ignorer dossiers et métadonnées macOS
        if entry.endswith("/") or "__MACOSX" in entry:
            continue
        basename = entry.split("/")[-1]
        # Ignorer fichiers cachés (ex: .DS_Store) mais pas les entrées "./photo.jpg"
        if not basename or basename.startswith("."):
            continue
        stem, ext = os.path.splitext(basename)
        if ext.lower() not in _IMAGE_EXTS:
            continue
        info = zf.getinfo(entry)
        if info.file_size > _ZIP_MAX_SINGLE_FILE:
            skipped.append(basename)
            continue
        sku = stem
        filename = f"{sku}{ext.lower()}"
        data = zf.read(entry)
        _save_photo(filename, data)
        matched.append({"sku": sku, "url": f"/api/photos/{filename}", "filename": filename})

    return {"matched": matched, "skipped": skipped, "count": len(matched)}


@app.get("/api/photos/list")
async def list_photos(request: Request):
    # This endpoint lives under the public /api/photos/ prefix (so images load in
    # <img> tags without auth), but it leaks every user's filenames/SKUs — gate it
    # behind an admin JWT explicitly.
    from app.services.auth import is_admin, _decode_token_data
    auth = request.headers.get("Authorization", "")
    token = auth.split(" ", 1)[1] if auth.startswith("Bearer ") else ""
    email = verify_token(token) if token else None
    if not email:
        raise HTTPException(401, "Non authentifié")
    jwt_data = _decode_token_data(token) or {}
    if not (bool(jwt_data.get("adm")) or is_admin(email)):
        raise HTTPException(403, "Accès réservé aux administrateurs")
    from app.services.storage import USE_R2, LOCAL_DIR
    files = storage.list_keys()
    image_exts = set(_IMAGE_EXTS)
    image_files = [f for f in files if os.path.splitext(f)[1].lower() in image_exts]
    return {
        "backend": "r2" if USE_R2 else "disk",
        "location": f"r2://{os.getenv('R2_BUCKET_NAME','')}" if USE_R2 else LOCAL_DIR,
        "count": len(image_files),
        "files": sorted(image_files),
    }


@app.get("/api/photos/{filename}")
async def serve_photo(filename: str):
    # When R2 public access is enabled, redirect directly to CDN (no server bandwidth)
    cdn = storage.public_url(filename)
    if cdn:
        return Response(status_code=302, headers={"Location": cdn})
    photo = _load_photo(filename)
    if not photo:
        raise HTTPException(404, "Photo non trouvée")
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else "jpg"
    return Response(content=photo, media_type=_IMAGE_CONTENT_TYPES.get(ext, "image/jpeg"))


# ── URL Scraping ──────────────────────────────────────────────────────────────

def _reject_internal_url(url: str) -> None:
    """SSRF guard: block requests to private / loopback / link-local addresses
    (e.g. cloud metadata at 169.254.169.254, localhost, internal services)."""
    import ipaddress
    import socket
    from urllib.parse import urlparse
    parsed = urlparse(url)
    if parsed.scheme not in ("http", "https"):
        raise HTTPException(400, "URL invalide — utilisez http:// ou https://")
    host = parsed.hostname
    if not host:
        raise HTTPException(400, "URL invalide")
    try:
        infos = socket.getaddrinfo(host, None)
    except Exception:
        raise HTTPException(502, "Nom de domaine introuvable")
    for info in infos:
        ip = ipaddress.ip_address(info[4][0])
        if ip.is_private or ip.is_loopback or ip.is_link_local or ip.is_reserved or ip.is_multicast:
            raise HTTPException(400, "URL non autorisée (adresse interne)")


@app.post("/api/scrape-url")
async def scrape_url_endpoint(request: Request):
    """Fetch a product URL and extract name, brand, price, EAN, images, etc."""
    body = await request.json()
    url = (body.get("url") or "").strip()
    if not url.startswith("http"):
        raise HTTPException(400, "URL invalide — elle doit commencer par http:// ou https://")
    _reject_internal_url(url)
    # Follow redirects manually so each hop is re-validated against the SSRF guard
    # (a public site could otherwise 302 to an internal address).
    try:
        async with httpx.AsyncClient(
            follow_redirects=False, timeout=15.0,
            headers={"User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            )},
        ) as client:
            current = url
            for _ in range(6):
                resp = await client.get(current)
                if resp.is_redirect and resp.headers.get("location"):
                    current = str(resp.next_request.url) if resp.next_request else resp.headers["location"]
                    _reject_internal_url(current)
                    continue
                break
            resp.raise_for_status()
    except HTTPException:
        raise
    except httpx.TimeoutException:
        raise HTTPException(504, "Le site met trop de temps à répondre (délai 15s dépassé)")
    except httpx.HTTPStatusError as e:
        raise HTTPException(502, f"Le site a renvoyé une erreur {e.response.status_code}")
    except Exception as e:
        raise HTTPException(502, f"Impossible d'accéder à cette URL : {e}")

    from app.utils.url_scraper import scrape_product
    try:
        data = scrape_product(resp.text, url)
    except ValueError as e:
        raise HTTPException(422, str(e))
    return data


# ── Ingestion ─────────────────────────────────────────────────────────────────

MAX_UPLOAD_SIZE = 10 * 1024 * 1024  # 10 MB


class FilePreviewRequest(BaseModel):
    filename: str
    content_b64: str


class FileIngestRequest(BaseModel):
    filename: str
    content_b64: str
    custom_mapping: Optional[dict] = None


def _decode_upload(filename: str, content_b64: str) -> bytes:
    try:
        content = base64.b64decode(content_b64)
    except Exception:
        raise HTTPException(400, "Contenu base64 invalide")
    if not content:
        raise HTTPException(400, "Fichier vide")
    if len(content) > MAX_UPLOAD_SIZE:
        raise HTTPException(413, "Fichier trop volumineux (max 10 Mo)")
    return content


@app.post("/api/ingest/preview")
async def ingest_preview(payload: FilePreviewRequest):
    """Return headers, sample rows and auto-mapping info without full parsing."""
    content = _decode_upload(payload.filename, payload.content_b64)
    try:
        result = get_headers_and_sample(payload.filename or "upload.csv", content)
    except ValueError as e:
        raise HTTPException(400, str(e))
    except Exception as e:
        raise HTTPException(500, f"Erreur de prévisualisation: {e}")
    return result


@app.post("/api/ingest", response_model=List[RawProduct])
async def ingest_file(payload: FileIngestRequest):
    content = _decode_upload(payload.filename, payload.content_b64)
    try:
        products = parse_file(payload.filename or "upload.csv", content, custom_mapping=payload.custom_mapping)
    except ValueError as e:
        raise HTTPException(400, str(e))
    except Exception as e:
        raise HTTPException(500, f"Erreur de parsing: {e}")
    if not products:
        raise HTTPException(422, "Aucun produit trouvé dans le fichier")
    return products


# ── Async job store ───────────────────────────────────────────────────────────

_jobs: dict = {}  # job_id → {status, progress, total, result, error, created_at}

# ── Keyword suggestion cache ──────────────────────────────────────────────────

_kw_cache: dict = {}  # "query:marketplace" → {ts, keywords}
_KW_CACHE_TTL = 1800  # 30 min


def _cleanup_jobs():
    cutoff = time.time() - 3600
    for k in [k for k, v in _jobs.items() if (v.get("created_at") or 0) < cutoff]:
        del _jobs[k]


async def _run_generation_job(job_id: str, request: GenerationRequest, email: str):
    _jobs[job_id]["status"] = "running"
    log.info("job started", extra={"job_id": job_id})
    target_mkts = request.marketplaces or [request.marketplace]

    # Group products by parent_sku to detect variation groups
    groups = group_by_parent(request.products)

    # Build list of products to send to AI (one per group: synthetic parent or standalone)
    products_for_ai = []
    variation_groups: dict = {}  # parent_sku → List[RawProduct] (only for variation groups)
    for parent_sku, group_children in groups.items():
        is_variation = len(group_children) > 1 or group_children[0].parent_sku
        if is_variation:
            parent_product = build_parent_product(parent_sku, group_children)
            products_for_ai.append(parent_product)
            variation_groups[parent_sku] = group_children
        else:
            products_for_ai.append(group_children[0])

    n_total = len(products_for_ai) * len(target_mkts)
    done_count = 0

    def on_progress(done: int, _total: int):
        nonlocal done_count
        done_count += 1
        if job_id in _jobs:
            _jobs[job_id]["progress"] = done_count

    try:
        tasks = [
            generate_listings_batch(
                products=products_for_ai,
                marketplace=mkt,
                focus_keywords=request.focus_keywords or [],
                style_tone=request.style_tone,
                brand_voice=request.brand_voice,
                on_progress=on_progress,
            )
            for mkt in target_mkts
        ]
        results_all = await asyncio.gather(*tasks)

        all_listings, all_failed, total_in, total_out = [], [], 0, 0
        for listings, failed, tokens in results_all:
            # Expand variation groups into parent + child listings
            expanded = []
            for listing in listings:
                if listing.sku in variation_groups:
                    parent_l, child_ls = expand_to_variation_listings(listing, variation_groups[listing.sku])
                    expanded.append(parent_l)
                    expanded.extend(child_ls)
                else:
                    expanded.append(listing)
            all_listings.extend(expanded)
            all_failed.extend(failed)
            total_in  += tokens.get("input_tokens", 0)
            total_out += tokens.get("output_tokens", 0)

        # Count SKUs (children count individually)
        sku_count = sum(1 + len(l.children) if l.is_parent else 1 for l in all_listings)
        if sku_count and email:
            log_usage(email, "sku_generated", sku_count)
        if email:
            if total_in:  log_usage(email, "tokens_in",  total_in)
            if total_out: log_usage(email, "tokens_out", total_out)

        # Quota 80% alert: check after logging usage
        if email and sku_count:
            try:
                from app.db import get_db as _get_db
                usage_data = get_user_usage(email)
                skus_used = usage_data.get("skus_used", 0)
                skus_quota = usage_data.get("skus_quota", 0)
                if skus_quota and skus_used >= skus_quota * 0.8:
                    from datetime import datetime as _dt
                    _annual = usage_data.get("annual_pool", False)
                    current_period = _dt.utcnow().strftime("%Y" if _annual else "%Y-%m")
                    _conn = _get_db()
                    _row = _conn.execute(
                        "SELECT quota_alert_sent, quota_alert_period FROM users WHERE email = ?", (email,)
                    ).fetchone()
                    alert_sent = int((_row["quota_alert_sent"] if _row else None) or 0)
                    alert_period = (_row["quota_alert_period"] if _row else None) or ""
                    _conn.close()
                    if not alert_sent or alert_period != current_period:
                        from app.services.email import send_quota_alert
                        plan_label = usage_data.get("plan_label", "")
                        asyncio.get_event_loop().run_in_executor(
                            None, send_quota_alert, email, skus_used, skus_quota, plan_label
                        )
                        _conn2 = _get_db()
                        _conn2.execute(
                            "UPDATE users SET quota_alert_sent = 1, quota_alert_period = ? WHERE email = ?",
                            (current_period, email),
                        )
                        _conn2.commit()
                        _conn2.close()
                # Quota 100% : email upsell (1 seul envoi par période)
                if skus_quota and skus_used >= skus_quota:
                    from datetime import datetime as _dt2
                    _annual2 = usage_data.get("annual_pool", False)
                    period2 = _dt2.utcnow().strftime("%Y" if _annual2 else "%Y-%m")
                    _c3 = _get_db()
                    _r3 = _c3.execute(
                        "SELECT quota_full_period FROM users WHERE email = ?", (email,)
                    ).fetchone()
                    full_period = (_r3["quota_full_period"] if _r3 else None) or ""
                    _c3.close()
                    if full_period != period2:
                        from app.services.email import send_quota_reached
                        from app.services.usage import get_upgrade_suggestion
                        asyncio.get_event_loop().run_in_executor(
                            None, send_quota_reached, email, skus_quota,
                            usage_data.get("plan_label", ""),
                            get_upgrade_suggestion(usage_data.get("plan", "")),
                        )
                        _c4 = _get_db()
                        _c4.execute(
                            "UPDATE users SET quota_full_period = ? WHERE email = ?",
                            (period2, email),
                        )
                        _c4.commit()
                        _c4.close()
            except Exception:
                pass

        result = GenerationResult(
            listings=all_listings, failed=all_failed,
            total=n_total, success_count=len(all_listings),
            marketplace=target_mkts[0],
        )
        result_dict = result.model_dump()
        _jobs[job_id].update({"status": "done", "progress": n_total,
                               "result": result_dict})
        log.info("job done", extra={"job_id": job_id, "user": email})
        try:
            update_job_db(job_id, status="done", progress=n_total,
                          result_json=json.dumps(result_dict, default=str))
        except Exception:
            pass
        if email:
            try:
                save_generation(
                    user_email=email,
                    batch_id=job_id,
                    marketplace=str(target_mkts[0].value),
                    listings=result_dict.get("listings", []),
                )
            except Exception:
                pass
            # Notify client integrations (PIM, sites…) — fire-and-forget
            try:
                from app.services.webhooks import fire_and_forget
                fire_and_forget(email, "listing.generated", {
                    "batch_id": job_id,
                    "marketplace": str(target_mkts[0].value),
                    "product_count": len(result_dict.get("listings", [])),
                    "listings": result_dict.get("listings", []),
                })
            except Exception:
                pass
            # Email notification for large batches (>= 5 SKUs)
            if n_total >= 5:
                try:
                    from app.services.email import send_batch_complete
                    plan_label = get_user_usage(email).get("plan_label", "")
                    asyncio.get_event_loop().run_in_executor(
                        None, send_batch_complete, email, n_total, plan_label
                    )
                except Exception:
                    pass
    except Exception as e:
        _jobs[job_id].update({"status": "failed", "error": str(e)})
        log.error("job failed", extra={"job_id": job_id, "user": email})
        try:
            update_job_db(job_id, status="failed", error=str(e))
        except Exception:
            pass


# ── Generation ────────────────────────────────────────────────────────────────

@app.post("/api/generate")
async def generate(request: GenerationRequest, req: Request):
    if not os.getenv("ANTHROPIC_API_KEY"):
        raise HTTPException(503, "ANTHROPIC_API_KEY manquante")
    if not request.products:
        raise HTTPException(400, "Liste de produits vide")

    email = getattr(req.state, "user_email", None)
    if email:
        usage = get_user_usage(email)
        remaining = usage["skus_quota"] - usage["skus_used"]
        if remaining <= 0:
            raise HTTPException(429,
                f"Quota atteint — plan {usage['plan_label']} : {usage['skus_quota']} SKU/mois. "
                "Contactez-nous pour passer à l'offre supérieure.")
        if len(request.products) > remaining:
            raise HTTPException(429,
                f"Quota insuffisant — il vous reste {remaining} SKU ce mois "
                f"(plan {usage['plan_label']} : {usage['skus_quota']}/mois). "
                f"Vous demandez {len(request.products)} SKU.")
        if not _rate_limit(f"{email}:generate", limit=20, window=3600):
            raise HTTPException(429, "Trop de générations — maximum 20/heure. Réessayez dans quelques minutes.")

    n_mkts = len(request.marketplaces) if request.marketplaces else 1
    groups = group_by_parent(request.products)
    n_total = len(groups) * n_mkts
    job_id = str(uuid4())
    _jobs[job_id] = {"status": "pending", "progress": 0,
                     "total": n_total, "created_at": time.time()}
    _cleanup_jobs()
    try:
        save_job(job_id, email, "generation", n_total)
    except Exception:
        pass
    asyncio.create_task(_run_generation_job(job_id, request, email))
    return {"job_id": job_id, "total": n_total}


async def _run_catalog_sync_job(job_id: str, email: str, marketplace_str: str):
    """Background: pull the seller catalog via the SP-API report (slow), persist,
    and expose the rich products. Run as a job so the long report generation
    doesn't hit the HTTP gateway timeout."""
    from app.services.amazon_sp import fetch_seller_catalog
    from app.db import save_catalog_items
    _jobs[job_id]["status"] = "running"
    try:
        try:
            mkt = Marketplace(marketplace_str)
        except ValueError:
            mkt = Marketplace.AMAZON_FR
        items = await fetch_seller_catalog(email, mkt)
        count = save_catalog_items(email, marketplace_str, items)
        products = [{
            "sku": it.get("sku", ""),
            "name": it.get("title", "") or it.get("sku", ""),
            "brand": it.get("brand", "") or "",
            "ean": it.get("ean", "") or "",
            "description": it.get("description", "") or "",
            "features": it.get("bullet_points", []) or [],
            "images": it.get("images", []) or [],
            "asin": it.get("asin", "") or "",
            "price": it.get("price"),
            "status": it.get("status", "") or "",
            "quantity": it.get("quantity"),
        } for it in items]
        active = sum(1 for it in items if (it.get("status") or "") == "active")
        in_stock = sum(1 for it in items if (it.get("quantity") or 0) > 0)
        inactive = len(items) - active
        _jobs[job_id].update({"status": "done",
                              "result": {"synced": count, "marketplace": marketplace_str,
                                         "products": products,
                                         "stats": {"total": len(items), "active": active,
                                                   "inactive": inactive, "in_stock": in_stock}}})
    except RuntimeError as e:
        _jobs[job_id].update({"status": "failed", "error": str(e)})
    except Exception as e:
        _jobs[job_id].update({"status": "failed", "error": f"{type(e).__name__}"})


def start_catalog_sync_job(email: str, marketplace_str: str) -> str:
    job_id = f"catsync_{uuid4().hex[:12]}"
    _jobs[job_id] = {"status": "pending", "created_at": time.time()}
    asyncio.create_task(_run_catalog_sync_job(job_id, email, marketplace_str))
    return job_id


async def _run_catalog_sales_job(job_id: str, email: str):
    """Background: pull Sales & Traffic for every marketplace present in the
    catalog and store per-ASIN metrics, then match to the catalog."""
    import datetime as _d
    from app.services.amazon_sp import fetch_orders_sales
    from app.db import get_catalog_marketplaces, save_catalog_sales, get_catalog_sales_overview
    from app.models import Marketplace
    _jobs[job_id]["status"] = "running"
    try:
        markets = get_catalog_marketplaces(email)
        if not markets:
            _jobs[job_id].update({"status": "done", "result": {"reason": "no_catalog"}})
            return
        end = _d.date.today()
        start = end - _d.timedelta(days=30)
        total_saved = 0
        for mkt_str in markets:
            try:
                mkt = Marketplace(mkt_str)
            except ValueError:
                continue
            # All-Orders report → units + revenue per ASIN (no Brand Analytics)
            metrics = await fetch_orders_sales(
                email, mkt, start.isoformat() + "T00:00:00Z", end.isoformat() + "T00:00:00Z")
            total_saved += save_catalog_sales(email, mkt_str, metrics, start.isoformat(), end.isoformat())
        overview = get_catalog_sales_overview(email)
        _jobs[job_id].update({"status": "done",
                              "result": {"saved": total_saved, "overview": overview}})
    except RuntimeError as e:
        s = str(e)
        if s == "report_pending":
            msg = "Le rapport ventes met trop de temps — relancez dans une minute."
        elif "403" in s or "Unauthorized" in s or "forbidden" in s.lower():
            msg = ("Rôle SP-API « Commandes/Rapports » manquant sur l'app Amazon. "
                   "Ajoutez-le dans Developer Central puis reconnectez le compte, "
                   "ou importez le rapport de ventes en CSV (onglet Suivi).")
        else:
            msg = s
        _jobs[job_id].update({"status": "failed", "error": msg})
    except Exception as e:
        _jobs[job_id].update({"status": "failed", "error": type(e).__name__})


def _parse_business_report_rows(csv_text: str) -> list:
    """Parse an Amazon Business Report 'Ventes et trafic par ASIN' CSV into a list
    of per-row records {sku, asin, units, revenue, sessions}. Order-independent
    column detection, B2B columns skipped, FR/EN accents handled."""
    import csv as _csv, io as _io, unicodedata as _ud
    def _norm(k):
        s = ((k or "").lower().strip().replace("﻿", "")
             .replace("–", "-").replace("—", "-").replace("’", "'").replace("‘", "'")
             .replace("\xa0", " "))
        return "".join(c for c in _ud.normalize("NFKD", s) if not _ud.combining(c))
    reader = _csv.DictReader(_io.StringIO(csv_text))
    def _find(row, includes, excludes=()):
        for k, v in row.items():
            if all(t in k for t in includes) and not any(t in k for t in excludes):
                if v not in (None, "", "-"):
                    return v
        return ""
    def _num(s):
        try:
            return float((s or "").replace("%", "").replace(",", ".").replace(" ", "").strip() or 0)
        except ValueError:
            return 0.0
    def _eur(s):
        s = (s or "").replace("€", "").replace("$", "").replace("\xa0", "").replace(" ", "").strip()
        if "," in s and "." in s:
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", ".")
        try:
            return float(s)
        except ValueError:
            return 0.0
    rows = []
    for raw in reader:
        row = {_norm(k): (v.strip() if isinstance(v, str) else v) for k, v in raw.items()}
        asin = (_find(row, ["asin", "enfant"]) or _find(row, ["asin", "child"])
                or _find(row, ["asin", "parent"]) or _find(row, ["asin"], ["b2b"]) or "").upper()
        sku = (_find(row, ["sku"], ["b2b"]) or "").strip()
        if not asin and not sku:
            continue
        units = int(_num(_find(row, ["unites", "commandees"], ["b2b"])
                         or _find(row, ["units", "ordered"], ["b2b"])))
        revenue = _eur(_find(row, ["ventes", "produits", "commandes"], ["b2b", "nombre", "total de"])
                       or _find(row, ["chiffre", "affaires", "commandes"], ["b2b"])
                       or _find(row, ["ordered", "product", "sales"], ["b2b"]))
        sessions = int(_num(_find(row, ["sessions", "total"], ["b2b", "pourcentage", "vues"])
                            or _find(row, ["sessions"], ["b2b", "pourcentage", "vues"])))
        rows.append({"sku": sku, "asin": asin, "units": units,
                     "revenue": round(revenue, 2), "sessions": sessions})
    return rows


def _parse_business_report_asin_csv(csv_text: str) -> dict:
    """Aggregate the report per child ASIN: {asin: {units_ordered, revenue,
    sessions, conversion_rate}}. Units/revenue summed, sessions maxed (they repeat
    across a child's SKU rows)."""
    out: dict = {}
    for r in _parse_business_report_rows(csv_text):
        asin = r["asin"]
        if not asin or len(asin) < 5:
            continue
        a = out.setdefault(asin, {"units_ordered": 0, "revenue": 0.0, "sessions": 0, "conversion_rate": 0.0})
        a["units_ordered"] += r["units"]
        a["revenue"] = round(a["revenue"] + r["revenue"], 2)
        a["sessions"] = max(a["sessions"], r["sessions"])
    for a in out.values():
        a["conversion_rate"] = round(a["units_ordered"] / a["sessions"] * 100, 2) if a["sessions"] else 0.0
    return out




@app.post("/api/catalog/sales-import-csv")
async def catalog_sales_import_csv(request: Request):
    """Import a Sales & Traffic report CSV and cross-reference it with the catalog.
    Matches primarily by SKU (both sides carry it reliably), falling back to ASIN,
    and backfills the catalog's child ASIN from the report so the sales tiles line
    up even when the sync stored a wrong/empty ASIN. Needs no SP-API sales role."""
    from app.db import (get_catalog_summary, get_catalog, save_catalog_sales,
                        get_catalog_sales_overview, backfill_catalog_asins)
    body = await request.json()
    csv_text = body.get("csv_text", "")
    if not csv_text:
        raise HTTPException(400, "CSV vide")
    email = request.state.user_email
    rows = _parse_business_report_rows(csv_text)
    if not rows:
        raise HTTPException(400, "Aucune donnée de vente détectée — utilisez le rapport « Ventes et trafic par ASIN ».")

    # Aggregate the report by SKU and by ASIN (units/revenue summed, sessions maxed).
    def _agg(store, key, r):
        a = store.setdefault(key, {"units_ordered": 0, "revenue": 0.0, "sessions": 0,
                                   "conversion_rate": 0.0, "asin": r["asin"]})
        a["units_ordered"] += r["units"]
        a["revenue"] = round(a["revenue"] + r["revenue"], 2)
        a["sessions"] = max(a["sessions"], r["sessions"])
        if r["asin"]:
            a["asin"] = r["asin"]
    by_sku: dict = {}
    by_asin: dict = {}
    asin_count = set()
    for r in rows:
        if r["sku"]:
            _agg(by_sku, r["sku"].upper(), r)
        if r["asin"]:
            _agg(by_asin, r["asin"].upper(), r)
            asin_count.add(r["asin"].upper())

    import datetime as _d
    end = _d.date.today().isoformat()
    start = (_d.date.today() - _d.timedelta(days=30)).isoformat()

    by_mkt: dict = {}           # marketplace -> {asin: metrics} to persist
    backfill: dict = {}         # marketplace -> {sku: asin} to repair catalog
    matched = 0
    for mkt in (get_catalog_summary(email) or {}):
        for it in get_catalog(email, mkt):
            sku = (it.get("sku") or "").strip()
            catalog_asin = (it.get("asin") or "").upper()
            m = by_sku.get(sku.upper()) if sku else None       # reliable: match by SKU
            if not m and catalog_asin:
                m = by_asin.get(catalog_asin)                  # fallback: match by ASIN
            if not m:
                continue
            store_asin = catalog_asin or (m.get("asin") or "").upper()
            if not store_asin:
                continue
            metrics = {k: m[k] for k in ("units_ordered", "revenue", "sessions", "conversion_rate")}
            metrics["conversion_rate"] = round(metrics["units_ordered"] / metrics["sessions"] * 100, 2) \
                if metrics["sessions"] else 0.0
            by_mkt.setdefault(mkt, {})[store_asin] = metrics
            if not catalog_asin and sku:                       # sync had no ASIN → repair it
                backfill.setdefault(mkt, {})[sku] = store_asin
            matched += 1

    for mkt, mm in by_mkt.items():
        save_catalog_sales(email, mkt, mm, start, end)
    repaired = 0
    for mkt, updates in backfill.items():
        repaired += backfill_catalog_asins(email, mkt, updates)

    return {"imported": matched, "total_in_csv": len(asin_count),
            "repaired_asins": repaired,
            "overview": get_catalog_sales_overview(email)}


@app.get("/api/catalog/export.csv")
async def catalog_export_csv(request: Request):
    """Download the full catalogue health as CSV: fiche, status, price, stock, sales."""
    import csv as _csv, io as _io
    from fastapi.responses import StreamingResponse
    from app.db import get_catalog_full_export
    rows = get_catalog_full_export(request.state.user_email)
    buf = _io.StringIO()
    buf.write("﻿")  # BOM so Excel opens accents/€ correctly
    cols = ["marketplace", "sku", "asin", "ean", "title", "status",
            "quantity", "price", "units_ordered", "revenue"]
    headers = {"marketplace": "Marketplace", "sku": "SKU", "asin": "ASIN", "ean": "EAN",
               "title": "Titre", "status": "Statut", "quantity": "Stock", "price": "Prix",
               "units_ordered": "Unités vendues", "revenue": "CA"}
    w = _csv.DictWriter(buf, fieldnames=cols, delimiter=";")
    w.writerow(headers)
    for r in rows:
        w.writerow({k: r.get(k, "") for k in cols})
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]), media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="sante-catalogue-synqio.csv"'})


@app.post("/api/catalog/sales-sync")
async def catalog_sales_sync(request: Request):
    """Kick off the catalog-wide Sales & Traffic pull (background job)."""
    job_id = f"catsales_{uuid4().hex[:12]}"
    _jobs[job_id] = {"status": "pending", "created_at": time.time()}
    asyncio.create_task(_run_catalog_sales_job(job_id, request.state.user_email))
    return {"job_id": job_id}


@app.get("/api/jobs/{job_id}")
async def get_job(job_id: str):
    job = _jobs.get(job_id)
    if not job:
        try:
            db_jobs = load_recent_jobs(limit=500)
            job = db_jobs.get(job_id)
        except Exception:
            pass
    if not job:
        raise HTTPException(404, "Job introuvable ou expiré (1h max)")
    return job


# ── Publish ───────────────────────────────────────────────────────────────────

def _log_amazon_pushes(email: str, result: "PublishResult", marketplace) -> None:
    """Journal successfully published fiches (never breaks the publish flow)."""
    try:
        from app.db import log_pushes
        skus = [r.get("sku") for r in (result.report or []) if r.get("sku")]
        log_pushes(email, "amazon", skus, status="published",
                   marketplace=getattr(marketplace, "value", str(marketplace)))
    except Exception:
        pass


@app.post("/api/publish", response_model=PublishResult)
async def publish(request_data: PublishRequest, request: Request):
    if not request_data.listings:
        raise HTTPException(400, "Aucune fiche à publier")
    result = await publish_listings(
        listings=request_data.listings,
        marketplace=request_data.marketplace,
        dry_run=request_data.dry_run,
        user_email=getattr(request.state, "user_email", None),
    )
    if not request_data.dry_run:
        _log_amazon_pushes(request.state.user_email, result, request_data.marketplace)
    return result


async def _run_bulk_publish_job(job_id: str, request_data: PublishRequest, email: str):
    """Bulk publish via the Amazon Feeds API (1 upload for N fiches), with an
    automatic per-SKU fallback while the 'Product Listing' role isn't granted."""
    from app.services.amazon_sp import publish_via_feed
    _jobs[job_id]["status"] = "running"
    try:
        try:
            res = await publish_via_feed(
                request_data.listings, request_data.marketplace, user_email=email)
            mode = "feed"
        except RuntimeError as e:
            s = str(e)
            if s == "feeds_forbidden":
                # Role missing → per-SKU path still works for accounts that have it;
                # surface the reason either way so the user knows what happened.
                res = await publish_listings(
                    listings=request_data.listings, marketplace=request_data.marketplace,
                    dry_run=False, user_email=email)
                mode = "per_sku_fallback"
            elif s == "feed_pending":
                _jobs[job_id].update({"status": "failed", "error":
                    "Amazon met plus de 10 min à traiter le feed — réessayez dans quelques minutes "
                    "(le traitement continue côté Amazon)."})
                return
            else:
                raise
        _log_amazon_pushes(email, res, request_data.marketplace)
        _jobs[job_id].update({"status": "done", "result": {
            "mode": mode, **res.model_dump(),
            **({"note": "Rôle « Product Listing » manquant pour la Feeds API — "
                        "publication fiche par fiche utilisée à la place."}
               if mode == "per_sku_fallback" else {}),
        }})
    except Exception as e:
        _jobs[job_id].update({"status": "failed", "error": str(e)[:300] or type(e).__name__})


@app.post("/api/publish/bulk")
async def publish_bulk(request_data: PublishRequest, request: Request):
    """Publish a large batch through the Feeds API as a background job.
    Poll GET /api/jobs/{job_id} for the per-SKU report."""
    if not request_data.listings:
        raise HTTPException(400, "Aucune fiche à publier")
    if request_data.dry_run:
        return await publish_listings(
            listings=request_data.listings, marketplace=request_data.marketplace,
            dry_run=True, user_email=getattr(request.state, "user_email", None))
    job_id = f"pubfeed_{uuid4().hex[:12]}"
    _jobs[job_id] = {"status": "pending", "created_at": time.time()}
    asyncio.create_task(_run_bulk_publish_job(job_id, request_data, request.state.user_email))
    return {"job_id": job_id, "count": len(request_data.listings)}


from app.utils.amazon_template_filler import fill_amazon_template
# ── Export ────────────────────────────────────────────────────────────────────

@app.post("/api/export/csv")
async def export_csv(listings: List[AmazonListing]):
    return Response(content=to_csv_bytes(listings), media_type="text/csv",
                    headers={"Content-Disposition": "attachment; filename=listings.csv"})


@app.post("/api/export/json")
async def export_json_file(listings: List[AmazonListing]):
    return Response(content=to_json_bytes(listings), media_type="application/json",
                    headers={"Content-Disposition": "attachment; filename=listings.json"})


@app.post("/api/export/fill-amazon-template")
async def fill_amazon_template_endpoint(request: Request):
    """Accept multipart: amazon_template (file) + listings (JSON string)."""
    form = await request.form()
    template_file = form.get("amazon_template")
    listings_json = form.get("listings")
    if not template_file or not listings_json:
        raise HTTPException(400, "Fichier template et listings requis")
    template_bytes = await template_file.read()
    import json as _json
    raw = _json.loads(listings_json)
    listings = [AmazonListing(**item) for item in raw]
    compliance_raw = form.get("compliance_data")
    compliance_data = _json.loads(compliance_raw) if compliance_raw else None
    try:
        image_urls = _get_image_urls_for_skus([l.sku for l in listings])
        filled_bytes = fill_amazon_template(template_bytes, listings, image_urls=image_urls,
                                            compliance_data=compliance_data, for_new_products=True)
    except ValueError as e:
        # Offer template used for new products, or other content-level problem.
        raise HTTPException(422, str(e))
    except Exception as e:
        raise HTTPException(422, f"Erreur lors du remplissage : {e}")
    # Auto-save to library if this product type isn't there yet
    try:
        from app.utils.amazon_template_filler import get_product_type_from_bytes, is_offer_only_template
        pt = get_product_type_from_bytes(template_bytes)
        # "PRODUCT" is the generic Listing Loader (offer) template — not a real
        # category, don't pollute the category-template library with it. Also skip
        # any offer-only template (no product columns) regardless of its type code.
        if pt and pt != "PRODUCT" and not is_offer_only_template(template_bytes) and not amazon_template_exists(pt):
            fname = getattr(template_file, "filename", None) or f"{pt}.xlsm"
            data_b64 = base64.b64encode(template_bytes).decode()
            label = pt.replace("_", " ").title()
            save_amazon_template(str(uuid4()), label, pt, fname, data_b64,
                                 getattr(request.state, "user_email", ""))
            log.info(f"[Templates] Auto-saved from upload: {pt}")
    except Exception:
        pass
    fn = getattr(template_file, "filename", None) or "amazon_template_rempli.xlsm"
    return Response(
        content=filled_bytes,
        media_type="application/vnd.ms-excel.sheet.macroEnabled.12",
        headers={"Content-Disposition": f"attachment; filename={fn}"},
    )


def _get_image_urls_for_skus(skus: list) -> dict:
    """Look up which generated images exist in storage for each SKU and return their
    public URLs. When a SKU has no generated 'main', its OWN uploaded photo
    ({sku}.jpg/png/webp) fills the main slot — so '100% ma photo' flows into
    exports, feeds and the public catalogue without any generation."""
    if not skus:
        return {}
    from app.services import storage as _storage
    from app.services.image_gen import AMAZON_IMAGE_TYPES
    base = os.getenv("PUBLIC_BASE_URL", "https://synqio.io").rstrip("/")
    # Fetch all keys once to avoid N×7 round trips
    every_key = _storage.list_keys()
    all_keys = {k for k in every_key if k.startswith("gen_")}
    own_by_sku = {}
    for k in every_key:
        if k.startswith("gen_"):
            continue
        stem, ext = os.path.splitext(k)
        if ext.lower() in _IMAGE_EXTS:
            own_by_sku.setdefault(stem, k)
    def _url(filename):
        cdn = _storage.public_url(filename)
        return cdn if cdn else f"{base}/api/photos/{filename}"
    result: dict = {}
    for sku in skus:
        imgs: dict = {}
        for img_type in AMAZON_IMAGE_TYPES:
            img_id = img_type["id"]
            filename = f"gen_{sku}_{img_id}.png"
            if filename in all_keys:
                imgs[img_id] = _url(filename)
        if "hero" not in imgs and sku in own_by_sku:
            imgs["hero"] = _url(own_by_sku[sku])
        if imgs:
            result[sku] = imgs
    return result


# ── Amazon template library ───────────────────────────────────────────────────

@app.get("/api/templates")
async def api_list_templates(request: Request):
    return JSONResponse(list_amazon_templates())


@app.get("/api/templates/attributes")
async def api_template_attributes(request: Request, product_type: str = "", tpl_id: str = ""):
    """Dynamic per-category attribute list (label + accepted values + level) for
    the compliance form. Resolved from the stored Amazon template of the given
    category (product_type) or template id."""
    from app.utils.amazon_template_filler import extract_template_attributes
    tpl = None
    if tpl_id:
        tpl = get_amazon_template(tpl_id)
    if not tpl and product_type:
        tpl = get_amazon_template_by_type(product_type.strip().upper())
    if not tpl:
        return JSONResponse({"product_type": product_type.upper(), "attributes": []})
    try:
        data = base64.b64decode(tpl["data_b64"])
        attrs = extract_template_attributes(data)
    except Exception as e:
        log.error(f"[templates] attribute extraction failed: {e}")
        attrs = []
    return JSONResponse({"product_type": tpl.get("product_type", ""), "attributes": attrs})


@app.post("/api/admin/templates")
async def api_upload_template(request: Request):
    if not getattr(request.state, "is_admin", False):
        raise HTTPException(403, "Réservé aux administrateurs")
    form = await request.form()
    template_file = form.get("template")
    label = (form.get("label") or "").strip()
    product_type = (form.get("product_type") or "").strip().upper()
    if not template_file or not label or not product_type:
        raise HTTPException(400, "template, label et product_type requis")
    data = await template_file.read()
    # Reject offer/Listing-Loader templates here: this library is for category
    # templates that CREATE products. An offer template only updates existing
    # catalogue products and would later fail with Amazon errors 8560/13013.
    from app.utils.amazon_template_filler import is_offer_only_template, OFFER_TEMPLATE_FOR_NEW_PRODUCTS_MSG
    if is_offer_only_template(data):
        raise HTTPException(422, OFFER_TEMPLATE_FOR_NEW_PRODUCTS_MSG)
    data_b64 = base64.b64encode(data).decode()
    tpl_id = str(uuid4())
    save_amazon_template(tpl_id, label, product_type,
                         getattr(template_file, "filename", f"{product_type}.xlsm"),
                         data_b64,
                         getattr(request.state, "user_email", "admin"))
    return JSONResponse({"id": tpl_id, "label": label, "product_type": product_type})


@app.delete("/api/admin/templates/{tpl_id}")
async def api_delete_template(tpl_id: str, request: Request):
    if not getattr(request.state, "is_admin", False):
        raise HTTPException(403, "Réservé aux administrateurs")
    ok = delete_amazon_template(tpl_id)
    if not ok:
        raise HTTPException(404, "Template introuvable")
    return JSONResponse({"ok": True})


@app.post("/api/export/fill-from-library/{tpl_id}")
async def fill_from_library(tpl_id: str, request: Request):
    """Fill a stored template with the supplied listings and return the filled bytes."""
    body = await request.json()
    raw = body.get("listings", [])
    if not raw:
        raise HTTPException(400, "Listings requis")
    listings = [AmazonListing(**item) for item in raw]
    compliance_data = body.get("compliance_data") or None
    tpl = get_amazon_template(tpl_id)
    if not tpl:
        raise HTTPException(404, "Template introuvable")
    template_bytes = base64.b64decode(tpl["data_b64"])
    try:
        image_urls = _get_image_urls_for_skus([l.sku for l in listings])
        filled_bytes = fill_amazon_template(template_bytes, listings, image_urls=image_urls,
                                            compliance_data=compliance_data, for_new_products=True)
    except ValueError as e:
        # Offer template stored as a category template, or other content problem.
        raise HTTPException(422, str(e))
    except Exception as e:
        raise HTTPException(422, f"Erreur lors du remplissage : {e}")
    fn = tpl.get("filename") or "amazon_template_rempli.xlsm"
    return Response(
        content=filled_bytes,
        media_type="application/vnd.ms-excel.sheet.macroEnabled.12",
        headers={"Content-Disposition": f"attachment; filename={fn}"},
    )


@app.post("/api/export/flat-file")
async def export_flat_file(listings: List[AmazonListing], request: Request):
    skus = [l.sku for l in listings]
    image_urls = _get_image_urls_for_skus(skus)
    return Response(
        content=to_amazon_flat_file_xlsx(listings, image_urls=image_urls),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=amazon_nouveaux_produits.xlsx"},
    )


@app.post("/api/export/listing-loader")
async def export_listing_loader(listings: List[AmazonListing]):
    return Response(
        content=to_listing_loader_xlsx(listings),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=amazon_produits_existants.xlsx"},
    )


@app.post("/api/export/channel/{channel}")
async def export_channel(channel: str, listings: List[AmazonListing], request: Request):
    """Multichannel export — PIM / Shopify / WooCommerce / PrestaShop / Akeneo /
    Google Merchant. Generated image URLs are resolved from storage per SKU."""
    from app.utils.export_channels import CHANNELS
    spec = CHANNELS.get(channel)
    if not spec:
        raise HTTPException(404, f"Canal inconnu : {channel} — disponibles : {', '.join(CHANNELS)}")
    if not listings:
        raise HTTPException(400, "Aucune fiche à exporter")
    image_urls = _get_image_urls_for_skus([l.sku for l in listings])
    try:
        from app.db import log_pushes
        log_pushes(request.state.user_email, channel, [l.sku for l in listings], status="export")
    except Exception:
        pass  # journaling must never break an export
    return Response(
        content=spec["builder"](listings, image_urls=image_urls),
        media_type=spec["media_type"],
        headers={"Content-Disposition": f'attachment; filename="{spec["filename"]}"'},
    )


# ── Médiathèque — every image (AI-generated + own uploads), per fiche ────────

def _user_skus(email: str) -> set:
    """All SKUs belonging to this account: generated listings + synced catalog."""
    skus = set()
    try:
        from app.routes.public_api import _latest_listings
        skus |= {l.get("sku") for l in _latest_listings(email) if l.get("sku")}
    except Exception:
        pass
    try:
        from app.db import get_catalog_summary, get_catalog
        for mkt in (get_catalog_summary(email) or {}):
            skus |= {i.get("sku") for i in get_catalog(email, mkt) if i.get("sku")}
    except Exception:
        pass
    return skus


@app.get("/api/media/library")
async def media_library(request: Request):
    """Media bank grouped by SKU: AI-generated images (by slot) + own uploads."""
    from app.services import storage as _storage
    from app.services.image_gen import AMAZON_IMAGE_TYPES
    email = request.state.user_email
    skus = _user_skus(email)
    base = os.getenv("PUBLIC_BASE_URL", "https://synqio.io").rstrip("/")
    def _url(fn):
        cdn = _storage.public_url(fn)
        return cdn if cdn else f"{base}/api/photos/{fn}"
    type_ids = [t["id"] for t in AMAZON_IMAGE_TYPES]
    items: dict = {}
    for key in _storage.list_keys():
        stem, ext = os.path.splitext(key)
        if ext.lower() not in _IMAGE_EXTS:
            continue
        if key.startswith("gen_"):
            # gen_{sku}_{type}.png — type is the last _-segment matching a known slot
            body = stem[4:]
            for tid in type_ids:
                if body.endswith("_" + tid):
                    sku = body[: -len(tid) - 1]
                    if sku in skus:
                        items.setdefault(sku, {"sku": sku, "generated": {}, "own": []})
                        items[sku]["generated"][tid] = {"url": _url(key), "filename": key}
                    break
        else:
            if stem in skus:
                items.setdefault(stem, {"sku": stem, "generated": {}, "own": []})
                items[stem]["own"].append({"url": _url(key), "filename": key})
    out = sorted(items.values(), key=lambda i: i["sku"])
    n_gen = sum(len(i["generated"]) for i in out)
    n_own = sum(len(i["own"]) for i in out)
    return {"items": out, "skus": len(out), "generated": n_gen, "own": n_own}


@app.post("/api/media/upload")
async def media_upload(request: Request, sku: str = ""):
    """Upload one photo for a SKU (query param). Stored as {sku}.{ext} — the same
    convention as the ZIP/Drive imports, so it feeds generation and exports."""
    sku = (sku or "").strip()
    if not sku:
        raise HTTPException(400, "Paramètre sku requis")
    if "/" in sku or "\\" in sku or sku.startswith("gen_"):
        raise HTTPException(400, "SKU invalide")
    content = await request.body()
    if not content:
        raise HTTPException(400, "Fichier vide")
    if len(content) > 15 * 1024 * 1024:
        raise HTTPException(413, "Image trop volumineuse (max 15 Mo)")
    ctype = (request.headers.get("Content-Type") or "").lower()
    ext = {"image/jpeg": ".jpg", "image/png": ".png", "image/webp": ".webp"}.get(ctype)
    if not ext:  # sniff magic bytes as fallback
        if content[:3] == b"\xff\xd8\xff":
            ext = ".jpg"
        elif content[:8] == b"\x89PNG\r\n\x1a\n":
            ext = ".png"
        elif content[:4] == b"RIFF" and content[8:12] == b"WEBP":
            ext = ".webp"
        else:
            raise HTTPException(415, "Format non supporté (JPEG, PNG ou WebP)")
    filename = f"{sku}{ext}"
    _save_photo(filename, content, {"jpg": "image/jpeg", "png": "image/png",
                                    "webp": "image/webp"}[ext[1:]])
    return {"sku": sku, "filename": filename, "url": f"/api/photos/{filename}"}


@app.delete("/api/media/{filename}")
async def media_delete(filename: str, request: Request):
    """Delete one image — only if its SKU belongs to the requesting account."""
    from app.services import storage as _storage
    from app.services.image_gen import AMAZON_IMAGE_TYPES
    stem, ext = os.path.splitext(filename)
    if ext.lower() not in _IMAGE_EXTS:
        raise HTTPException(400, "Fichier non-image")
    sku = stem
    if filename.startswith("gen_"):
        body = stem[4:]
        sku = next((body[: -len(t["id"]) - 1] for t in AMAZON_IMAGE_TYPES
                    if body.endswith("_" + t["id"])), body)
    if sku not in _user_skus(request.state.user_email):
        raise HTTPException(403, "Cette image n'appartient pas à votre compte")
    if not _storage.delete(filename):
        raise HTTPException(404, "Image introuvable")
    return {"deleted": filename}


@app.get("/api/content/library")
async def content_library(request: Request):
    """Text bank: the latest generated content per fiche (title, bullets,
    description, keywords) — everything copyable/reusable."""
    from app.routes.public_api import _latest_listings
    items = []
    for l in _latest_listings(request.state.user_email):
        items.append({
            "sku": l.get("sku") or "", "marketplace": l.get("marketplace") or "",
            "title": l.get("title") or "", "brand": l.get("brand") or "",
            "bullet_points": l.get("bullet_points") or [],
            "description": l.get("description") or "",
            "backend_keywords": l.get("backend_keywords") or "",
            "price": l.get("price"),
        })
    return {"items": items, "count": len(items)}


# ── Hub de diffusion — where every fiche lives ───────────────────────────────

@app.get("/api/hub/overview")
async def hub_overview(request: Request):
    """Diffusion dashboard: per-channel counts, feed stats and the public
    catalogue link stats — the 'where are my fiches' view."""
    from app.db import get_push_overview, get_feed_token, get_or_create_catalog_link
    email = request.state.user_email
    out = {"channels": get_push_overview(email)}
    try:
        ft = get_feed_token(email)
        if ft:
            out["feeds"] = {"fetch_count": ft.get("fetch_count") or 0,
                            "last_channel": ft.get("last_channel") or "",
                            "last_fetched_at": ft.get("last_fetched_at") or ""}
    except Exception:
        pass
    try:
        cl = get_or_create_catalog_link(email)
        base = os.getenv("APP_URL", "https://synqio.io").rstrip("/")
        out["catalog"] = {"url": f"{base}/catalogue/{cl['token']}",
                          "view_count": cl.get("view_count") or 0,
                          "last_viewed_at": cl.get("last_viewed_at") or "",
                          "title": cl.get("title") or "", "subtitle": cl.get("subtitle") or "",
                          "contact_email": cl.get("contact_email") or "",
                          "website": cl.get("website") or "", "phone": cl.get("phone") or "",
                          "logo_url": cl.get("logo_url") or ""}
    except Exception:
        pass
    return out


@app.get("/api/hub/pushes/{channel}")
async def hub_pushes(channel: str, request: Request):
    from app.db import get_pushes_detail
    return {"channel": channel,
            "items": get_pushes_detail(request.state.user_email, channel, limit=200)}


class CatalogMeta(BaseModel):
    title: str = ""
    subtitle: str = ""
    contact_email: str = ""
    website: str = ""
    phone: str = ""


@app.post("/api/catalog-link/logo")
async def catalog_link_logo(request: Request):
    """Upload the brand logo shown on the public catalogue (PNG/JPEG/WebP)."""
    import hashlib
    from app.db import set_catalog_link_meta, get_or_create_catalog_link
    content = await request.body()
    if not content:
        raise HTTPException(400, "Fichier vide")
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(413, "Logo trop volumineux (max 5 Mo)")
    if content[:8] == b"\x89PNG\r\n\x1a\n":
        ext, ct = ".png", "image/png"
    elif content[:3] == b"\xff\xd8\xff":
        ext, ct = ".jpg", "image/jpeg"
    elif content[:4] == b"RIFF" and content[8:12] == b"WEBP":
        ext, ct = ".webp", "image/webp"
    else:
        raise HTTPException(415, "Format non supporté (PNG, JPEG ou WebP)")
    email = request.state.user_email
    filename = f"catlogo_{hashlib.md5(email.encode()).hexdigest()[:16]}{ext}"
    _save_photo(filename, content, ct)
    cl = get_or_create_catalog_link(email)
    set_catalog_link_meta(email, cl.get("title") or "", cl.get("subtitle") or "",
                          cl.get("contact_email") or "", cl.get("website") or "",
                          cl.get("phone") or "", logo_url=f"/api/photos/{filename}")
    return {"logo_url": f"/api/photos/{filename}"}


class SiteImportRequest(BaseModel):
    url: str


@app.post("/api/catalog-from-site")
async def catalog_from_site(req: SiteImportRequest, request: Request):
    """Import a client's products straight from their website URL (Shopify's
    public products.json, or WooCommerce's public Store API) and build their
    catalogue: listings saved as a batch, first image stored per SKU so the
    flipbook shows real photos."""
    import re as _re
    email = request.state.user_email
    if not _rate_limit(f"{email}:site-import", limit=6, window=3600):
        raise HTTPException(429, "Limite de 6 imports/heure — réessayez plus tard")
    base = req.url.strip()
    if not base.startswith("http"):
        base = "https://" + base
    base = base.rstrip("/")
    strip = lambda h: _re.sub(r"\s+", " ", _re.sub(r"<[^>]+>", " ", h or "")).strip()
    products = []
    async with httpx.AsyncClient(timeout=25, follow_redirects=True,
                                 headers={"User-Agent": "SynqIO-CatalogBot/1.0"}) as client:
        # 1. Shopify public catalogue endpoint
        try:
            r = await client.get(f"{base}/products.json?limit=250")
            data = r.json() if r.status_code == 200 else {}
            for p in (data.get("products") or [])[:120]:
                v = (p.get("variants") or [{}])[0]
                sku = _re.sub(r"[^A-Za-z0-9_.-]", "-", (v.get("sku") or p.get("handle") or ""))[:60]
                if not sku:
                    continue
                products.append({
                    "sku": sku, "marketplace": "amazon_fr",
                    "title": strip(p.get("title") or "")[:200],
                    "brand": (p.get("vendor") or "")[:80],
                    "category": (p.get("product_type") or "")[:80],
                    "description": strip(p.get("body_html") or "")[:2000],
                    "bullet_points": [], "backend_keywords": "",
                    "price": float(v["price"]) if v.get("price") else None,
                    "_img": ((p.get("images") or [{}])[0].get("src") or ""),
                })
        except Exception:
            pass
        # 2. WooCommerce public Store API
        if not products:
            try:
                r = await client.get(f"{base}/wp-json/wc/store/v1/products?per_page=100")
                arr = r.json() if r.status_code == 200 else []
                for p in (arr or [])[:120]:
                    sku = _re.sub(r"[^A-Za-z0-9_.-]", "-", (p.get("sku") or p.get("slug") or ""))[:60]
                    if not sku:
                        continue
                    pr = (p.get("prices") or {})
                    minor = int(pr.get("currency_minor_unit") or 2)
                    price = int(pr["price"]) / (10 ** minor) if pr.get("price") else None
                    cats = p.get("categories") or []
                    imgs = p.get("images") or []
                    products.append({
                        "sku": sku, "marketplace": "amazon_fr",
                        "title": strip(p.get("name") or "")[:200], "brand": "",
                        "category": (cats[0].get("name") if cats else "")[:80],
                        "description": strip(p.get("description") or p.get("short_description") or "")[:2000],
                        "bullet_points": [], "backend_keywords": "", "price": price,
                        "_img": (imgs[0].get("src") if imgs else ""),
                    })
            except Exception:
                pass
        if not products:
            raise HTTPException(422, "Impossible de lire les produits de ce site — Shopify et "
                                     "WooCommerce sont supportés (le catalogue public doit être accessible).")
        # 3. Store the first photo per SKU (own-photo convention → hero du flipbook)
        photos = 0
        for p in products[:60]:
            u = p.pop("_img", "")
            if not u:
                continue
            try:
                ir = await client.get(u)
                if ir.status_code == 200 and len(ir.content) < 8 * 1024 * 1024:
                    ext = ".png" if ir.content[:8] == b"\x89PNG\r\n\x1a\n" else ".jpg"
                    _save_photo(f"{p['sku']}{ext}", ir.content,
                                "image/png" if ext == ".png" else "image/jpeg")
                    photos += 1
            except Exception:
                continue
        for p in products:
            p.pop("_img", None)
    from app.db import save_generation, get_or_create_catalog_link
    batch_id = f"site_{uuid4().hex[:10]}"
    save_generation(email, batch_id, "amazon_fr", products, "Import site web")
    cl = get_or_create_catalog_link(email)
    base_url = os.getenv("APP_URL", "https://synqio.io").rstrip("/")
    return {"imported": len(products), "photos": photos,
            "catalog_url": f"{base_url}/catalogue/{cl['token']}"}


@app.post("/api/catalog-link/regenerate")
async def catalog_link_regenerate(request: Request):
    from app.db import regenerate_catalog_link
    cl = regenerate_catalog_link(request.state.user_email)
    base = os.getenv("APP_URL", "https://synqio.io").rstrip("/")
    return {"url": f"{base}/catalogue/{cl['token']}"}


@app.put("/api/catalog-link")
async def catalog_link_meta(meta: CatalogMeta, request: Request):
    from app.db import set_catalog_link_meta
    set_catalog_link_meta(request.state.user_email, meta.title.strip(), meta.subtitle.strip(),
                          meta.contact_email.strip(), meta.website.strip(), meta.phone.strip())
    return {"ok": True}


def _render_public_catalog(owner: dict, listings: list, image_urls: dict) -> str:
    """Calaméo-style public flipbook: cover page, 4 products per page with 3D
    page-turn, photo lightbox (all visuals per product), contact back page.
    Embeddable in an <iframe> on any website."""
    import html as _html
    e = _html.escape
    brands = sorted({(l.get("brand") or "").strip() for l in listings if (l.get("brand") or "").strip()})
    title = owner.get("title") or (f"Catalogue {brands[0]}" if len(brands) == 1 else "Notre catalogue")
    import datetime as _d
    subtitle = owner.get("subtitle") or f"Collection {_d.date.today().year} — {len(listings)} produits"
    groups: dict = {}
    for l in listings:
        groups.setdefault((l.get("category") or "Autres").strip() or "Autres", []).append(l)
    cur = {"amazon_uk": "£", "amazon_pl": "zł", "amazon_se": "kr"}

    def product_card(l):
        imgs = (image_urls or {}).get(l.get("sku"), []) or []
        img = imgs[0] if imgs else ""
        price = l.get("price")
        sym = cur.get(l.get("marketplace") or "", "€")
        imgs_attr = e(json.dumps(imgs)) if imgs else ""
        pic = (f'<img loading="lazy" src="{e(img)}" alt="" class="zoom" data-imgs="{imgs_attr}" '
               f'data-title="{e(l.get("title") or "")}">' if img
               else '<div class="noimg">Photo à venir</div>')
        return (f'<div class="prd">{pic}<h3>{e(l.get("title") or l.get("sku") or "")}</h3>'
                + (f'<div class="price">{price:.2f} {sym}</div>' if isinstance(price, (int, float)) else '')
                + '</div>')

    logo = f'<img class="logo" src="{e(owner.get("logo_url") or "")}" alt="">' if owner.get("logo_url") else ''
    pages = [f'<div class="page cover">{logo}<h1>{e(title)}</h1><p>{e(subtitle)}</p>'
             f'<span class="hintturn">Cliquez ou utilisez les flèches pour tourner les pages</span></div>']
    for cat, items in groups.items():
        for i in range(0, len(items), 4):
            chunk = items[i:i + 4]
            pages.append(f'<div class="page"><div class="cat">{e(cat)}</div>'
                         f'<div class="pgrid">{"".join(product_card(l) for l in chunk)}</div></div>')
    pages.append(f'<div class="page cover back">{logo}<h1>Merci</h1>{_catalog_contact_html(owner)}'
                 '<p class="credit">Catalogue généré avec SynqIO</p></div>')
    # paper grain overlay on every page
    pages = [p[:-6] + '<i class="gr"></i></div>' for p in pages]
    n = len(pages)
    if n % 2:
        pages.append('<div class="page blank"></div>')
    sheets = ""
    S = len(pages) // 2
    for k in range(S):
        sheets += (f'<div class="sheet" data-s="{k}">'
                   f'<div class="pg f">{pages[2*k]}</div>'
                   f'<div class="pg b">{pages[2*k+1]}</div></div>')

    return f'''<!DOCTYPE html><html lang="fr"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<meta name="robots" content="noindex">
<title>{e(title)}</title>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:'Segoe UI',system-ui,sans-serif;background:linear-gradient(125deg,#171236,#2a1f5c);min-height:100vh;display:flex;flex-direction:column;align-items:center;justify-content:center;padding:18px}}
.stage{{position:relative;width:min(1160px,96vw);aspect-ratio:3/2;max-height:82vh;perspective:3200px}}
.stage.single{{width:min(560px,96vw);aspect-ratio:3/4;max-height:84vh;perspective:none}}
.spine{{position:absolute;left:50%;top:2%;bottom:2%;width:0;box-shadow:0 0 34px 7px rgba(0,0,0,.4);z-index:1}}
.single .spine{{display:none}}
.sheet{{position:absolute;left:50%;top:0;height:100%;width:50%;transform-origin:left center;transform-style:preserve-3d;transition:transform 1.1s cubic-bezier(.35,.06,.15,1)}}
.sheet.flip{{transform:rotateY(-180deg)}}
.sheet.moving{{z-index:99!important}}
.pg{{position:absolute;inset:0;backface-visibility:hidden}}
.pg.b{{transform:rotateY(180deg)}}
.page{{position:absolute;inset:0;background:#fff;border-radius:4px 14px 14px 4px;box-shadow:0 16px 50px rgba(0,0,0,.4);overflow:hidden;display:flex;flex-direction:column;padding:24px 26px}}
.pg.b .page{{border-radius:14px 4px 4px 14px}}
.page::before{{content:'';position:absolute;left:0;top:0;bottom:0;width:16px;background:linear-gradient(90deg,rgba(0,0,0,.12),transparent);pointer-events:none;z-index:2}}
.pg.b .page::before{{left:auto;right:0;background:linear-gradient(-90deg,rgba(0,0,0,.12),transparent)}}
.gr{{position:absolute;inset:0;pointer-events:none;opacity:.4;mix-blend-mode:overlay;z-index:1;background-image:url("data:image/svg+xml;utf8,<svg xmlns='http://www.w3.org/2000/svg' width='140' height='140'><filter id='n'><feTurbulence type='fractalNoise' baseFrequency='0.85' numOctaves='2' stitchTiles='stitch'/><feColorMatrix type='saturate' values='0'/></filter><rect width='140' height='140' filter='url(%23n)' opacity='0.45'/></svg>")}}
.sheet.moving .page{{box-shadow:-30px 22px 60px rgba(0,0,0,.5)}}
.turn{{position:absolute;left:50%;top:0;height:100%;width:50%;z-index:120;pointer-events:none;transform-style:preserve-3d}}
.strip{{position:absolute;inset:0;transform-origin:left center;transform-style:preserve-3d;will-change:transform}}
.strip .pg{{position:absolute;inset:0;backface-visibility:hidden}}
@keyframes flexf{{0%{{transform:perspective(2200px) rotateY(0)}}30%{{transform:perspective(2200px) rotateY(19deg) rotateZ(-1.6deg) skewY(-1.1deg)}}66%{{transform:perspective(2200px) rotateY(7deg) rotateZ(-.5deg)}}86%{{transform:perspective(2200px) rotateY(-3.5deg)}}94%{{transform:perspective(2200px) rotateY(1.2deg)}}100%{{transform:none}}}}
@keyframes flexb{{0%{{transform:perspective(2200px) rotateY(0)}}30%{{transform:perspective(2200px) rotateY(-19deg) rotateZ(1.6deg) skewY(1.1deg)}}66%{{transform:perspective(2200px) rotateY(-7deg) rotateZ(.5deg)}}86%{{transform:perspective(2200px) rotateY(3.5deg)}}94%{{transform:perspective(2200px) rotateY(-1.2deg)}}100%{{transform:none}}}}
.sheet.moving .pg .page::after{{content:'';position:absolute;inset:0;pointer-events:none;z-index:3;background:linear-gradient(100deg,transparent 32%,rgba(255,255,255,.38) 50%,transparent 68%);animation:sheen 1.1s ease both}}
@keyframes sheen{{0%{{transform:translateX(-70%);opacity:0}}30%{{opacity:1}}100%{{transform:translateX(70%);opacity:0}}}}
.blank{{background:linear-gradient(150deg,#241a4d,#4c37a3)}}
.cover{{align-items:center;justify-content:center;text-align:center;background:linear-gradient(150deg,#241a4d,#4c37a3);color:#fff}}
.cover h1{{font-size:clamp(24px,3.6vw,40px);font-weight:900;letter-spacing:-1px;margin:14px 0 8px}}
.cover p{{color:rgba(255,255,255,.75);font-size:15px}}
.cover .logo{{max-height:76px;max-width:220px;object-fit:contain}}
.hintturn{{position:absolute;bottom:20px;left:0;right:0;font-size:11.5px;color:rgba(255,255,255,.5)}}
.credit{{margin-top:22px;font-size:12px;color:rgba(255,255,255,.5)}}
.cat{{font-size:14px;font-weight:800;color:#6d28d9;background:#f3efff;border-radius:999px;padding:5px 15px;align-self:flex-start;margin-bottom:12px;z-index:2}}
.pgrid{{flex:1;display:grid;grid-template-columns:1fr 1fr;grid-template-rows:1fr 1fr;gap:14px;min-height:0}}
.prd{{display:flex;flex-direction:column;min-height:0;border:1px solid #f1eefb;border-radius:12px;padding:9px;overflow:hidden;background:#fff;z-index:2}}
.prd img,.noimg{{width:100%;flex:1;min-height:0;object-fit:cover;border-radius:8px;background:#f3efff}}
.prd img.zoom{{cursor:zoom-in}}
.noimg{{display:flex;align-items:center;justify-content:center;font-size:12px;font-weight:600;color:#8b7fc9;background:linear-gradient(135deg,#ede9fe,#ddd6fe)}}
.prd h3{{font-size:12px;font-weight:700;color:#1f2937;line-height:1.25;margin-top:7px;display:-webkit-box;-webkit-line-clamp:2;-webkit-box-orient:vertical;overflow:hidden}}
.prd .price{{color:#6d28d9;font-size:13.5px;font-weight:900;margin-top:2px}}
.single .sheet{{left:0;width:100%;transform:none!important;transition:none}}
.single .pg{{transform:none;backface-visibility:visible;opacity:0;pointer-events:none;transition:opacity .4s}}
.single .pg.cur{{opacity:1;pointer-events:auto}}
.single .pg .page,.single .pg.b .page{{border-radius:12px}}
.nav{{display:flex;align-items:center;gap:16px;margin-top:16px;flex-wrap:wrap;justify-content:center}}
.nav button{{background:rgba(255,255,255,.1);border:1px solid rgba(255,255,255,.25);color:#fff;border-radius:12px;padding:10px 20px;font-size:14px;font-weight:700;cursor:pointer;transition:background .15s}}
.nav button:hover{{background:rgba(255,255,255,.22)}}
.nav .cnt{{color:rgba(255,255,255,.65);font-size:13px;font-weight:600;min-width:70px;text-align:center}}
.side{{position:absolute;top:50%;transform:translateY(-50%);z-index:40;width:54px;height:54px;border-radius:50%;background:#fff;border:none;color:#4c37a3;font-size:26px;font-weight:900;cursor:pointer;box-shadow:0 10px 30px rgba(0,0,0,.4);display:flex;align-items:center;justify-content:center;transition:transform .15s}}
.side:hover{{transform:translateY(-50%) scale(1.12)}}
.side.prev{{left:-27px}}.side.next{{right:-27px}}
@media (max-width:700px){{.side.prev{{left:6px}}.side.next{{right:6px}}}}
.contact{{display:flex;gap:16px;flex-wrap:wrap;justify-content:center;margin-top:18px}}
.contact a{{color:#c4b5fd;font-weight:700;text-decoration:none;font-size:14px}}
.lb{{display:none;position:fixed;inset:0;background:rgba(12,8,32,.94);z-index:50;flex-direction:column;align-items:center;justify-content:center;padding:24px}}
.lb.open{{display:flex}}
.lb img{{max-width:min(92vw,900px);max-height:68vh;object-fit:contain;border-radius:14px;background:#fff}}
.lb .cap{{color:#e9e4ff;font-size:14px;font-weight:600;margin-top:12px;text-align:center;max-width:700px}}
.lb .lnav{{display:flex;gap:14px;margin-top:14px;align-items:center}}
.lb button{{background:rgba(255,255,255,.12);border:1px solid rgba(255,255,255,.25);color:#fff;border-radius:12px;padding:10px 18px;font-size:14px;font-weight:700;cursor:pointer}}
.lb .cnt{{color:rgba(255,255,255,.6);font-size:13px}}
.lb .x{{position:absolute;top:18px;right:22px}}
@media print{{body{{background:#fff;display:block}}.stage{{perspective:none;width:100%;aspect-ratio:auto;max-height:none}}.sheet{{position:static;transform:none!important;height:auto}}.pg{{position:static;transform:none!important;opacity:1!important;page-break-after:always;height:100vh}}.page{{position:relative;box-shadow:none}}.nav,.lb,.side,.spine{{display:none!important}}}}
</style></head><body>
<div class="stage" id="book"><div class="spine"></div>{sheets}
  <button class="side prev" onclick="event.stopPropagation();go(-1)">&#8249;</button>
  <button class="side next" onclick="event.stopPropagation();go(1)">&#8250;</button>
</div>
<div class="nav">
  <button onclick="go(-1)">&#8249; Précédente</button>
  <span class="cnt" id="cnt"></span>
  <button onclick="go(1)">Suivante &#8250;</button>
  <button id="snd" onclick="sndMuted=!sndMuted;this.textContent=sndMuted?&#39;Son : coupé&#39;:&#39;Son : activé&#39;" style="font-size:12px;padding:8px 14px">Son : activé</button>
</div>
<div class="lb" id="lb"><button class="x" onclick="lbClose()">&#10005; Fermer</button>
  <img id="lbImg" src="" alt=""><div class="cap" id="lbCap"></div>
  <div class="lnav"><button onclick="lbNav(-1)">&#8249;</button><span class="cnt" id="lbCnt"></span><button onclick="lbNav(1)">&#8250;</button></div>
</div>
<script>
var N={n},S={S},pos=0,sndMuted=false,_actx=null;
var stage=document.getElementById('book');
var sheets=document.querySelectorAll('.sheet');
var pgs=document.querySelectorAll('.pg');
function isSingle(){{return window.innerWidth<1000}}
function flipSound(){{
  if(sndMuted)return;
  try{{
    _actx=_actx||new (window.AudioContext||window.webkitAudioContext)();
    if(_actx.state==='suspended')_actx.resume();
    var t0=_actx.currentTime;
    function burst(start,dur,f1,f2,vol,curve){{
      var sr=_actx.sampleRate,buf=_actx.createBuffer(1,Math.floor(sr*dur),sr),d=buf.getChannelData(0);
      for(var i=0;i<d.length;i++){{var t=i/d.length;d[i]=(Math.random()*2-1)*Math.pow(Math.sin(Math.PI*Math.min(1,t/curve))* (1-t*0.55),2);}}
      var src=_actx.createBufferSource();src.buffer=buf;
      var lp=_actx.createBiquadFilter();lp.type='lowpass';lp.frequency.setValueAtTime(f2,start);lp.Q.value=0.4;
      var hp=_actx.createBiquadFilter();hp.type='highpass';hp.frequency.value=f1;hp.Q.value=0.3;
      var g=_actx.createGain();g.gain.setValueAtTime(vol,start);
      g.gain.exponentialRampToValueAtTime(0.008,start+dur);
      src.connect(hp);hp.connect(lp);lp.connect(g);g.connect(_actx.destination);src.start(start);
    }}
    burst(t0,0.55,200,850,0.16,0.45);        /* glissement feutré */
    burst(t0+0.62,0.12,300,1400,0.10,0.3);   /* la feuille se pose */
  }}catch(e){{}}
}}
function upd(){{
  if(isSingle()){{
    stage.classList.add('single');
    pgs.forEach(function(p,i){{p.classList.toggle('cur',i===pos);}});
    document.getElementById('cnt').textContent=(pos+1)+' / '+N;
  }}else{{
    stage.classList.remove('single');
    var fs=Math.floor((pos+1)/2);
    sheets.forEach(function(sh,k){{
      sh.classList.toggle('flip',k<fs);
      sh.style.zIndex = k<fs ? k : (S-k);
    }});
    var lbl = fs===0 ? '1' : (2*fs)+(2*fs+1<=N ? '-'+(2*fs+1) : '');
    document.getElementById('cnt').textContent=lbl+' / '+N;
  }}
}}
function go(d){{
  var old=pos;
  if(isSingle()){{pos=Math.max(0,Math.min(N-1,pos+d));}}
  else{{
    var fs=Math.floor((pos+1)/2)+d;
    fs=Math.max(0,Math.min(S,fs));
    pos = fs===0 ? 0 : Math.min(N-1,2*fs-1);
  }}
  if(pos===old)return;
  if(!isSingle()){{
    var ofs=Math.floor((old+1)/2),nfs=Math.floor((pos+1)/2);
    var mv=sheets[d>0?ofs:nfs];
    if(mv)curlTurn(mv,d);
  }}
  flipSound();upd();
}}
var NS=10,DUR=1200;
function curlTurn(mv,d){{
  var turn=document.createElement('div');turn.className='turn';
  var fF=mv.querySelector('.pg.f'),fB=mv.querySelector('.pg.b');
  var w=100/NS;
  var strips=[];
  for(var i=0;i<NS;i++){{
    var st=document.createElement('div');st.className='strip';
    var cf=fF.cloneNode(true),cb=fB.cloneNode(true);
    cf.style.clipPath='inset(0 '+(100-(i+1)*w)+'% 0 '+(i*w)+'%)';
    cb.style.clipPath='inset(0 '+(i*w)+'% 0 '+(100-(i+1)*w)+'%)';
    st.appendChild(cf);st.appendChild(cb);turn.appendChild(st);strips.push(st);
  }}
  stage.appendChild(turn);
  var oldTr=mv.style.transition;mv.style.transition='none';mv.style.visibility='hidden';
  var ease='cubic-bezier(.3,.05,.2,1)';
  strips.forEach(function(st,i){{
    var f=i/(NS-1);
    var kf = d>0 ? [
      {{transform:'rotateY(0deg)'}},
      {{transform:'rotateY('+(-(72-30*f))+'deg)',offset:.32}},
      {{transform:'rotateY('+(-(148-14*f))+'deg)',offset:.62}},
      {{transform:'rotateY('+(-(181+7*f))+'deg)',offset:.85}},
      {{transform:'rotateY(-180deg)'}}
    ] : [
      {{transform:'rotateY(-180deg)'}},
      {{transform:'rotateY('+(-(108+30*f))+'deg)',offset:.32}},
      {{transform:'rotateY('+(-(32+14*f))+'deg)',offset:.62}},
      {{transform:'rotateY('+((1+7*f))+'deg)',offset:.85}},
      {{transform:'rotateY(0deg)'}}
    ];
    st.animate(kf,{{duration:DUR,easing:ease,fill:'both'}});
  }});
  setTimeout(function(){{
    turn.remove();mv.style.visibility='';
    requestAnimationFrame(function(){{mv.style.transition=oldTr;}});
  }},DUR+40);
}}
upd();
window.addEventListener('resize',upd);
document.addEventListener('keydown',function(e){{
  if(document.getElementById('lb').classList.contains('open')){{
    if(e.key==='Escape')lbClose();if(e.key==='ArrowLeft')lbNav(-1);if(e.key==='ArrowRight')lbNav(1);
    return;
  }}
  if(e.key==='ArrowLeft')go(-1);if(e.key==='ArrowRight')go(1);
}});
stage.addEventListener('click',function(ev){{
  var t=ev.target;
  if(t.classList&&t.classList.contains('zoom')){{
    try{{_lbImgs=JSON.parse(t.getAttribute('data-imgs')||'[]')}}catch(e){{_lbImgs=[t.src]}}
    if(!_lbImgs.length)_lbImgs=[t.src];
    _lbIdx=0;document.getElementById('lbCap').textContent=t.getAttribute('data-title')||'';
    lbShow();document.getElementById('lb').classList.add('open');
    return;
  }}
  var r=this.getBoundingClientRect();
  go(ev.clientX - r.left > r.width/2 ? 1 : -1);
}});
var sx=null;
stage.addEventListener('touchstart',function(e){{sx=e.touches[0].clientX}});
stage.addEventListener('touchend',function(e){{
  if(sx===null)return;var dx=e.changedTouches[0].clientX-sx;sx=null;
  if(Math.abs(dx)>40)go(dx<0?1:-1);
}});
var _lbImgs=[],_lbIdx=0;
function lbShow(){{document.getElementById('lbImg').src=_lbImgs[_lbIdx];
  document.getElementById('lbCnt').textContent=(_lbIdx+1)+' / '+_lbImgs.length;}}
function lbNav(d){{_lbIdx=(_lbIdx+d+_lbImgs.length)%_lbImgs.length;lbShow();}}
function lbClose(){{document.getElementById('lb').classList.remove('open');}}
</script>
</body></html>'''


def _catalog_contact_html(owner: dict) -> str:
    import html as _html
    e = _html.escape
    parts = []
    if owner.get("website"):
        url = owner["website"] if owner["website"].startswith("http") else "https://" + owner["website"]
        label = url.replace("https://", "").replace("http://", "").rstrip("/")
        parts.append(f'<a href="{e(url)}" target="_blank" rel="noopener">{e(label)}</a>')
    if owner.get("contact_email"):
        parts.append(f'<a href="mailto:{e(owner["contact_email"])}">{e(owner["contact_email"])}</a>')
    if owner.get("phone"):
        parts.append(f'<a href="tel:{e(owner["phone"])}">{e(owner["phone"])}</a>')
    return f'<div class="contact">{"".join(parts)}</div>' if parts else ''


@app.get("/catalogue/{token}")
async def public_catalogue(token: str):
    """Public, tokenised, always-up-to-date brand catalogue (revocable link)."""
    from app.db import resolve_catalog_token
    from app.routes.public_api import _latest_listings
    row = resolve_catalog_token(token)
    if not row:
        raise HTTPException(404, "Catalogue introuvable")
    if not _rate_limit(f"cat:{token}", limit=300, window=3600):
        raise HTTPException(429, "Trop de requêtes")
    listings = _latest_listings(row["user_email"])
    from app.services.image_gen import AMAZON_IMAGE_TYPES
    raw_urls = _get_image_urls_for_skus([l.get("sku") for l in listings if l.get("sku")])
    order = [t["id"] for t in AMAZON_IMAGE_TYPES]
    image_urls = {sku: [d[t] for t in order if t in d] for sku, d in raw_urls.items()}
    return Response(content=_render_public_catalog(row, listings, image_urls),
                    media_type="text/html; charset=utf-8")


def _bundled_listing_loader_bytes() -> Optional[bytes]:
    """The generic Amazon Listing Loader (offer) template. Prefer the DB-seeded
    copy (reliable across deploys, same store the category templates use); fall
    back to the bundled file on disk."""
    # 1) DB (seeded from data/amazon_templates/_LISTING_LOADER.xlsm at startup)
    try:
        tpl = get_amazon_template_by_type("_LISTING_LOADER")
        if tpl and tpl.get("data_b64"):
            return base64.b64decode(tpl["data_b64"])
    except Exception:
        pass
    # 2) File fallback — try a few candidate locations
    here = os.path.dirname(os.path.abspath(__file__))
    candidates = [
        os.path.join(os.path.dirname(here), "data", "amazon_templates", "_LISTING_LOADER.xlsm"),
        os.path.join(os.getcwd(), "data", "amazon_templates", "_LISTING_LOADER.xlsm"),
    ]
    for path in candidates:
        try:
            with open(path, "rb") as f:
                return f.read()
        except Exception:
            continue
    return None


class SellerCentralPackRequest(BaseModel):
    listings: List[AmazonListing]
    compliance_data: Optional[dict] = None
    category_template_id: Optional[str] = None   # explicit category template to use


@app.post("/api/export/seller-central-pack")
async def export_seller_central_pack(req: SellerCentralPackRequest, request: Request):
    """One smart download: a ZIP with BOTH Amazon Excel files filled and ready —
    the category template (create new products) when available for the listings'
    category, and the Listing Loader (offers for existing catalog products) which
    always works. Includes a short instructions file.

    The category template is chosen by category_template_id when provided
    (explicit pick from the library); otherwise we fall back to matching the
    listings' category text against stored template product types."""
    import io as _io, zipfile as _zipfile
    listings = req.listings
    if not listings:
        raise HTTPException(400, "Aucune fiche à exporter")
    image_urls = _get_image_urls_for_skus([l.sku for l in listings])
    compliance = req.compliance_data or None

    buf = _io.BytesIO()
    included = []
    cat_included = False
    with _zipfile.ZipFile(buf, "w", _zipfile.ZIP_DEFLATED) as zf:
        # 1) Category template (create new products)
        #    Prefer the explicitly chosen template; else auto-match by category text.
        category = (listings[0].category or "").strip().upper()
        cat_tpl = None
        if req.category_template_id:
            cat_tpl = get_amazon_template(req.category_template_id)
        if not cat_tpl and category:
            cat_tpl = get_amazon_template_by_type(category)
        if cat_tpl:
            try:
                cat_bytes = base64.b64decode(cat_tpl["data_b64"])
                filled = fill_amazon_template(cat_bytes, listings,
                                              image_urls=image_urls, compliance_data=compliance,
                                              for_new_products=True)
                zf.writestr("1_NOUVEAUX_PRODUITS.xlsm", filled)
                included.append("1_NOUVEAUX_PRODUITS.xlsm (créer de nouveaux produits)")
                cat_included = True
            except Exception as e:
                log.error(f"[pack] category fill failed: {e}")

        # 2) Listing Loader (offers for existing catalog products) — always
        ll_bytes = _bundled_listing_loader_bytes()
        if ll_bytes:
            try:
                filled_ll = fill_amazon_template(ll_bytes, listings,
                                                 image_urls=image_urls, compliance_data=compliance)
                zf.writestr("2_PRODUITS_EXISTANTS_listing_loader.xlsm", filled_ll)
                included.append("2_PRODUITS_EXISTANTS_listing_loader.xlsm (offres pour produits déjà au catalogue)")
            except Exception as e:
                log.error(f"[pack] listing loader fill failed: {e}")

        if not included:
            raise HTTPException(422, "Aucun modèle disponible pour cette catégorie")

        # 3) Instructions
        cat_note = ("• 1_NOUVEAUX_PRODUITS.xlsm → Seller Central > Catalogue > Ajouter des "
                    "produits via un fichier > onglet « Téléverser » (créer de nouveaux produits).\n"
                    if cat_tpl else
                    "• (Modèle de catégorie non disponible pour cette catégorie — utilisez le "
                    "Listing Loader, ou ajoutez le modèle de votre catégorie dans SynqIO.)\n")
        readme = (
            "PACK SYNQIO — IMPORT SELLER CENTRAL\n"
            "===================================\n\n"
            "Deux fichiers selon votre cas :\n\n"
            + cat_note +
            "• 2_PRODUITS_EXISTANTS_listing_loader.xlsm → Seller Central > Catalogue > "
            "Ajouter des produits > « Mettre en vente des produits figurant dans le catalogue "
            "Amazon » (met à jour/crée l'offre par SKU : prix, stock, état, photos).\n\n"
            "Les photos sont incluses via des URLs publiques — Amazon les télécharge "
            "automatiquement à l'import.\n\n"
            "Astuce : si vos produits n'ont pas de code-barres (EAN/UPC), demandez une "
            "exemption GTIN dans Seller Central (Catalogue > Ajouter des produits > "
            "Demander une exemption de code-barres).\n"
        )
        zf.writestr("LISEZ-MOI.txt", readme)

    buf.seek(0)
    d = __import__("datetime").datetime.now()
    fn = f"synqio_seller_central_{d.strftime('%Y%m%d-%H%M%S')}.zip"
    return Response(
        content=buf.getvalue(),
        media_type="application/zip",
        headers={
            "Content-Disposition": f"attachment; filename={fn}",
            "X-Category-Included": "1" if cat_included else "0",
            "X-Category-Type": category or "",
            "Access-Control-Expose-Headers": "X-Category-Included, X-Category-Type",
        },
    )


@app.post("/api/export/variation-flat-file")
async def export_variation_flat_file(listings: List[AmazonListing], request: Request):
    skus = [l.sku for l in listings if l.is_parent or not l.parent_sku]
    image_urls = _get_image_urls_for_skus(skus)
    return Response(
        content=to_variation_flat_file_xlsx(listings, image_urls=image_urls),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=amazon_variations.xlsx"},
    )


# ── Images ───────────────────────────────────────────────────────────────────

class ImageRequest(BaseModel):
    sku: str
    product_name: str
    brand: str = ""
    category: str = ""
    features: List[str] = []
    color: str = ""
    material: str = ""
    selected_types: Optional[List[str]] = None
    reference_image_url: Optional[str] = None
    batch_id: Optional[str] = None   # generation history id to attach images to
    marketplace: Optional[str] = None


async def _persist_images(sku: str, images: list) -> list:
    """Download temp DALL-E URLs / decode base64 → save to storage permanently.
    Returns the same list with urls replaced by /api/photos/... paths."""
    import base64 as _b64
    for img in images:
        url = img.get("url") or ""
        if not url or not img.get("has_image"):
            continue
        ext = "png"
        filename = f"gen_{sku}_{img.get('id','img')}.{ext}"
        if storage.exists(filename):
            img["url"] = f"/api/photos/{filename}"
            continue
        try:
            if url.startswith("data:image"):
                data_part = url.split(",", 1)[1]
                _save_photo(filename, _b64.b64decode(data_part))
                img["url"] = f"/api/photos/{filename}"
            elif url.startswith("http"):
                async with httpx.AsyncClient(timeout=30) as client:
                    resp = await client.get(url)
                    if resp.status_code == 200:
                        _save_photo(filename, resp.content)
                        img["url"] = f"/api/photos/{filename}"
        except Exception:
            pass  # keep original url as fallback
    return images


async def _run_image_job(job_id: str, req: ImageRequest, email: str):
    _jobs[job_id]["status"] = "running"
    log.info("job started", extra={"job_id": job_id})
    try:
        # Résoudre la photo de référence :
        # - /api/photos/{nom} → bytes depuis _temp_photos (ZIP uploadé, zéro HTTP)
        # - URL externe → téléchargement dans image_gen
        ref_url = req.reference_image_url or None
        ref_bytes = None
        if ref_url and ref_url.startswith("/api/photos/"):
            fname = ref_url[len("/api/photos/"):]
            ref_bytes = _load_photo(fname)
            ref_url = None  # on a les bytes, inutile de télécharger

        images, img_tokens = await generate_product_images(
            sku=req.sku,
            product_name=req.product_name,
            brand=req.brand,
            category=req.category,
            features=req.features,
            color=req.color,
            material=req.material,
            selected_types=req.selected_types,
            reference_image_url=ref_url,
            reference_image_bytes=ref_bytes,
            marketplace=req.marketplace,
        )
        generated = [i for i in images if i.get("has_image")]
        if generated and email:
            log_usage(email, "image_generated", len(generated))
        if email and img_tokens:
            if img_tokens.get("input_tokens"):
                log_usage(email, "tokens_in", img_tokens["input_tokens"])
            if img_tokens.get("output_tokens"):
                log_usage(email, "tokens_out", img_tokens["output_tokens"])

        # Persist images to disk so URLs never expire
        images = await _persist_images(req.sku, images)

        # Attach images to generation history if batch_id provided
        if req.batch_id and email:
            try:
                from app.db import save_generation_images
                save_generation_images(req.batch_id, email, req.sku, images)
            except Exception:
                pass

        openai_ok = bool(os.getenv("OPENAI_API_KEY"))
        result_img = {
            "sku": req.sku,
            "images": images,
            "images_generated": openai_ok,
            "openai_configured": openai_ok,
            "total": len(images),
            "reference_image_used": img_tokens.get("reference_image_used", False),
        }
        _jobs[job_id].update({"status": "done", "result": result_img})
        log.info("job done", extra={"job_id": job_id, "user": email})
        try:
            update_job_db(job_id, status="done", result_json=json.dumps(result_img, default=str))
        except Exception:
            pass
    except Exception as e:
        _jobs[job_id].update({"status": "failed", "error": str(e)})
        log.error("job failed", extra={"job_id": job_id, "user": email})
        try:
            update_job_db(job_id, status="failed", error=str(e))
        except Exception:
            pass


@app.post("/api/generate-images")
async def generate_images(req: ImageRequest, request: Request):
    if not os.getenv("ANTHROPIC_API_KEY"):
        raise HTTPException(503, "ANTHROPIC_API_KEY manquante")
    email = getattr(request.state, "user_email", None)
    if email:
        usage = get_user_usage(email)
        imgs_remaining = max(0, usage.get("images_quota", 200) - usage.get("images_used", 0))
        if imgs_remaining <= 0:
            raise HTTPException(429,
                f"Quota d'images atteint ce mois (plan {usage['plan_label']}). "
                "Contactez-nous pour augmenter votre quota.")
        if not _rate_limit(f"{email}:images", limit=10, window=3600):
            raise HTTPException(429, "Trop de générations d'images — maximum 10/heure.")
    job_id = str(uuid4())
    _jobs[job_id] = {"status": "pending", "progress": 0, "total": 7, "created_at": time.time()}
    _cleanup_jobs()
    try:
        save_job(job_id, email, "images", 7)
    except Exception:
        pass
    asyncio.create_task(_run_image_job(job_id, req, email))
    return {"job_id": job_id}


class SingleImageRequest(BaseModel):
    sku: str
    image_id: str
    prompt: str


async def _run_single_image_job(job_id: str, req: SingleImageRequest, email: str):
    _jobs[job_id]["status"] = "running"
    log.info("job started", extra={"job_id": job_id})
    try:
        url = await _generate_image_dalle3(req.prompt, req.image_id)
        openai_ok = bool(os.getenv("OPENAI_API_KEY"))
        if url and email:
            log_usage(email, "image_generated", 1)
            try:
                from app.services.usage import log_image_regen
                log_image_regen(email, req.sku, req.image_id)
            except Exception:
                pass
        # Persist to disk
        if url:
            persisted = await _persist_images(req.sku, [{"id": req.image_id, "url": url, "has_image": True}])
            url = persisted[0]["url"] if persisted else url
        result_single = {
            "sku": req.sku,
            "image_id": req.image_id,
            "url": url,
            "prompt": req.prompt,
            "openai_configured": openai_ok,
        }
        _jobs[job_id].update({"status": "done", "result": result_single})
        log.info("job done", extra={"job_id": job_id, "user": email})
        try:
            update_job_db(job_id, status="done", result_json=json.dumps(result_single, default=str))
        except Exception:
            pass
    except Exception as e:
        _jobs[job_id].update({"status": "failed", "error": str(e)})
        log.error("job failed", extra={"job_id": job_id, "user": email})
        try:
            update_job_db(job_id, status="failed", error=str(e))
        except Exception:
            pass


@app.post("/api/generate-image-single")
async def generate_image_single(req: SingleImageRequest, request: Request):
    if not os.getenv("OPENAI_API_KEY"):
        raise HTTPException(503, "OPENAI_API_KEY manquante — impossible de régénérer l'image")
    email = getattr(request.state, "user_email", None)
    if email:
        from app.services.usage import get_image_regen_count, MAX_IMAGE_REGENS
        # Quota mensuel d'images du plan (les régénérations le consomment aussi)
        usage = get_user_usage(email)
        if usage.get("images_used", 0) >= usage.get("images_quota", 200):
            raise HTTPException(429,
                f"Quota d'images atteint ce mois (plan {usage['plan_label']}). "
                "Contactez-nous pour augmenter votre quota.")
        # Plafond de régénération par visuel : 3 max par image d'une fiche
        regens = get_image_regen_count(email, req.sku, req.image_id)
        if regens >= MAX_IMAGE_REGENS:
            raise HTTPException(429,
                f"Limite de régénération atteinte pour ce visuel ({MAX_IMAGE_REGENS} max par image). "
                "Modifiez le prompt d'un autre visuel ou contactez-nous si besoin.")
        if not _rate_limit(f"{email}:images", limit=10, window=3600):
            raise HTTPException(429, "Trop de générations d'images — maximum 10/heure.")
    job_id = str(uuid4())
    _jobs[job_id] = {"status": "pending", "progress": 0, "total": 1, "created_at": time.time()}
    _cleanup_jobs()
    try:
        save_job(job_id, email, "single_image", 1)
    except Exception:
        pass
    asyncio.create_task(_run_single_image_job(job_id, req, email))
    return {"job_id": job_id}


# ── Usage ─────────────────────────────────────────────────────────────────────

@app.get("/api/usage/me")
async def usage_me(request: Request):
    email = getattr(request.state, "user_email", None)
    if not email:
        raise HTTPException(401, "Non authentifié")
    return get_user_usage(email)


@app.get("/api/admin/usage")
async def usage_all(request: Request):
    from app.services.auth import is_admin, _decode_token_data
    email = getattr(request.state, "user_email", None)
    if not email:
        raise HTTPException(403, "Accès réservé aux administrateurs")
    # Accept JWT adm claim (survives DB wipes) or DB flag
    auth = request.headers.get("Authorization", "")
    jwt_data = _decode_token_data(auth.split(" ", 1)[1]) if auth.startswith("Bearer ") else {}
    if not (bool((jwt_data or {}).get("adm")) or is_admin(email)):
        raise HTTPException(403, "Accès réservé aux administrateurs")
    return get_all_users_usage()


@app.get("/api/admin/amazon-credentials")
async def amazon_credentials_check(request: Request):
    """Show current Amazon credentials and verify seller_id against the Listings API."""
    from app.services.auth import is_admin, _decode_token_data
    from app.services.amazon_sp import (
        _get_sp_credentials, _get_lwa_token, _assume_role, _sign_request,
        _sp_endpoint, MARKETPLACE_IDS, Marketplace as _Mkt,
    )
    from urllib.parse import quote
    import asyncio as _asyncio, logging as _log
    email = getattr(request.state, "user_email", None)
    auth = request.headers.get("Authorization", "")
    jwt_data = _decode_token_data(auth.split(" ", 1)[1]) if auth.startswith("Bearer ") else {}
    if not (bool((jwt_data or {}).get("adm")) or is_admin(email)):
        raise HTTPException(403, "Accès réservé aux administrateurs")

    creds = _get_sp_credentials(email)
    marketplace_id = MARKETPLACE_IDS[_Mkt.AMAZON_FR]
    seller_id = creds.get("seller_id", "")

    result = {
        "seller_id_in_db": seller_id,
        "marketplace_id": marketplace_id,
        "lwa_client_id": (creds.get("lwa_client_id") or "")[:8] + "...",
    }

    try:
        lwa_token = await _get_lwa_token(creds)
        loop = _asyncio.get_event_loop()
        temp_creds = await loop.run_in_executor(None, lambda: _assume_role(creds))

        # 1. Participations — proves auth works
        part_url = f"{_sp_endpoint()}/sellers/v1/marketplaceParticipations"
        part_h = _sign_request("GET", part_url, b"", temp_creds, lwa_token)
        async with httpx.AsyncClient(timeout=20) as c:
            part_resp = await c.get(part_url, headers=part_h)
        result["participations_status"] = part_resp.status_code
        result["participations"] = part_resp.json() if part_resp.status_code == 200 else part_resp.text[:300]

        # 2. Catalog list — verifies seller_id is the correct Merchant Token
        cat_url = f"{_sp_endpoint()}/listings/2021-08-01/items/{seller_id}?marketplaceIds={marketplace_id}&pageSize=1"
        cat_h = _sign_request("GET", cat_url, b"", temp_creds, lwa_token)
        async with httpx.AsyncClient(timeout=20) as c:
            cat_resp = await c.get(cat_url, headers=cat_h)
        result["catalog_list_status"] = cat_resp.status_code
        result["catalog_list_seller_id_valid"] = cat_resp.status_code != 400
        result["catalog_list_response"] = cat_resp.json() if cat_resp.content else {}

    except Exception as exc:
        result["error"] = str(exc)

    return result


@app.post("/api/admin/amazon-credentials")
async def amazon_credentials_update(request: Request):
    """Update seller_id and/or refresh_token in amazon_credentials for this user."""
    from app.services.auth import is_admin, _decode_token_data
    from app.db import get_db
    email = getattr(request.state, "user_email", None)
    auth = request.headers.get("Authorization", "")
    jwt_data = _decode_token_data(auth.split(" ", 1)[1]) if auth.startswith("Bearer ") else {}
    if not (bool((jwt_data or {}).get("adm")) or is_admin(email)):
        raise HTTPException(403, "Accès réservé aux administrateurs")
    body = await request.json()
    new_seller_id = (body.get("seller_id") or "").strip()
    new_refresh_token = (body.get("refresh_token") or "").strip()
    if not new_seller_id and not new_refresh_token:
        raise HTTPException(400, "seller_id or refresh_token required")
    conn = get_db()
    updated = {}
    if new_seller_id:
        conn.execute(
            "UPDATE amazon_credentials SET seller_id = ? WHERE user_email = ?",
            (new_seller_id, email),
        )
        updated["seller_id"] = new_seller_id
    if new_refresh_token:
        from app.services.crypto import encrypt_token
        conn.execute(
            "UPDATE amazon_credentials SET refresh_token = ? WHERE user_email = ?",
            (encrypt_token(new_refresh_token), email),
        )
        updated["refresh_token"] = new_refresh_token[:12] + "..."
    conn.commit()
    conn.close()
    return {"updated": True, **updated}


@app.get("/api/admin/amazon-product-types")
async def amazon_product_types_search(request: Request, keywords: str = "collar,pet,animal,dog"):
    """Diagnostic: list valid Amazon product types for this marketplace."""
    from app.services.auth import is_admin, _decode_token_data
    from app.services.amazon_sp import search_product_types_for_marketplace
    from app.models import Marketplace
    email = getattr(request.state, "user_email", None)
    auth = request.headers.get("Authorization", "")
    jwt_data = _decode_token_data(auth.split(" ", 1)[1]) if auth.startswith("Bearer ") else {}
    if not (bool((jwt_data or {}).get("adm")) or is_admin(email)):
        raise HTTPException(403, "Accès réservé aux administrateurs")
    kw_list = [k.strip() for k in keywords.split(",") if k.strip()]
    results = await search_product_types_for_marketplace(kw_list, Marketplace.AMAZON_FR, email)
    return results


@app.get("/api/image-types")
async def get_image_types():
    return AMAZON_IMAGE_TYPES


# ── Marketplaces ──────────────────────────────────────────────────────────────

# ── Generation history ────────────────────────────────────────────────────────

@app.get("/api/history")
async def get_history(request: Request):
    email = request.state.user_email
    return list_generations(email)


@app.get("/api/history/{batch_id}")
async def get_history_batch(batch_id: str, request: Request):
    email = request.state.user_email
    try:
        batch = get_generation(batch_id, email)
    except Exception as e:
        log.error(f"get_generation failed: {type(e).__name__}: {e}", extra={"email": email})
        raise HTTPException(500, f"{type(e).__name__}: {str(e)[:300]}")
    if not batch:
        raise HTTPException(status_code=404, detail="Lot introuvable")
    return batch


@app.delete("/api/history/{batch_id}")
async def delete_history_batch(batch_id: str, request: Request):
    email = request.state.user_email
    if not delete_generation(batch_id, email):
        raise HTTPException(status_code=404, detail="Lot introuvable")
    return {"ok": True}


class LabelUpdate(BaseModel):
    label: str


@app.patch("/api/history/{batch_id}/label")
async def update_history_label(batch_id: str, body: LabelUpdate, request: Request):
    email = request.state.user_email
    if not update_generation_label(batch_id, email, body.label):
        raise HTTPException(status_code=404, detail="Lot introuvable")
    return {"ok": True}


class ListingsUpdate(BaseModel):
    listings: list


@app.put("/api/history/{batch_id}/listings")
async def update_history_listings(batch_id: str, body: ListingsUpdate, request: Request):
    """Persist client-side edits (notably declination children) back to the
    stored batch so a later restore matches what the user sees in the UI."""
    email = request.state.user_email
    if not update_generation_listings(batch_id, email, body.listings):
        raise HTTPException(status_code=404, detail="Lot introuvable")
    return {"ok": True}


@app.get("/api/marketplaces")
async def list_marketplaces():
    return [
        {"id": "amazon_fr",  "name": "Amazon France",    "flag": "🇫🇷"},
        {"id": "amazon_de",  "name": "Amazon Allemagne",  "flag": "🇩🇪"},
        {"id": "amazon_it",  "name": "Amazon Italie",     "flag": "🇮🇹"},
        {"id": "amazon_es",  "name": "Amazon Espagne",    "flag": "🇪🇸"},
        {"id": "amazon_uk",  "name": "Amazon UK",         "flag": "🇬🇧"},
        {"id": "amazon_nl",  "name": "Amazon Pays-Bas",   "flag": "🇳🇱"},
        {"id": "amazon_se",  "name": "Amazon Suède",      "flag": "🇸🇪"},
        {"id": "amazon_pl",  "name": "Amazon Pologne",    "flag": "🇵🇱"},
        {"id": "amazon_be",  "name": "Amazon Belgique",   "flag": "🇧🇪"},
    ]


# ── Help chatbot ──────────────────────────────────────────────────────────────

_HELP_SYSTEM = """Tu es l'assistant d'aide de SynqIO, une application de création de fiches produits Amazon optimisées.

Tu réponds UNIQUEMENT sur deux sujets :
1. Comment utiliser l'interface SynqIO (boutons, étapes, workflow, fonctionnalités)
2. Les règles et bonnes pratiques Amazon pour bien vendre sur Seller Central

Tu ne réponds PAS sur : la technique, le code, les APIs, l'infrastructure, le prix des plans ou toute autre question hors-sujet. Si quelqu'un te pose une question hors-sujet, redirige poliment vers ces deux domaines.
Tu ne mentionnes jamais de technologie sous-jacente.
Tu réponds toujours en français, de façon claire, concise et bienveillante. Utilise des listes à puces quand c'est utile.

=== GUIDE SYNQIO ===
Workflow en 3 étapes :
1. Importer ses produits — glisser-déposer un fichier Excel/CSV (colonnes : SKU, Titre, Description, EAN, URL image) ou décrire ses produits via le Chat IA
2. Générer les fiches — cliquer sur « Générer » pour obtenir titre optimisé, 5 bullet points, description A+, mots-clés backend
3. Exporter et mettre en ligne — bouton « Exporter Excel » → déposer le fichier dans Seller Central via Catalogue > Ajouter des produits via importation

Fonctionnalités clés :
- Score SEO (0–100) : évalue la qualité de chaque fiche selon les critères Amazon
- Génération de visuels : 7 images IA par produit (hero fond blanc, lifestyle, infographie, dimensions, packaging)
- Bouton « Générer toutes les photos » : lance toutes les générations en arrière-plan, vous pouvez continuer à travailler
- Historique : retrouvez vos générations précédentes depuis l'icône horloge
- Multi-sessions : chaque nouvelle génération crée un onglet séparé, accessible via les chips en haut des résultats
- Suppression d'une fiche : bouton ✕ sur la carte produit
- Chat IA : décrivez vos produits en langage naturel, l'assistant extrait les informations et lance la génération

=== RÈGLES AMAZON ===
Titre :
- Max 75 caractères (nouvelle règle Amazon 2026, espaces compris) — un titre plus long sera tronqué par Amazon
- Structure : Marque + Type produit + Attribut clé + Couleur/Matière + Taille/Quantité
- Interdit : prix, promotions (« soldes », « promo »), livraison gratuite, majuscules excessives, superlatifs (« meilleur », « n°1 »)
- L'outil "Migration des titres" de SynqIO permet de raccourcir tous ses titres en masse automatiquement

Bullet points :
- Exactement 5 bullets, max 500 caractères chacun
- Commencer par une majuscule, pas de point final
- Mettre les bénéfices consommateur en premier, les caractéristiques techniques ensuite
- Éviter la répétition avec le titre

Description :
- Max 2 000 caractères
- Peut contenir du HTML basique : <b>, <br>, <ul><li>
- Développer les usages, le contexte d'utilisation, les arguments de réassurance

Images :
- Image principale (hero) : fond blanc obligatoire, pas de texte ni logo, produit occupe 85 % du cadre
- Format : JPEG ou PNG, min 1 000×1 000 px, recommandé 2 000×2 000 px
- Max 7 images au total par fiche
- Images secondaires : lifestyle, infographies, comparatifs autorisés

Mots-clés backend :
- Max 249 bytes (pas caractères) au total
- Séparés par des espaces (pas de virgules)
- Ne pas répéter les mots déjà présents dans le titre ou les bullets
- Inclure variantes orthographiques, synonymes, usages alternatifs

Catégories :
- Choisir la plus précise possible (sous-catégorie > catégorie générique)
- Une mauvaise catégorie = moins de visibilité dans les recherches
- Vérifier les attributs obligatoires de la catégorie dans le flat file

Flat file (fichier d'importation) :
- Télécharger le template depuis Seller Central : Catalogue > Télécharger une feuille de calcul
- Sélectionner la bonne catégorie avant de télécharger le template
- SynqIO exporte directement au bon format : Exporter > Excel
- Importer via : Seller Central > Catalogue > Ajouter des produits via importation > Charger un fichier

EAN / GTIN :
- Obligatoire dans la plupart des catégories
- Exemption GTIN possible mais déconseillée (moins de visibilité)
- L'EAN doit correspondre au GS1 officiel du produit
"""

class HelpMessage(BaseModel):
    role: str   # "user" | "assistant"
    content: str

class HelpChatRequest(BaseModel):
    messages: list[HelpMessage]

@app.post("/api/help-chat")
async def help_chat(req_body: HelpChatRequest, request: Request):
    if not os.getenv("ANTHROPIC_API_KEY"):
        raise HTTPException(503, "Assistant indisponible")
    from app.services.ai_agent import get_client
    messages = [{"role": m.role, "content": m.content} for m in req_body.messages[-12:]]
    resp = await get_client().messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=600,
        system=[{"type": "text", "text": _HELP_SYSTEM, "cache_control": {"type": "ephemeral"}}],
        messages=messages,
    )
    return {"reply": resp.content[0].text}


# ── Audit ──────────────────────────────────────────────────────────────────────

_AUDIT_SYSTEM = """Tu es un expert Amazon SEO. Analyse la fiche produit fournie et retourne UNIQUEMENT un objet JSON valide, sans texte avant ni après, sans bloc markdown.

Structure JSON exacte :
{
  "global_score": <entier 0-100>,
  "title":    {"score": <0-100>, "issues": [<max 3 strings courtes>], "suggestions": [<max 3 strings courtes>]},
  "bullets":  {"score": <0-100>, "issues": [<max 3 strings courtes>], "suggestions": [<max 3 strings courtes>]},
  "keywords": {"score": <0-100>, "issues": [<max 3 strings courtes>], "suggestions": [<max 3 strings courtes>]},
  "images":   {"score": <0-100>, "issues": [<max 2 strings courtes>], "suggestions": [<max 2 strings courtes>]},
  "summary": "<2 phrases max>"
}

Critères de scoring :

TITRE (poids 30%) :
RÈGLE AMAZON 2026 : le titre doit faire ≤75 caractères (espaces compris). Un titre >75 chars sera tronqué par Amazon.
- Longueur ≤75 chars → 40 pts ; 76-100 chars → 15 pts (trop long, sera tronqué) ; >100 chars → 0 pts
- Mots-clés principaux dans les 60 premiers chars → 30 pts
- Pas de majuscules abusives, caractères interdits (!?$¡¿), superlatifs (meilleur, n°1), répétitions → 30 pts
NE PAS recommander d'allonger le titre si celui-ci est déjà ≤75 chars — c'est correct selon la norme actuelle.

BULLETS (poids 30%) :
- 5 bullets → 25 pts ; 4 → 15 pts ; <4 → 5 pts
- Longueur moyenne >150 chars → 25 pts ; 80-150 → 15 pts ; <80 → 5 pts
- Bénéfice client (pas juste technique) dans ≥3 bullets → 25 pts
- Mots-clés secondaires variés → 25 pts

KEYWORDS BACKEND (poids 20%) :
- Vide → 0
- ≤249 bytes → 40 pts
- Pas de répétition avec le titre → 30 pts
- Diversité et pertinence → 30 pts

IMAGES (poids 20%) :
- 7+ → 100 pts ; 5-6 → 70 pts ; 3-4 → 40 pts ; 1-2 → 20 pts ; 0 → 0 pts

GLOBAL : titre×0.3 + bullets×0.3 + keywords×0.2 + images×0.2 (si keywords vide : titre×0.375 + bullets×0.375 + images×0.25).
Chaque string de issues/suggestions : <80 chars, actionnable, en français."""

_IMPROVE_SYSTEM = """Tu es un expert Amazon SEO. Améliore le champ demandé d'une fiche Amazon.
Retourne UNIQUEMENT un objet JSON valide, sans texte avant ni après :
- Pour "title"    → {"improved": "<titre optimisé>", "explanation": "<1 phrase>"}
- Pour "bullets"  → {"improved": ["bullet 1","bullet 2","bullet 3","bullet 4","bullet 5"], "explanation": "<1 phrase>"}
- Pour "keywords" → {"improved": "<keywords ≤249 bytes, séparés par espaces>", "explanation": "<1 phrase>"}

Règles :
- Titre : ≤75 chars (nouvelle règle Amazon 2026), mots-clés principaux en tête, Marque + Type produit + Attribut clé. Ne JAMAIS dépasser 75 caractères.
- Bullets : 5 bullets, 150-200 chars chacun, commencer par un bénéfice client, pas de point final
- Keywords : ≤249 bytes UTF-8, pas de doublons avec le titre, pluriels/synonymes/longue traîne"""


def _scrape_amazon_for_audit(html_text: str) -> dict:
    """Extract title, bullets and image URLs from Amazon product page static HTML."""
    import re as _re, html as _html
    def _clean(t): return _re.sub(r'<[^>]+>', '', _html.unescape(t or '')).strip()

    result: dict = {"title": "", "bullets": [], "image_urls": []}

    # ── Title ──────────────────────────────────────────────────────────────────
    # Primary: id="productTitle"
    m = _re.search(r'<span[^>]+id=["\']productTitle["\'][^>]*>(.*?)</span>', html_text, _re.DOTALL | _re.IGNORECASE)
    if m:
        result["title"] = _clean(m.group(1))

    # Fallback: og:title
    if not result["title"]:
        for pat in (
            r'<meta[^>]+property=["\']og:title["\'][^>]+content=["\']([^"\']+)',
            r'<meta[^>]+content=["\']([^"\']+)["\'][^>]+property=["\']og:title["\']',
        ):
            mm = _re.search(pat, html_text, _re.IGNORECASE)
            if mm:
                result["title"] = _html.unescape(mm.group(1).strip())
                break

    # ── Bullets ────────────────────────────────────────────────────────────────
    # Strategy: find the feature-bullets section first, then extract list items
    feature_section = ""
    fs = _re.search(
        r'id=["\']feature-bullets["\'][^>]*>(.*?)(?:id=["\'](?:productDescription|aplus)["\']|</div>\s*</div>\s*</div>)',
        html_text, _re.DOTALL | _re.IGNORECASE,
    )
    if fs:
        feature_section = fs.group(1)

    search_area = feature_section or html_text
    bullet_matches = _re.findall(
        r'<span[^>]+class=["\'][^"\']*a-list-item[^"\']*["\'][^>]*>(.*?)</span>',
        search_area, _re.DOTALL | _re.IGNORECASE,
    )
    seen_b: set = set()
    for raw in bullet_matches:
        b = _clean(raw)
        # Skip very short strings or strings that look like navigation / UI labels
        if len(b) > 20 and b not in seen_b and not b.startswith(("Voir ", "En savoir", "See more", "Read more")):
            seen_b.add(b)
            result["bullets"].append(b)
        if len(result["bullets"]) >= 5:
            break

    # ── Images ─────────────────────────────────────────────────────────────────
    # Primary: colorImages JS object (hiRes > large > main)
    for key in ("hiRes", "large", "main"):
        img_urls = _re.findall(rf'"{key}"\s*:\s*"(https://[^"]+\.(?:jpg|jpeg|png)[^"]*)"', html_text, _re.IGNORECASE)
        if img_urls:
            seen_u: set = set()
            for u in img_urls:
                base = _re.sub(r'\._[A-Z0-9_,]+_\.', '.', u)  # strip size suffix
                if base not in seen_u:
                    seen_u.add(base)
                    result["image_urls"].append(base)
                if len(result["image_urls"]) >= 9:
                    break
            break

    # Fallback: data-old-hires attribute
    if not result["image_urls"]:
        for pat in (
            r'data-old-hires=["\']([^"\']+)["\']',
            r'"landingAsinColor"[^}]*?"large"\s*:\s*"(https://[^"]+)"',
        ):
            mm = _re.search(pat, html_text)
            if mm:
                result["image_urls"] = [mm.group(1)]
                break

    return result


class AuditRequest(BaseModel):
    title: str = ""
    bullets: list[str] = []
    backend_keywords: str = ""
    description: str = ""
    image_count: int = 0
    image_urls: list[str] = []
    marketplace: str = "amazon_fr"


class AuditImproveRequest(BaseModel):
    field: str
    title: str = ""
    bullets: list[str] = []
    backend_keywords: str = ""
    marketplace: str = "amazon_fr"


_AMAZON_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
    "Accept-Language": "fr-FR,fr;q=0.9,en-US;q=0.8,en;q=0.7",
    "Accept-Encoding": "gzip, deflate, br",
    "Cache-Control": "no-cache",
    "Pragma": "no-cache",
    "Sec-Fetch-Dest": "document",
    "Sec-Fetch-Mode": "navigate",
    "Sec-Fetch-Site": "none",
    "Upgrade-Insecure-Requests": "1",
}


@app.post("/api/audit/prefill")
async def audit_prefill(request: Request):
    """Scrape an Amazon product URL, OR parse raw HTML if html_content is provided."""
    body = await request.json()
    html_content = (body.get("html_content") or "").strip()

    # Mode 2: user pasted raw HTML — parse directly, no HTTP fetch
    if html_content:
        data = _scrape_amazon_for_audit(html_content)
        if not data["title"] and not data["bullets"] and not data["image_urls"]:
            raise HTTPException(422, "Aucune donnée trouvée dans ce HTML. Vérifiez que vous avez bien copié le code source complet (Ctrl+U).")
        return {**data, "source": "html_paste"}

    # Mode 1: fetch from URL
    url = (body.get("url") or "").strip()
    if not url.startswith("http"):
        raise HTTPException(400, "URL invalide")

    html_text = None
    last_err = ""
    # Try main URL then mobile variant
    candidates = [url]
    if "amazon." in url:
        mob = url.replace("www.amazon.", "www.amazon.").replace("//amazon.", "//www.amazon.")
        # Try with /dp/ short form if it's a long URL
        import re as _re2
        dp_m = _re2.search(r'/dp/([A-Z0-9]{10})', url)
        if dp_m:
            base = _re2.match(r'(https?://[^/]+)', url)
            if base:
                candidates.insert(0, f"{base.group(1)}/dp/{dp_m.group(1)}")

    for candidate_url in candidates:
        try:
            async with httpx.AsyncClient(
                follow_redirects=True, timeout=12.0, headers=_AMAZON_HEADERS,
            ) as client:
                resp = await client.get(candidate_url)
                resp.raise_for_status()
                html_text = resp.text
                break
        except httpx.TimeoutException:
            last_err = "Délai dépassé"
        except httpx.HTTPStatusError as e:
            last_err = f"Erreur HTTP {e.response.status_code}"
        except Exception as e:
            last_err = str(e)

    if not html_text:
        raise HTTPException(502, f"Impossible d'accéder à l'URL ({last_err}). Utilisez le fallback 'Coller le source HTML'.")

    # Detect bot-detection / CAPTCHA
    bot_signals = [
        "api.security-challenge.aws.com",
        "Enter the characters you see below",
        "robot or an automated system",
        "captcha",
        "px-captcha",
    ]
    if any(s.lower() in html_text.lower() for s in bot_signals):
        raise HTTPException(
            403,
            "BLOCKED",  # frontend detects this code to show the fallback UI
        )

    data = _scrape_amazon_for_audit(html_text)
    if not data["title"] and not data["bullets"]:
        raise HTTPException(
            422,
            "BLOCKED",  # same fallback — page likely JS-rendered
        )
    return {**data, "source": "url_scrape"}


@app.post("/api/audit")
async def audit_listing(req: AuditRequest, request: Request):
    if not os.getenv("ANTHROPIC_API_KEY"):
        raise HTTPException(503, "Service indisponible")
    from app.services.ai_agent import get_client
    import json as _json

    lang = "français" if "fr" in req.marketplace else ("anglais" if "uk" in req.marketplace else "langue locale du marché")
    bullets_text = "\n".join(f"• {b}" for b in req.bullets if b.strip()) or "(aucun bullet)"
    kw_bytes = len(req.backend_keywords.encode("utf-8"))
    # Effective image count: prefer URL list length, fallback to slider
    effective_img_count = len(req.image_urls) if req.image_urls else req.image_count

    text_prompt = f"""Marketplace : {req.marketplace} (langue fiches : {lang})

TITRE ({len(req.title)} chars) :
{req.title or "(vide)"}

BULLETS ({len([b for b in req.bullets if b.strip()])}/5) :
{bullets_text}

KEYWORDS BACKEND ({kw_bytes}/249 bytes) :
{req.backend_keywords or "(vide)"}

DESCRIPTION ({len(req.description)} chars) :
{req.description[:1000] or "(vide)"}

IMAGES : {effective_img_count} image(s)"""

    # Build message content — download images as base64 for vision (max 4)
    urls_to_analyze = [u for u in req.image_urls if u.startswith("http")][:4]
    vision_used = False

    if urls_to_analyze:
        img_blocks: list = []
        async with httpx.AsyncClient(timeout=8.0) as img_client:
            for img_url in urls_to_analyze:
                try:
                    ir = await img_client.get(img_url, headers={"Referer": "https://www.amazon.fr/"})
                    if ir.status_code == 200:
                        ct = ir.headers.get("content-type", "image/jpeg").split(";")[0].strip()
                        if ct not in ("image/jpeg", "image/png", "image/gif", "image/webp"):
                            ct = "image/jpeg"
                        img_blocks.append({
                            "type": "image",
                            "source": {
                                "type": "base64",
                                "media_type": ct,
                                "data": base64.b64encode(ir.content).decode(),
                            },
                        })
                except Exception:
                    pass  # skip unreachable images silently

        if img_blocks:
            vision_used = True
            content: list = img_blocks + [{
                "type": "text",
                "text": text_prompt + "\n\nNOTE : Les images ci-dessus sont les vraies images produit Amazon. Analyse-les visuellement pour scorer la dimension IMAGES : fond blanc (image 1), produit bien visible, visuels lifestyle, infographie, angles multiples. Ajuste le score et les issues/suggestions en conséquence.",
            }]
        else:
            content = text_prompt
    else:
        content = text_prompt

    import re as _re_json
    resp = await get_client().messages.create(
        model="claude-haiku-4-5-20251001" if not vision_used else "claude-sonnet-4-6",
        max_tokens=1500,
        system=[{"type": "text", "text": _AUDIT_SYSTEM, "cache_control": {"type": "ephemeral"}}],
        messages=[{"role": "user", "content": content}],
    )
    raw = resp.content[0].text.strip()
    # Strip markdown code fences if Claude wraps its JSON
    raw = _re_json.sub(r'^```(?:json)?\s*', '', raw, flags=_re_json.MULTILINE)
    raw = _re_json.sub(r'\s*```$', '', raw, flags=_re_json.MULTILINE).strip()
    try:
        result = _json.loads(raw)
        result["vision_used"] = vision_used
        return result
    except Exception:
        raise HTTPException(500, "Erreur d'analyse — réessayez")


# ── Performance tracking ──────────────────────────────────────────────────────

class TrackedListingRequest(BaseModel):
    sku: str = ""
    asin: str = ""
    title: str = ""
    marketplace: str = "amazon_fr"
    published_at: str = ""
    seo_score: int = 0


class SnapshotRequest(BaseModel):
    snapshot_date: str
    period_label: str = ""
    sessions: int = 0
    page_views: int = 0
    units_ordered: int = 0
    conversion_rate: float = 0.0
    revenue: float = 0.0
    keyword: str = ""
    keyword_rank: Optional[int] = None
    notes: str = ""


@app.get("/api/tracking")
async def tracking_list(request: Request):
    return list_tracked_listings(request.state.user_email)


@app.get("/api/tracking/summary")
async def tracking_summary(request: Request):
    return get_tracking_summary(request.state.user_email)


@app.post("/api/tracking")
async def tracking_add(req: TrackedListingRequest, request: Request):
    if not req.asin and not req.sku:
        raise HTTPException(400, "ASIN ou SKU requis")
    return add_tracked_listing(
        request.state.user_email, req.sku, req.asin, req.title,
        req.marketplace, req.published_at, req.seo_score,
    )


@app.delete("/api/tracking/{listing_id}")
async def tracking_delete(listing_id: str, request: Request):
    if not delete_tracked_listing(listing_id, request.state.user_email):
        raise HTTPException(404, "Fiche introuvable")
    return {"ok": True}


@app.get("/api/tracking/{listing_id}/snapshots")
async def tracking_snapshots(listing_id: str, request: Request):
    return get_snapshots(listing_id, request.state.user_email)


@app.post("/api/tracking/{listing_id}/snapshots")
async def tracking_add_snapshot(listing_id: str, req: SnapshotRequest, request: Request):
    return add_snapshot(
        listing_id, request.state.user_email, req.snapshot_date, req.period_label,
        req.sessions, req.page_views, req.units_ordered, req.conversion_rate,
        req.revenue, req.keyword, req.keyword_rank, req.notes,
    )


@app.delete("/api/tracking/{listing_id}/snapshots/{snap_id}")
async def tracking_delete_snapshot(listing_id: str, snap_id: int, request: Request):
    if not delete_snapshot(snap_id, request.state.user_email):
        raise HTTPException(404, "Snapshot introuvable")
    return {"ok": True}


@app.post("/api/tracking/{listing_id}/import-csv")
async def tracking_import_csv(listing_id: str, request: Request):
    """Import Amazon Business Report CSV (Detail Page Sales and Traffic By Date)."""
    import csv as _csv, io as _io
    body = await request.json()
    csv_text = body.get("csv_text", "")
    if not csv_text:
        raise HTTPException(400, "CSV vide")
    email = request.state.user_email

    reader = _csv.DictReader(_io.StringIO(csv_text))
    # Normalise header names: lowercase + strip + normalize dashes and quotes
    def _norm(k):
        return (k.lower().strip()
                .replace("﻿", "")          # BOM
                .replace("–", "-")          # en-dash → hyphen
                .replace("—", "-")          # em-dash → hyphen
                .replace("’", "'")          # right single quote → apostrophe
                .replace("‘", "'")          # left single quote → apostrophe
                .replace(" ", " ")          # non-breaking space → space
                )
    rows = [{_norm(k): v.strip() for k, v in row.items()} for row in reader]

    # Column name aliases
    def _pick(row, *keys):
        for k in keys:
            v = row.get(k, "")
            if v and v != "-": return v
        return ""

    import datetime as _dt

    imported = 0
    for row in rows:
        # Date column: present in "by date" reports, absent in "by ASIN" reports
        date_val = _pick(row, "date", "day", "jour", "période", "period")
        if date_val:
            snap_date = None
            for fmt in ("%m/%d/%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
                try:
                    snap_date = str(_dt.datetime.strptime(date_val, fmt).date())
                    break
                except ValueError:
                    continue
            if not snap_date:
                continue
        else:
            # ASIN-level report: no date column — use today as snapshot date
            snap_date = str(_dt.date.today())

        sessions   = int(float(_pick(row,
            "sessions", "sessions - total") or 0))
        page_views = int(float(_pick(row,
            "page views", "page_views",
            "vues de la page - total",
            "vues de page") or 0))
        units      = int(float(_pick(row,
            "units ordered", "units_ordered",
            "unites commandees", "unites commandees - total",
            "unites commandees - total") or 0))
        conv_str   = _pick(row,
            "unit session percentage", "taux de conversion", "conversion rate",
            "pourcentage de session d'unite",
            "pourcentage par session d'unites",
            "pourcentage de session d'unite - total",
            "taux d'acquisition des commandes - total",
            "pourcentage de sessions d'unites - total") or ""
        if conv_str:
            conv = float(conv_str.replace("%", "").replace(",", ".").replace(" ", "").strip() or "0")
        elif sessions > 0 and units > 0:
            conv = round(units / sessions * 100, 2)
        else:
            conv = 0.0
        rev_str    = _pick(row,
            "ordered product sales", "revenue",
            "ventes de produits commandes",
            "chiffre d'affaires des produits commandes - total",
            "chiffre d'affaires des produits commandes") or "0"
        def _parse_eur(s):
            s = s.replace("€","").replace("$","").replace("\xa0","").replace(" ","").replace("\u00a0","").replace(" ","").strip()
            if "," in s and "." in s:
                s = s.replace(".", "").replace(",", ".")
            else:
                s = s.replace(",", ".")
            try: return float(s)
            except: return 0.0
        revenue = _parse_eur(rev_str)

        # Skip rows with no meaningful data
        if sessions == 0 and units == 0 and page_views == 0:
            continue

        add_snapshot(listing_id, email, snap_date, "", sessions, page_views,
                     units, conv, revenue, "", None, "Import CSV")
        imported += 1

    return {"imported": imported}


@app.post("/api/audit/improve")
async def audit_improve(req: AuditImproveRequest, request: Request):
    if not os.getenv("ANTHROPIC_API_KEY"):
        raise HTTPException(503, "Service indisponible")
    if req.field not in ("title", "bullets", "keywords"):
        raise HTTPException(400, "Champ invalide")
    from app.services.ai_agent import get_client
    import json as _json
    lang = "français" if "fr" in req.marketplace else ("anglais" if "uk" in req.marketplace else "langue locale du marché")
    bullets_text = "\n".join(f"• {b}" for b in req.bullets if b.strip()) or "(aucun)"
    prompt = f"""Marketplace : {req.marketplace} (langue : {lang})
Champ à améliorer : {req.field}

Fiche actuelle :
TITRE : {req.title or "(vide)"}
BULLETS :
{bullets_text}
KEYWORDS : {req.backend_keywords or "(vide)"}"""

    resp = await get_client().messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=1200,
        system=[{"type": "text", "text": _IMPROVE_SYSTEM, "cache_control": {"type": "ephemeral"}}],
        messages=[{"role": "user", "content": prompt}],
    )
    raw = resp.content[0].text.strip()
    import re as _re_imp
    raw = _re_imp.sub(r'^```(?:json)?\s*', '', raw, flags=_re_imp.MULTILINE)
    raw = _re_imp.sub(r'\s*```$', '', raw, flags=_re_imp.MULTILINE).strip()
    try:
        return _json.loads(raw)
    except Exception:
        raise HTTPException(500, "Erreur d'amélioration — réessayez")

# ── Title reformat (Amazon 2026-07-27 ≤75-char migration) ────────────────────

class ReformatTitleRequest(BaseModel):
    title: str
    brand: str = ""
    category: str = ""
    bullets: list[str] = []
    item_highlights: str = ""
    backend_keywords: str = ""
    marketplace: str = "amazon_fr"


_REFORMAT_SYSTEM = """Tu es un expert en conformité des fiches Amazon.
Nouvelle politique Amazon (27/07/2026) : le titre doit faire 75 caractères MAXIMUM (espaces compris),
et un nouveau champ "Item Highlights" de 125 caractères MAXIMUM (texte continu, sans puces, indexable)
récupère les bénéfices et qualificatifs retirés du titre raccourci.

À partir de la fiche fournie, tu produis :
1. Un titre conforme ≤ 75 caractères : Marque → Type produit → Attribut clé → Couleur → Taille. Mot-clé
   principal en tête. Aucun caractère interdit (! $ ? _ {} ^), jamais tout en majuscules, pas de superlatif.
2. Un champ item_highlights ≤ 125 caractères : texte fluide, 1-3 phrases, bénéfices + mots-clés secondaires
   issus de l'ancien titre/bullets, sans puces, sans prix ni promotion.

Conserve la langue d'origine de la fiche. Compte les caractères avant de répondre.
Réponds UNIQUEMENT en JSON valide : {"title": "...", "item_highlights": "..."}"""


async def _reformat_title_core(
    title: str, brand: str = "", category: str = "", bullets: list = None,
    item_highlights: str = "", backend_keywords: str = "", marketplace: str = "amazon_fr",
) -> dict:
    """Shorten a title to ≤75 chars + produce Item Highlights via Claude.
    Returns {title, item_highlights, seo_score[, media_exempt]}. Raises on AI failure."""
    from app.services.ai_agent import (
        get_client, is_media_category, MEDIA_TITLE_MAX,
        constraints_for_category, MARKETPLACE_CONSTRAINTS, _compute_seo_score,
    )
    from app.models import Marketplace
    import json as _json
    import re as _re_rf

    bullets = bullets or []
    try:
        mkt = Marketplace(marketplace)
    except ValueError:
        mkt = Marketplace.AMAZON_FR
    base_constraints = MARKETPLACE_CONSTRAINTS.get(mkt, MARKETPLACE_CONSTRAINTS[Marketplace.AMAZON_FR])
    constraints = constraints_for_category(base_constraints, category)

    def _scored(t: str, h: str, **extra) -> dict:
        seo = _compute_seo_score({
            "title": t, "item_highlights": h,
            "bullet_points": bullets,
            "backend_keywords": backend_keywords or "",
            "description": "",
        }, constraints)
        return {"title": t, "item_highlights": h, "seo_score": seo, **extra}

    # Media categories are exempt from the 75-char rule
    if is_media_category(category):
        return _scored(title[:MEDIA_TITLE_MAX], (item_highlights or "")[:125], media_exempt=True)

    # Already compliant with a filled highlights field → nothing to do
    if len(title) <= 75 and item_highlights:
        return _scored(title, item_highlights[:125])

    bullets_text = "\n".join(f"• {b}" for b in bullets if b and str(b).strip()) or "(aucun)"
    prompt = f"""Marketplace : {marketplace}
Marque : {brand or "(non précisée)"}
Catégorie : {category or "(non précisée)"}
Mots-clés backend : {(backend_keywords or "(vides)")[:300]}

TITRE ACTUEL ({len(title)} car.) : {title}
BULLETS :
{bullets_text}
ITEM HIGHLIGHTS ACTUEL : {item_highlights or "(vide)"}"""

    async def _call_once() -> Optional[dict]:
        resp = await get_client().messages.create(
            model="claude-sonnet-4-6",
            max_tokens=600,
            system=[{"type": "text", "text": _REFORMAT_SYSTEM, "cache_control": {"type": "ephemeral"}}],
            messages=[
                {"role": "user", "content": prompt},
                # Prefill the reply with "{" so the model is forced to emit JSON
                # (eliminates the prose/markdown replies that broke parsing before).
                {"role": "assistant", "content": "{"},
            ],
        )
        raw = ""
        for block in (resp.content or []):
            txt = getattr(block, "text", None)
            if txt and txt.strip():
                raw = txt.strip()
                break
        raw = "{" + raw  # re-add the prefilled opening brace
        raw = _re_rf.sub(r'^```(?:json)?\s*', '', raw, flags=_re_rf.MULTILINE)
        raw = _re_rf.sub(r'\s*```\s*$', '', raw, flags=_re_rf.MULTILINE).strip()
        try:
            return _json.loads(raw)
        except _json.JSONDecodeError:
            m = _re_rf.search(r'\{.*"title".*\}', raw, _re_rf.DOTALL)
            if m:
                try:
                    return _json.loads(m.group())
                except _json.JSONDecodeError:
                    return None
        return None

    # Retry transient API errors AND non-JSON replies — a single miss across a
    # 70-row catalogue is what produced crude truncations + empty highlights.
    data = None
    for attempt in range(3):
        try:
            data = await _call_once()
        except Exception as e:
            data = None
            log.warning("[reformat_title] attempt %d failed: %s", attempt + 1, str(e)[:160])
        if data and str(data.get("title", "")).strip():
            break
        if attempt < 2:
            await asyncio.sleep(0.6 * (attempt + 1))

    if not data or not str(data.get("title", "")).strip():
        # Last resort: clean word-boundary truncation + a highlight built from the
        # descriptive tail that no longer fits the title (never leave it empty).
        short = title[:75].rsplit(" ", 1)[0] if len(title) > 75 else title
        tail = _re_rf.sub(r"\s+", " ", title[len(short):]).strip(" -|,;:.")
        fallback_hl = (item_highlights or tail or title)[:125]
        return _scored(short, fallback_hl)

    # Enforce hard limits server-side regardless of model output
    return _scored(
        str(data.get("title", title))[:75],
        str(data.get("item_highlights", item_highlights or ""))[:125],
    )


@app.post("/api/reformat-title")
async def reformat_title(req: ReformatTitleRequest, request: Request):
    """Shorten an existing title to ≤75 chars and produce the Item Highlights field.
    Media categories keep the 200-char allowance (returned mostly unchanged)."""
    if not os.getenv("ANTHROPIC_API_KEY"):
        raise HTTPException(503, "Service indisponible")
    try:
        return await _reformat_title_core(
            title=req.title, brand=req.brand, category=req.category,
            bullets=req.bullets, item_highlights=req.item_highlights,
            backend_keywords=req.backend_keywords, marketplace=req.marketplace,
        )
    except HTTPException:
        raise
    except Exception:
        raise HTTPException(500, "Erreur de reformatage — réessayez")


# ── Catalog title migration (bulk, for existing Amazon listings) ─────────────

class TitleMigrationParseRequest(BaseModel):
    filename: str
    content_b64: str


# Column aliases found in Seller Central "All Listings" reports (FR/EN) and
# generic catalog exports.
_TM_SKU_COLS    = ("seller-sku", "sku", "item_sku", "item-sku", "sku vendeur", "seller_sku")
_TM_TITLE_COLS  = ("item-name", "item_name", "title", "titre", "product-name", "product_name",
                   "nom", "nom du produit", "item-description", "name")
_TM_ASIN_COLS   = ("asin1", "asin", "asin 1")
_TM_BRAND_COLS  = ("brand", "brand-name", "brand_name", "marque")
_TM_CAT_COLS    = ("product-type", "product_type", "category", "catégorie", "categorie", "item-type", "type")


def _tm_find_col(headers: list, candidates: tuple) -> Optional[int]:
    low = [str(h or "").strip().lower() for h in headers]
    # Exact match first
    for cand in candidates:
        if cand in low:
            return low.index(cand)
    # Fallback: tolerate punctuation/space variants (e.g. "item name", "item_name",
    # "nom de l'article") by normalising separators to '-'.
    def _norm(s: str) -> str:
        return re.sub(r"[\s_]+", "-", s.strip().lower()).strip("-")
    low_n = [_norm(h) for h in low]
    for cand in candidates:
        cn = _norm(cand)
        if cn in low_n:
            return low_n.index(cn)
    return None


def _tm_find_header_row(data_rows: list) -> int:
    """Amazon reports & category templates often carry metadata/label rows before
    the real header (e.g. row 1 = 'TemplateType=...', then technical field names a
    few rows down). Scan the first rows and return the index of the one that holds
    a recognisable title column — falling back to row 0."""
    for i, row in enumerate(data_rows[:15]):
        if _tm_find_col(list(row), _TM_TITLE_COLS) is not None:
            return i
    return 0


@app.post("/api/title-migration/parse")
async def title_migration_parse(payload: TitleMigrationParseRequest, request: Request):
    """Parse a Seller Central listings report (CSV/TSV/XLSX) or any file with
    SKU + title columns. Returns rows annotated with compliance status."""
    content = _decode_upload(payload.filename, payload.content_b64)
    fname = (payload.filename or "report.csv").lower()
    from app.services.ai_agent import is_media_category

    rows = []
    if fname.endswith((".xlsx", ".xlsm", ".xls")):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        data_rows = [[c if c is not None else "" for c in r] for r in ws.iter_rows(values_only=True)]
        wb.close()
    else:
        # Seller Central reports are TSV (.txt); also accept ; and , CSV
        text = None
        for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
            try:
                text = content.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        if text is None:
            raise HTTPException(400, "Encodage de fichier non reconnu")
        first_line = text.splitlines()[0] if text.splitlines() else ""
        delim = "\t" if "\t" in first_line else (";" if first_line.count(";") > first_line.count(",") else ",")
        import csv as _csv
        data_rows = list(_csv.reader(io.StringIO(text), delimiter=delim))

    if not data_rows:
        raise HTTPException(422, "Fichier vide")

    # Amazon files may carry metadata/label rows before the real header — locate it.
    header_idx = _tm_find_header_row(data_rows)
    headers = [str(h or "") for h in data_rows[header_idx]]
    sku_i   = _tm_find_col(headers, _TM_SKU_COLS)
    title_i = _tm_find_col(headers, _TM_TITLE_COLS)
    if title_i is None:
        preview = ", ".join(h for h in headers[:12] if h.strip()) or "(aucune colonne lisible)"
        raise HTTPException(422, "Colonne titre introuvable — le fichier doit contenir une colonne "
                                 "'item-name' (rapport Seller Central) ou 'titre'/'title'. "
                                 f"Colonnes détectées : {preview}")
    asin_i  = _tm_find_col(headers, _TM_ASIN_COLS)
    brand_i = _tm_find_col(headers, _TM_BRAND_COLS)
    cat_i   = _tm_find_col(headers, _TM_CAT_COLS)

    MAX_ROWS = 2000
    for r in data_rows[header_idx + 1:]:
        if len(rows) >= MAX_ROWS:
            break
        def _cell(i):
            return str(r[i]).strip() if i is not None and i < len(r) and r[i] is not None else ""
        title = _cell(title_i)
        if not title:
            continue
        sku = _cell(sku_i)
        category = _cell(cat_i)
        media = is_media_category(category)
        limit = 200 if media else 75
        rows.append({
            "sku": sku or _cell(asin_i) or f"ROW-{len(rows)+1}",
            "asin": _cell(asin_i),
            "brand": _cell(brand_i),
            "category": category,
            "title": title,
            "title_len": len(title),
            "media_exempt": media,
            "compliant": len(title) <= limit,
        })

    if not rows:
        raise HTTPException(422, "Aucune ligne exploitable trouvée dans le fichier")
    non_compliant = sum(1 for x in rows if not x["compliant"])
    return {"rows": rows, "total": len(rows), "non_compliant": non_compliant,
            "truncated": len(data_rows) - 1 > MAX_ROWS}


class TitleMigrationRow(BaseModel):
    sku: str
    asin: str = ""
    title: str
    brand: str = ""
    category: str = ""


class TitleMigrationRunRequest(BaseModel):
    rows: list[TitleMigrationRow]
    marketplace: str = "amazon_fr"


async def _run_title_migration_job(job_id: str, rows: list, marketplace: str, email: str):
    _jobs[job_id]["status"] = "running"
    results = []
    sem = asyncio.Semaphore(8)
    done = 0

    async def _one(row: TitleMigrationRow):
        nonlocal done
        async with sem:
            out = {"sku": row.sku, "asin": row.asin, "old_title": row.title,
                   "brand": row.brand, "category": row.category}
            try:
                res = await _reformat_title_core(
                    title=row.title, brand=row.brand, category=row.category,
                    marketplace=marketplace,
                )
                out.update(res)
                out["ok"] = True
            except Exception as e:
                out.update({"ok": False, "error": str(e)[:200],
                            "title": row.title, "item_highlights": ""})
            results.append(out)
            done += 1
            _jobs[job_id]["progress"] = done

    await asyncio.gather(*[_one(r) for r in rows])
    ok_count = sum(1 for r in results if r.get("ok"))
    if email:
        try:
            log_usage(email, "title_migration", ok_count)
        except Exception:
            pass
    # Preserve input order for the before/after table
    order = {r.sku: i for i, r in enumerate(rows)}
    results.sort(key=lambda x: order.get(x["sku"], 1_000_000))
    _jobs[job_id].update({"status": "done", "result": {"rows": results, "ok": ok_count,
                                                       "failed": len(results) - ok_count}})


@app.post("/api/title-migration/run")
async def title_migration_run(req: TitleMigrationRunRequest, request: Request):
    if not os.getenv("ANTHROPIC_API_KEY"):
        raise HTTPException(503, "Service indisponible")
    if not req.rows:
        raise HTTPException(400, "Aucune ligne à traiter")
    if len(req.rows) > 2000:
        raise HTTPException(413, "Maximum 2000 titres par migration — divisez le fichier")
    email = request.state.user_email
    job_id = str(uuid4())
    _jobs[job_id] = {"status": "pending", "progress": 0, "total": len(req.rows),
                     "created_at": time.time()}
    _cleanup_jobs()
    try:
        save_job(job_id, email, "title_migration", len(req.rows))
    except Exception:
        pass
    asyncio.create_task(_run_title_migration_job(job_id, req.rows, req.marketplace, email))
    return {"job_id": job_id, "total": len(req.rows)}


class TitleMigrationExportRequest(BaseModel):
    rows: list[dict]  # [{sku, title, item_highlights}]


@app.post("/api/title-migration/export")
async def title_migration_export(req: TitleMigrationExportRequest, request: Request):
    """Build a Seller Central partial-update inventory file with the new titles,
    ready to upload in 'Ajouter des produits via un fichier'.

    Built from a REAL Amazon template (modern feedType=256 structure), NOT a
    hand-crafted fptcustom file. The legacy 'TemplateType=fptcustom' flat file is
    deprecated and was rejected every time (90012 'sku absent' / 90061 'invalid
    header') because the modern format identifies the SKU via
    'contribution_sku#1.value', the title via 'item_name#1.value' and the update
    mode via '::record_action' — names that can't be reproduced by hand because
    the settings/signature row is required. We load Amazon's own template, clear
    the example row and write one partial-update row per SKU (title only),
    leaving product_type empty so Amazon patches the existing listing by SKU
    across all categories. Updates both the title (item_name) and the new Item
    Highlights field ("Point fort de l'article" = title_differentiation)."""
    import openpyxl as _openpyxl
    from app.utils.amazon_template_filler import (
        _parse_template_settings, _find_template_worksheet, _build_col_index,
    )

    # Load a real Amazon category template (has item_name + contribution_sku).
    tpl_bytes = None
    try:
        _t = get_amazon_template_by_type("ANIMAL_COLLAR")
        if _t and _t.get("data_b64"):
            tpl_bytes = base64.b64decode(_t["data_b64"])
    except Exception:
        pass
    if not tpl_bytes:
        for _p in ("data/amazon_templates/ANIMAL_COLLAR.xlsm",
                   os.path.join(os.getcwd(), "data", "amazon_templates", "ANIMAL_COLLAR.xlsm")):
            try:
                with open(_p, "rb") as _f:
                    tpl_bytes = _f.read()
                break
            except Exception:
                continue
    if not tpl_bytes:
        raise HTTPException(503, "Modèle Amazon indisponible pour la migration des titres.")

    wb = _openpyxl.load_workbook(io.BytesIO(tpl_bytes), keep_vba=True, data_only=False)
    settings, settings_ws = _parse_template_settings(wb)
    attr_row = settings["attributeRow"]
    data_row = settings["dataRow"]
    ws = settings_ws or _find_template_worksheet(wb, attr_row)
    if ws is None:
        raise HTTPException(503, "Structure du modèle Amazon non reconnue.")
    col_index = _build_col_index(ws, attr_row)
    sku_col = col_index.get("contribution_sku#1.value")
    name_col = col_index.get("item_name#1.value")
    ra_col = col_index.get("::record_action")
    ptype_col = col_index.get("product_type#1.value")
    # Item Highlights (new July-2026 norm) = "Point fort de l'article" in the
    # template, attribute title_differentiation#1.value.
    hl_col = col_index.get("title_differentiation#1.value")
    if not (sku_col and name_col):
        raise HTTPException(503, "Colonnes SKU/titre introuvables dans le modèle Amazon.")
    # product_type#1.value is REQUIRED even for a partial update (error 90041).
    # Default to the template's own type; honour a real Amazon product-type CODE
    # from the source report (e.g. PET_TAG) when the row carries one.
    default_pt = (settings.get("productType") or "ANIMAL_COLLAR").upper()

    # Clear any example/demo rows Amazon ships from data_row downward.
    for r in range(data_row, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).value = None

    count = 0
    for r in req.rows:
        sku = str(r.get("sku", "")).strip()
        title = str(r.get("title", "")).strip()
        if not sku or not title:
            continue
        row_num = data_row + count
        ws.cell(row_num, sku_col).value = sku
        ws.cell(row_num, name_col).value = title[:200]
        highlights = str(r.get("item_highlights", "")).strip()
        if hl_col and highlights:
            ws.cell(row_num, hl_col).value = highlights[:125]
        if ra_col:
            ws.cell(row_num, ra_col).value = "partial_update"
        if ptype_col:
            row_pt = str(r.get("product_type", "")).strip()
            ws.cell(row_num, ptype_col).value = (
                row_pt if re.fullmatch(r"[A-Z][A-Z0-9_]+", row_pt) else default_pt
            )
        count += 1
    if not count:
        raise HTTPException(400, "Aucune ligne valide à exporter")

    _buf = io.BytesIO()
    wb.save(_buf)
    return Response(
        content=_buf.getvalue(),
        media_type="application/vnd.ms-excel.sheet.macroEnabled.12",
        headers={"Content-Disposition": 'attachment; filename="synqio_migration_titres_amazon.xlsm"'},
    )


# ── Saved sessions ────────────────────────────────────────────────────────────

class SaveSessionRequest(BaseModel):
    name: str
    data: dict


def _strip_base64_images(obj, _depth=0):
    """Recursively remove base64 data URLs from session data to keep payload small."""
    if _depth > 10:
        return obj
    if isinstance(obj, str):
        return "" if obj.startswith("data:image") else obj
    if isinstance(obj, list):
        return [_strip_base64_images(i, _depth + 1) for i in obj]
    if isinstance(obj, dict):
        return {k: _strip_base64_images(v, _depth + 1) for k, v in obj.items()}
    return obj


@app.post("/api/sessions/save")
async def api_save_session(req: SaveSessionRequest, request: Request):
    email = request.state.user_email
    session_id = str(uuid.uuid4())
    try:
        clean_data = _strip_base64_images(req.data)
        save_session(session_id, email, req.name.strip() or "Session sans nom", clean_data)
    except Exception as e:
        log.error(f"save_session failed: {type(e).__name__}: {e}", extra={"email": email})
        raise HTTPException(500, f"{type(e).__name__}: {str(e)[:300]}")
    return {"ok": True, "id": session_id}


@app.get("/api/sessions")
async def api_list_sessions(request: Request):
    try:
        return list_saved_sessions(request.state.user_email)
    except Exception as e:
        log.error(f"list_sessions failed: {type(e).__name__}: {e}", extra={"email": request.state.user_email})
        raise HTTPException(500, f"{type(e).__name__}: {str(e)[:300]}")


@app.get("/api/sessions/{session_id}")
async def api_get_session(session_id: str, request: Request):
    sess = get_saved_session(session_id, request.state.user_email)
    if not sess:
        raise HTTPException(404, "Session introuvable")
    return sess


@app.delete("/api/sessions/{session_id}")
async def api_delete_session(session_id: str, request: Request):
    if not delete_saved_session(session_id, request.state.user_email):
        raise HTTPException(404, "Session introuvable")
    return {"ok": True}


# ── Agency workspaces (multi-client) ─────────────────────────────────────────

class WorkspaceCreate(BaseModel):
    name: str


class WorkspaceRename(BaseModel):
    name: str


@app.get("/api/workspaces")
async def workspaces_list(request: Request):
    """List the agency's client workspaces. Uses the owner account, so it returns
    the same list whatever workspace is currently active."""
    from app.db import list_workspaces
    return {"workspaces": list_workspaces(request.state.owner_email)}


def _require_agency(email: str):
    from app.db import get_db
    conn = get_db()
    row = conn.execute("SELECT is_admin, agency_enabled FROM users WHERE email = ?", (email,)).fetchone()
    conn.close()
    if not (row and (row["is_admin"] or row["agency_enabled"])):
        raise HTTPException(403, "Mode agence non activé pour ce compte")


@app.post("/api/workspaces")
async def workspaces_create(body: WorkspaceCreate, request: Request):
    from app.db import list_workspaces, create_workspace
    owner = request.state.owner_email
    _require_agency(owner)
    name = (body.name or "").strip()[:60]
    if not name:
        raise HTTPException(400, "Nom d'espace requis")
    if len(list_workspaces(owner)) >= 50:
        raise HTTPException(400, "Maximum 50 espaces client")
    ws_id = f"ws_{uuid4().hex}"
    create_workspace(ws_id, owner, name)
    return {"id": ws_id, "name": name}


@app.patch("/api/workspaces/{ws_id}")
async def workspaces_rename(ws_id: str, body: WorkspaceRename, request: Request):
    from app.db import rename_workspace
    name = (body.name or "").strip()[:60]
    if not name:
        raise HTTPException(400, "Nom requis")
    if not rename_workspace(ws_id, request.state.owner_email, name):
        raise HTTPException(404, "Espace introuvable")
    return {"ok": True, "name": name}


@app.delete("/api/workspaces/{ws_id}")
async def workspaces_delete(ws_id: str, request: Request):
    from app.db import delete_workspace
    if not delete_workspace(ws_id, request.state.owner_email):
        raise HTTPException(404, "Espace introuvable")
    return {"deleted": True}


# ── Public API keys & webhooks management (session-authed, feeds the UI) ─────

class ApiKeyCreateRequest(BaseModel):
    label: str = ""


@app.get("/api/apikeys")
async def apikeys_list(request: Request):
    from app.db import list_api_keys
    return {"keys": list_api_keys(request.state.user_email)}


@app.post("/api/apikeys")
async def apikeys_create(body: ApiKeyCreateRequest, request: Request):
    import hashlib as _hl
    from app.db import list_api_keys, create_api_key
    email = request.state.user_email
    active = [k for k in list_api_keys(email) if not k.get("revoked")]
    if len(active) >= 5:
        raise HTTPException(400, "Maximum 5 clés actives — révoquez-en une d'abord")
    key = f"sk_synqio_{uuid4().hex}{uuid4().hex[:8]}"
    key_id = f"key_{uuid4().hex[:12]}"
    create_api_key(key_id, email, _hl.sha256(key.encode()).hexdigest(),
                   prefix=key[:16] + "…", label=(body.label or "").strip()[:60])
    # Plaintext returned ONCE — only the hash is stored
    return {"id": key_id, "key": key, "label": body.label}


@app.delete("/api/apikeys/{key_id}")
async def apikeys_revoke(key_id: str, request: Request):
    from app.db import revoke_api_key
    if not revoke_api_key(key_id, request.state.user_email):
        raise HTTPException(404, "Clé introuvable")
    return {"revoked": True}


class WebhookCreateRequest(BaseModel):
    url: str
    events: list[str] = ["listing.generated"]


@app.get("/api/webhooks")
async def webhooks_list(request: Request):
    from app.db import list_webhooks
    return {"webhooks": list_webhooks(request.state.user_email)}


@app.post("/api/webhooks")
async def webhooks_create(body: WebhookCreateRequest, request: Request):
    import secrets as _sec
    from app.db import list_webhooks, create_webhook
    email = request.state.user_email
    url = (body.url or "").strip()
    if not url.startswith("https://"):
        raise HTTPException(400, "L'URL du webhook doit être en https://")
    if len(list_webhooks(email)) >= 10:
        raise HTTPException(400, "Maximum 10 webhooks par compte")
    events = [e for e in body.events if e == "listing.generated"] or ["listing.generated"]
    hook_id = f"wh_{uuid4().hex[:12]}"
    secret = f"whsec_{_sec.token_urlsafe(24)}"
    create_webhook(hook_id, email, url, secret, ",".join(events))
    return {"id": hook_id, "url": url, "events": events, "secret": secret}


@app.delete("/api/webhooks/{hook_id}")
async def webhooks_delete(hook_id: str, request: Request):
    from app.db import delete_webhook
    if not delete_webhook(hook_id, request.state.user_email):
        raise HTTPException(404, "Webhook introuvable")
    return {"deleted": True}


def _feed_urls(token: str) -> dict:
    from app.utils.export_channels import CHANNELS
    base = os.getenv("PUBLIC_BASE_URL", "https://synqio.io").rstrip("/")
    return {ch: f"{base}/api/feed/{token}/{ch}" for ch in CHANNELS}


@app.get("/api/feedtoken")
async def feedtoken_get(request: Request):
    """Return the user's feed token + per-channel URLs (create it on first call)."""
    import secrets as _sec
    from app.db import get_feed_token, set_feed_token
    email = request.state.user_email
    row = get_feed_token(email)
    if not row:
        token = f"feed_{_sec.token_urlsafe(24)}"
        set_feed_token(email, token)
        row = {"token": token, "last_fetched_at": None, "last_channel": "", "fetch_count": 0}
    return {**row, "urls": _feed_urls(row["token"])}


@app.post("/api/feedtoken/regenerate")
async def feedtoken_regenerate(request: Request):
    """Mint a new token — the old feed URLs stop working immediately."""
    import secrets as _sec
    from app.db import set_feed_token
    email = request.state.user_email
    token = f"feed_{_sec.token_urlsafe(24)}"
    set_feed_token(email, token)
    return {"token": token, "urls": _feed_urls(token),
            "last_fetched_at": None, "last_channel": "", "fetch_count": 0}


# ── Native push connectors (Shopify, WooCommerce) ────────────────────────────

def _connector_config(email: str, platform: str) -> dict:
    from app.db import get_connector
    from app.services.crypto import decrypt_token
    row = get_connector(email, platform)
    if not row:
        raise HTTPException(404, f"Connecteur {platform} non configuré")
    try:
        return json.loads(decrypt_token(row["config_enc"]) or "{}")
    except Exception:
        raise HTTPException(500, "Configuration du connecteur illisible — reconfigurez-le")


@app.get("/api/connectors")
async def connectors_list(request: Request):
    from app.db import list_connectors
    from app.services.connectors import PLATFORMS
    configured = {c["platform"]: c for c in list_connectors(request.state.user_email)}
    return {"platforms": [
        {"platform": p, "label": spec["label"], "fields": spec["fields"],
         "configured": p in configured,
         "last_push_at": str(configured[p]["last_push_at"]) if p in configured and configured[p]["last_push_at"] else None,
         "last_status": configured[p]["last_status"] if p in configured else ""}
        for p, spec in PLATFORMS.items()
    ]}


@app.post("/api/connectors/{platform}")
async def connectors_save(platform: str, request: Request):
    """Validate credentials against the platform, then store them encrypted."""
    from app.db import set_connector
    from app.services.connectors import PLATFORMS
    from app.services.crypto import encrypt_token
    spec = PLATFORMS.get(platform)
    if not spec:
        raise HTTPException(404, f"Plateforme inconnue : {platform}")
    body = await request.json()
    cfg = {f: str(body.get(f, "")).strip() for f in spec["fields"]}
    if not all(cfg.values()):
        raise HTTPException(400, f"Champs requis : {', '.join(spec['fields'])}")
    try:
        info = await spec["test"](cfg)
    except ValueError as e:
        raise HTTPException(400, str(e))
    except Exception:
        raise HTTPException(502, "Connexion à la plateforme impossible — vérifiez l'URL")
    set_connector(request.state.user_email, platform,
                  encrypt_token(json.dumps(cfg, ensure_ascii=False)))
    return {"ok": True, "platform": platform, "info": info}


@app.delete("/api/connectors/{platform}")
async def connectors_delete(platform: str, request: Request):
    from app.db import delete_connector
    if not delete_connector(request.state.user_email, platform):
        raise HTTPException(404, "Connecteur introuvable")
    return {"deleted": True}


@app.post("/api/connectors/{platform}/push")
async def connectors_push(platform: str, listings: List[AmazonListing], request: Request):
    """Push listings into the client's store (draft status, create-or-update)."""
    from app.db import update_connector_status
    from app.services.connectors import PLATFORMS, MAX_PUSH
    spec = PLATFORMS.get(platform)
    if not spec:
        raise HTTPException(404, f"Plateforme inconnue : {platform}")
    if not listings:
        raise HTTPException(400, "Aucune fiche à publier")
    email = request.state.user_email
    if not _rate_limit(f"{email}:connector-push", limit=10, window=3600):
        raise HTTPException(429, "Limite de 10 publications/heure — réessayez plus tard")
    cfg = _connector_config(email, platform)
    image_urls = _get_image_urls_for_skus([l.sku for l in listings])
    results = await spec["push"](cfg, listings, image_urls=image_urls)
    created = sum(1 for r in results if r["action"] == "created")
    updated = sum(1 for r in results if r["action"] == "updated")
    errors = sum(1 for r in results if r["action"] == "error")
    update_connector_status(email, platform, f"{created}+{updated}ok/{errors}err")
    try:
        from app.db import log_pushes
        ok_skus = [r.get("sku") for r in results if r.get("action") in ("created", "updated")]
        log_pushes(email, platform, ok_skus, status="connector")
    except Exception:
        pass
    return {"created": created, "updated": updated, "errors": errors,
            "skipped": max(0, len(listings) - len(results)),
            "max_per_push": MAX_PUSH, "results": results}


@app.post("/api/webhooks/{hook_id}/test")
async def webhooks_test(hook_id: str, request: Request):
    """Send a signed test event to ONE webhook so the client can validate their receiver."""
    from app.db import get_active_webhooks, list_webhooks
    from app.services.webhooks import _deliver_one
    email = request.state.user_email
    hook = next((h for h in get_active_webhooks(email, "listing.generated") if h["id"] == hook_id), None)
    if not hook:
        if any(h["id"] == hook_id for h in list_webhooks(email)):
            raise HTTPException(400, "Webhook inactif ou non abonné à listing.generated")
        raise HTTPException(404, "Webhook introuvable")
    payload = {
        "event": "listing.generated", "test": True,
        "batch_id": "test_batch", "marketplace": "amazon_fr", "product_count": 1,
        "listings": [{"sku": "TEST-SKU", "title": "Produit de test SynqIO",
                      "bullet_points": ["Ceci est un événement de test"],
                      "description": "Événement envoyé depuis SynqIO pour valider votre intégration.",
                      "brand": "SynqIO", "category": "Test"}],
    }
    await _deliver_one(hook, "listing.generated", json.dumps(payload, ensure_ascii=False).encode())
    from app.db import list_webhooks as _lw
    status = next((h.get("last_status") for h in _lw(email) if h["id"] == hook_id), "")
    return {"sent": True, "last_status": status}


# ── Generated images persistence ──────────────────────────────────────────────

class SaveImagesRequest(BaseModel):
    images: list[dict]  # [{sku, slot, data_url}]


@app.post("/api/images/save")
async def api_save_images(req: SaveImagesRequest, request: Request):
    email = request.state.user_email
    saved = 0
    for img in req.images:
        sku = img.get("sku", "").strip()
        slot = img.get("slot", "").strip()
        url = img.get("data_url", "")
        if sku and slot and url.startswith("data:image"):
            save_generated_image(email, sku, slot, url)
            saved += 1
    return {"ok": True, "saved": saved}


@app.get("/api/images/{sku}")
async def api_get_images(sku: str, request: Request):
    return get_generated_images(request.state.user_email, sku)


class BulkImagesRequest(BaseModel):
    skus: list[str]


@app.post("/api/images/bulk")
async def api_get_images_bulk(req: BulkImagesRequest, request: Request):
    return get_generated_images_bulk(request.state.user_email, req.skus)


# ── Keyword suggestions (Amazon autocomplete proxy) ──────────────────────────

_KW_DOMAINS = {
    "amazon_fr": ("amazon.fr",     "A13V1IB3VIYZZH"),
    "amazon_de": ("amazon.de",     "A1PA6795UKMFR9"),
    "amazon_uk": ("amazon.co.uk",  "A1F83G8C2ARO7P"),
    "amazon_it": ("amazon.it",     "APJ6JRA9NG5V4"),
    "amazon_es": ("amazon.es",     "A1RKKUPIHCS9HS"),
    "amazon_com":("amazon.com",    "ATVPDKIKX0DER"),
}

@app.get("/api/keywords/suggest")
async def keywords_suggest(query: str, marketplace: str = "amazon_fr"):
    q = query.strip()
    if not q or len(q) < 2:
        return {"keywords": []}

    cache_key = f"{q.lower()}:{marketplace}"
    cached = _kw_cache.get(cache_key)
    if cached and (time.time() - cached["ts"]) < _KW_CACHE_TTL:
        return {"keywords": cached["keywords"]}

    domain, mid = _KW_DOMAINS.get(marketplace, ("amazon.fr", "A13V1IB3VIYZZH"))
    # Call autocomplete with base query + letter suffixes for more variety
    prefixes = [q] + [f"{q} {c}" for c in "abcdefghijklmnopqrstuvwxyz"[:8]]
    seen: set = set()
    keywords: list = []

    async with httpx.AsyncClient(timeout=6) as client:
        for prefix in prefixes:
            try:
                resp = await client.get(
                    f"https://completion.{domain}/api/2017/suggestions",
                    params={"mid": mid, "alias": "aps", "prefix": prefix, "limit": 11},
                    headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"},
                )
                if resp.is_success:
                    for s in resp.json().get("suggestions", []):
                        val = s.get("value", "").strip()
                        if val and val.lower() not in seen:
                            seen.add(val.lower())
                            keywords.append(val)
            except Exception:
                continue

    keywords = keywords[:30]
    # Prune old cache entries
    cutoff = time.time() - _KW_CACHE_TTL * 2
    for k in [k for k, v in _kw_cache.items() if v["ts"] < cutoff]:
        del _kw_cache[k]
    _kw_cache[cache_key] = {"ts": time.time(), "keywords": keywords}

    return {"keywords": keywords}


# ── A/B Testing ───────────────────────────────────────────────────────────────

class ABTestGenerateRequest(BaseModel):
    listing: dict
    marketplace: str = "amazon_fr"
    focus_keywords: Optional[List[str]] = None
    style_tone: str = "professionnel"


class ABTestExperimentCreate(BaseModel):
    exp_id: str
    asin: str = ""
    experiment_name: str = ""
    duration_days: int = 30


async def _run_ab_job(job_id: str, req: ABTestGenerateRequest, email: str):
    if job_id not in _jobs:
        _jobs[job_id] = {"status": "running", "progress": 0, "total": 2, "created_at": time.time()}
    else:
        _jobs[job_id]["status"] = "running"
    try:
        result = await generate_ab_variants(
            listing_data=req.listing,
            marketplace_str=req.marketplace,
            focus_keywords=req.focus_keywords or [],
            style_tone=req.style_tone,
        )
        _jobs[job_id].update({"status": "done", "result": result})
        if email:
            tokens = result.get("tokens", {})
            if tokens.get("input_tokens"):
                log_usage(email, "tokens_in", tokens["input_tokens"])
            if tokens.get("output_tokens"):
                log_usage(email, "tokens_out", tokens["output_tokens"])
    except Exception as e:
        log.error("ab_job failed", extra={"job_id": job_id, "error": str(e), "type": type(e).__name__})
        _jobs[job_id].update({"status": "failed", "error": f"{type(e).__name__}: {str(e)[:300]}"})


@app.post("/api/ab-test/generate")
async def ab_test_generate(req: ABTestGenerateRequest, request: Request):
    if not os.getenv("ANTHROPIC_API_KEY"):
        raise HTTPException(503, "ANTHROPIC_API_KEY manquante")
    email = getattr(request.state, "user_email", None)
    if email and not _rate_limit(f"{email}:ab_test", limit=10, window=3600):
        raise HTTPException(429, "Maximum 10 générations A/B par heure")
    job_id = str(uuid4())
    _jobs[job_id] = {"status": "pending", "progress": 0, "total": 2, "created_at": time.time()}
    _cleanup_jobs()
    asyncio.create_task(_run_ab_job(job_id, req, email or ""))
    return {"job_id": job_id}


@app.post("/api/ab-test/experiments")
async def ab_test_save_experiment(body: dict, request: Request):
    """Save a generated pair (variant A + B) and optionally push to Amazon MYE."""
    import uuid as _uuid
    email = request.state.user_email
    exp_id = str(_uuid.uuid4())
    sku = body.get("sku", "")
    marketplace = body.get("marketplace", "amazon_fr")
    name = (body.get("name") or f"Test A/B — {sku or 'SKU'}").strip()[:200]
    variant_a = body.get("variant_a", {})
    variant_b = body.get("variant_b", {})
    asin = (body.get("asin") or "").strip().upper()

    save_ab_experiment(exp_id, email, sku, asin, marketplace, name, variant_a, variant_b)

    amazon_result = None
    if asin and body.get("push_to_amazon"):
        try:
            amazon_result = await create_amazon_experiment(
                user_email=email,
                asin=asin,
                marketplace_str=marketplace,
                experiment_name=name,
                control_title=variant_a.get("title", ""),
                treatment_title=variant_b.get("title", ""),
                duration_days=int(body.get("duration_days", 30)),
            )
            update_ab_experiment(
                exp_id, email,
                amazon_experiment_id=amazon_result["experiment_id"],
                amazon_status=amazon_result["amazon_status"],
                asin=asin,
            )
        except ValueError as e:
            return {"ok": True, "id": exp_id, "amazon_error": str(e)}
        except Exception as e:
            return {"ok": True, "id": exp_id, "amazon_error": f"Erreur: {e}"}

    return {"ok": True, "id": exp_id, "amazon": amazon_result}


@app.get("/api/ab-test/experiments")
async def ab_test_list_experiments(request: Request):
    return list_ab_experiments(request.state.user_email)


@app.get("/api/ab-test/experiments/{exp_id}")
async def ab_test_get_experiment(exp_id: str, request: Request):
    email = request.state.user_email
    exp = get_ab_experiment(exp_id, email)
    if not exp:
        raise HTTPException(404, "Expérience introuvable")

    amazon_id = exp.get("amazon_experiment_id", "")
    if amazon_id:
        status_data = await get_amazon_experiment_status(email, amazon_id, exp.get("marketplace", "amazon_fr"))
        new_status = status_data.get("status") or status_data.get("amazon_status", "")
        if new_status and new_status != exp.get("amazon_status"):
            update_ab_experiment(exp_id, email, amazon_status=new_status)
            exp["amazon_status"] = new_status
        exp["amazon_data"] = status_data

    return exp


@app.delete("/api/ab-test/experiments/{exp_id}")
async def ab_test_delete_experiment(exp_id: str, request: Request):
    email = request.state.user_email
    exp = get_ab_experiment(exp_id, email)
    if not exp:
        raise HTTPException(404, "Expérience introuvable")

    amazon_id = exp.get("amazon_experiment_id", "")
    if amazon_id and exp.get("amazon_status") == "RUNNING":
        try:
            await cancel_amazon_experiment(email, amazon_id)
        except Exception:
            pass

    if not delete_ab_experiment(exp_id, email):
        raise HTTPException(404, "Expérience introuvable")
    return {"ok": True}


@app.patch("/api/ab-test/experiments/{exp_id}/winner")
async def ab_test_set_winner(exp_id: str, body: dict, request: Request):
    email = request.state.user_email
    winner = body.get("winner", "")
    if winner not in ("A", "B", ""):
        raise HTTPException(400, "winner doit être 'A', 'B' ou ''")
    if not update_ab_experiment(exp_id, email, winner=winner):
        raise HTTPException(404, "Expérience introuvable")
    return {"ok": True}


# ── Review monitoring ──────────────────────────────────────────────────────────

_REVIEWS_SYSTEM = """Tu es un expert Amazon listing. Analyse les avis clients négatifs fournis et retourne UNIQUEMENT un objet JSON valide, sans texte avant ni après, sans bloc markdown.

Structure JSON exacte :
{
  "summary": "<2 phrases max résumant les problèmes principaux>",
  "issues": [
    {
      "theme": "<nom court du problème, ex: Taille sous-estimée>",
      "frequency": "frequent" | "occasional" | "rare",
      "impact": "high" | "medium" | "low",
      "customer_words": "<verbatim court représentatif>",
      "field": "title" | "bullets" | "description" | "keywords" | "images",
      "suggestion": "<amélioration concrète et actionnable pour ce champ de la fiche>",
      "urgency": "critical" | "important" | "minor"
    }
  ]
}

Règles :
- frequency: "frequent" si ≥3 mentions similaires, "occasional" si 2, "rare" si 1
- field: champ de la fiche produit le plus pertinent à corriger pour ce problème
- suggestion: formule une action précise, ex: "Préciser dans les bullets : Taille réelle S = 38 FR — consultez le guide des tailles inclus"
- Trie les issues par urgency (critical d'abord) puis par impact (high d'abord)
- Retourne entre 2 et 8 issues maximum"""


class ReviewsAnalyzeRequest(BaseModel):
    reviews: str
    asin: str = ""
    marketplace: str = "amazon_fr"
    product_name: str = ""


async def _run_reviews_job(job_id: str, req: ReviewsAnalyzeRequest, email: str):
    if job_id not in _jobs:
        _jobs[job_id] = {"status": "running", "progress": 0, "total": 1, "created_at": time.time()}
    else:
        _jobs[job_id]["status"] = "running"
    try:
        from app.services.ai_agent import get_client
        import json as _json
        import re as _re
        lang = "français" if "fr" in req.marketplace else ("anglais" if "uk" in req.marketplace else "langue locale du marché")
        product_ctx = f"Produit : {req.product_name}\n" if req.product_name else ""
        asin_ctx = f"ASIN : {req.asin}\n" if req.asin else ""
        prompt = f"""{product_ctx}{asin_ctx}Marketplace : {req.marketplace} (langue : {lang})

Avis négatifs à analyser :
{req.reviews[:6000]}"""
        resp = await get_client().messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=2000,
            system=[{"type": "text", "text": _REVIEWS_SYSTEM, "cache_control": {"type": "ephemeral"}}],
            messages=[{"role": "user", "content": prompt}],
        )
        raw = resp.content[0].text.strip()
        raw = _re.sub(r'^```(?:json)?\s*', '', raw, flags=_re.MULTILINE)
        raw = _re.sub(r'\s*```$', '', raw, flags=_re.MULTILINE).strip()
        result = _json.loads(raw)
        _jobs[job_id].update({"status": "done", "progress": 1, "result": result})
        if email:
            tokens_in = getattr(resp.usage, "input_tokens", 0)
            tokens_out = getattr(resp.usage, "output_tokens", 0)
            if tokens_in: log_usage(email, "tokens_in", tokens_in)
            if tokens_out: log_usage(email, "tokens_out", tokens_out)
    except Exception as e:
        log.error("reviews_job failed", extra={"job_id": job_id, "error": str(e)})
        _jobs[job_id].update({"status": "failed", "error": f"{type(e).__name__}: {str(e)[:300]}"})


@app.post("/api/reviews/analyze")
async def reviews_analyze(req: ReviewsAnalyzeRequest, request: Request):
    if not os.getenv("ANTHROPIC_API_KEY"):
        raise HTTPException(503, "ANTHROPIC_API_KEY manquante")
    if not req.reviews.strip() or len(req.reviews.strip()) < 30:
        raise HTTPException(400, "Veuillez coller au moins quelques avis")
    email = getattr(request.state, "user_email", None)
    if email and not _rate_limit(f"{email}:reviews_analyze", limit=20, window=3600):
        raise HTTPException(429, "Maximum 20 analyses par heure")
    job_id = str(uuid4())
    _jobs[job_id] = {"status": "pending", "progress": 0, "total": 1, "created_at": time.time()}
    _cleanup_jobs()
    asyncio.create_task(_run_reviews_job(job_id, req, email or ""))
    return {"job_id": job_id}


@app.get("/api/reviews/fetch")
async def reviews_fetch(asin: str, marketplace: str = "amazon_fr"):
    """Scrape negative Amazon reviews (1★ + 2★) for a given ASIN."""
    import re as _re_rv
    asin = asin.strip().upper()
    if not asin or not _re_rv.match(r'^[A-Z0-9]{10}$', asin):
        raise HTTPException(400, "ASIN invalide (format attendu : B0XXXXXXXXX, 10 caractères)")

    domain = _KW_DOMAINS.get(marketplace, ("amazon.fr", "A13V1IB3VIYZZH"))[0]
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
        "Accept-Language": "fr-FR,fr;q=0.9,en-US;q=0.7,en;q=0.5",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Encoding": "gzip, deflate, br",
        "Connection": "keep-alive",
        "Upgrade-Insecure-Requests": "1",
    }

    reviews: list[str] = []
    blocked = False

    async with httpx.AsyncClient(timeout=20, headers=headers, follow_redirects=True) as client:
        for star_filter in ("one_star", "two_star"):
            for page in range(1, 4):
                url = (
                    f"https://www.{domain}/product-reviews/{asin}"
                    f"?filterByStar={star_filter}&pageNumber={page}&sortBy=recent"
                )
                try:
                    resp = await client.get(url)
                    if resp.status_code in (403, 503):
                        blocked = True
                        break
                    if not resp.is_success:
                        break
                    html = resp.text

                    # Detect anti-bot page
                    if any(kw in html.lower() for kw in ("captcha", "robot check", "automated access", "verify you're human")):
                        blocked = True
                        break

                    # Extract review body text
                    # Amazon wraps each review in: data-hook="review-body"...<span>text</span>
                    chunks = _re_rv.findall(
                        r'data-hook=["\']review-body["\'][^>]*>.*?<span[^>]*>(.*?)</span>',
                        html, _re_rv.DOTALL
                    )
                    page_reviews = []
                    for chunk in chunks:
                        text = _re_rv.sub(r'<[^>]+>', ' ', chunk)
                        text = _re_rv.sub(r'\s+', ' ', text).strip()
                        if text and len(text) > 15:
                            page_reviews.append(text)

                    if not page_reviews:
                        break  # No more reviews on this filter
                    reviews.extend(page_reviews)
                    if len(reviews) >= 60:
                        break
                except Exception:
                    break

            if blocked or len(reviews) >= 60:
                break

    if blocked:
        raise HTTPException(
            503,
            "BLOCKED: Amazon bloque les requêtes automatisées depuis ce serveur."
        )
    if not reviews:
        raise HTTPException(
            404,
            "Aucun avis 1★ ou 2★ trouvé pour ce produit (ASIN "
            + asin + "). Le produit n'a peut-être pas encore d'avis négatifs."
        )

    return {"reviews": reviews, "count": len(reviews), "asin": asin}


# ── Apify-backed review fetch (proxy infrastructure → no IP blocking) ────────────
# Direct scraping above gets blocked by Amazon from datacenter IPs. Apify runs the
# scrape through rotating residential proxies and returns clean review data.
# Configure via env: APIFY_TOKEN (required) + APIFY_REVIEWS_ACTOR (optional).

APIFY_REVIEWS_ACTOR = os.getenv("APIFY_REVIEWS_ACTOR", "junglee~amazon-reviews-scraper")

_APIFY_COUNTRY = {
    "fr": "FR", "de": "DE", "co.uk": "GB", "it": "IT",
    "es": "ES", "com": "US", "nl": "NL", "se": "SE", "pl": "PL",
}


def _apify_country_code(marketplace: str) -> str:
    domain = _KW_DOMAINS.get(marketplace, ("amazon.fr",))[0]
    suffix = domain.split("amazon.", 1)[-1] if "amazon." in domain else "fr"
    return _APIFY_COUNTRY.get(suffix, "FR")


def _apify_parse_rating(raw) -> Optional[float]:
    """Accept 5, '5', '5.0', '5,0 sur 5 étoiles', etc. → float or None."""
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    m = re.search(r'(\d+([.,]\d+)?)', str(raw))
    return float(m.group(1).replace(",", ".")) if m else None


def _apify_pick(item: dict, *keys):
    for k in keys:
        v = item.get(k)
        if v not in (None, "", []):
            return v
    return None


def _apify_build_payload(asins: list[str], marketplace: str, max_per: int) -> dict:
    domain = _KW_DOMAINS.get(marketplace, ("amazon.fr",))[0]
    product_urls = [{"url": f"https://www.{domain}/dp/{a}"} for a in asins]
    # Matches the Junglee "Amazon Reviews Scraper" schema (the default actor).
    # filterByRatings uses camelCase enums; marketplace is carried by URL domain.
    return {
        "productUrls": product_urls,
        "maxReviews": max_per,
        "filterByRatings": ["oneStar", "twoStar"],
        "sort": "recent",
        "reviewsUseProductVariantFilter": False,
        "scrapeProductDetails": True,
        "includeGdprSensitive": False,
        "proxyConfiguration": {"useApifyProxy": True},
    }


APIFY_PRODUCT_ACTOR = os.getenv("APIFY_PRODUCT_ACTOR", "junglee~amazon-product-details")


def _apify_parse_price(raw) -> Optional[float]:
    """Parse a price from many shapes: number, '19,99 €', '$1,299.00', {'value':..}."""
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return round(float(raw), 2)
    if isinstance(raw, dict):
        return _apify_parse_price(raw.get("value") or raw.get("amount") or raw.get("price"))
    m = re.search(r'\d[\d.,\s ]*', str(raw))
    if not m:
        return None
    num = m.group(0).replace(" ", "").replace(" ", "")
    if "," in num and "." in num:  # 1.299,00 (EU) or 1,299.00 (US)
        num = num.replace(".", "").replace(",", ".") if num.rfind(",") > num.rfind(".") else num.replace(",", "")
    elif "," in num:
        num = num.replace(",", ".")
    try:
        return round(float(num), 2)
    except ValueError:
        return None


def _apify_parse_products(items: list) -> dict:
    """Extract {asin: {asin, title, price, rating}} from a product-scraper dataset."""
    out: dict = {}
    for it in items or []:
        if not isinstance(it, dict):
            continue
        asin = str(_apify_pick(it, "asin", "productAsin", "asin1", "parentAsin") or "").upper().strip()
        if not asin or len(asin) < 5 or asin in out:
            continue
        name = _apify_pick(it, "title", "productTitle", "name", "productName", "product")
        if isinstance(name, dict):
            name = name.get("title") or name.get("name") or ""
        price = _apify_parse_price(_apify_pick(
            it, "price", "currentPrice", "priceAmount", "buyBoxPrice", "productPrice",
            "price_amount", "salePrice", "listPrice"))
        rating = _apify_parse_rating(_apify_pick(
            it, "productRating", "productOverallRating", "stars", "rating", "averageRating"))
        out[asin] = {"asin": asin, "title": str(name or "")[:200], "price": price, "rating": rating}
    return out


async def _apify_fetch_products(asins: list[str], marketplace: str) -> dict:
    """Scrape current product data (price/title/rating) for competitor ASINs."""
    token = os.getenv("APIFY_TOKEN")
    if not token:
        raise HTTPException(503, "APIFY_TOKEN manquant — configurez la clé Apify côté serveur.")
    domain = _KW_DOMAINS.get(marketplace, ("amazon.fr",))[0]
    payload = {
        "productUrls": [{"url": f"https://www.{domain}/dp/{a}"} for a in asins],
        "maxItems": len(asins),
        "proxyConfiguration": {"useApifyProxy": True},
    }
    base = "https://api.apify.com/v2"
    async with httpx.AsyncClient(timeout=120) as client:
        r = await client.post(f"{base}/acts/{APIFY_PRODUCT_ACTOR}/runs?token={token}", json=payload)
        if r.status_code in (401, 403):
            raise HTTPException(502, "APIFY_TOKEN invalide ou sans accès à cet acteur produit.")
        if r.status_code == 404:
            raise HTTPException(502, f"Acteur produit introuvable : '{APIFY_PRODUCT_ACTOR}' (APIFY_PRODUCT_ACTOR).")
        if not r.is_success:
            raise HTTPException(502, f"Apify a renvoyé {r.status_code} : {r.text[:200]}")
        data = (r.json() or {}).get("data", {})
        run_id, ds_id = data.get("id"), data.get("defaultDatasetId")
        status = data.get("status", "READY")
        if not run_id or not ds_id:
            raise HTTPException(502, "Réponse Apify inattendue au démarrage du run produit.")
        for _ in range(90):  # up to ~6 min
            if status == "SUCCEEDED":
                break
            if status in ("FAILED", "ABORTED", "TIMED-OUT", "TIMED_OUT", "TIMING-OUT"):
                raise HTTPException(502, f"Le run Apify produit s'est terminé en échec ({status}).")
            await asyncio.sleep(4)
            try:
                pr = await client.get(f"{base}/actor-runs/{run_id}?token={token}")
                status = (pr.json() or {}).get("data", {}).get("status", status)
            except httpx.HTTPError:
                continue
        else:
            raise HTTPException(504, "Le scraping produit Apify a dépassé le délai.")
        di = await client.get(f"{base}/datasets/{ds_id}/items?token={token}&clean=true&format=json")
        if not di.is_success:
            raise HTTPException(502, f"Lecture du dataset Apify impossible ({di.status_code}).")
        items = di.json()
    if not isinstance(items, list):
        items = items.get("items", []) if isinstance(items, dict) else []
    return _apify_parse_products(items)


def _apify_parse_int(raw) -> Optional[int]:
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return int(raw)
    m = re.search(r'[\d.,\s]+', str(raw))
    if not m:
        return None
    digits = re.sub(r'[^\d]', '', m.group(0))
    return int(digits) if digits else None


def _apify_parse_items(items: list) -> list[dict]:
    """Group raw Apify dataset items by ASIN. Keeps only 1★/2★ review text, but
    also captures the product-level overall rating and total review count
    (already present on each item thanks to scrapeProductDetails) so the UI can
    show the global score, not just the negatives."""
    by_asin: dict[str, dict] = {}
    for it in items or []:
        if not isinstance(it, dict):
            continue
        asin = str(_apify_pick(it, "asin", "productAsin", "asin1", "parentAsin") or "").upper().strip()
        if not asin or len(asin) < 5:
            continue
        name_raw = _apify_pick(
            it, "productTitle", "title", "productName", "name", "product",
        ) or asin
        if isinstance(name_raw, dict):
            name_raw = (name_raw.get("title") or name_raw.get("name")
                        or next((v for v in name_raw.values() if isinstance(v, str) and len(v) > 5), asin))
        name = str(name_raw).strip()
        slot = by_asin.setdefault(
            asin, {"asin": asin, "name": name or asin, "reviews": [],
                   "rating": None, "total_reviews": None})
        if not slot["name"] or slot["name"] == asin:
            slot["name"] = name or asin

        # Product-level score / volume — carried on every item via scrapeProductDetails
        if slot["rating"] is None:
            slot["rating"] = _apify_parse_rating(_apify_pick(
                it, "productRating", "productOverallRating", "overallRating",
                "averageRating", "productStarRating", "stars",
            ))
        if slot["total_reviews"] is None:
            slot["total_reviews"] = _apify_parse_int(_apify_pick(
                it, "countReviews", "reviewsCount", "productReviewsCount",
                "totalReviews", "countRatings", "ratingCount", "productTotalReviews",
            ))

        # Review text — keep only 1★/2★
        text = _apify_pick(
            it, "reviewDescription", "reviewText", "text", "review", "body",
            "content", "description", "reviewBody",
        )
        if not text or len(str(text).strip()) < 10:
            continue
        rating = _apify_parse_rating(_apify_pick(
            it, "ratingScore", "rating", "reviewRating", "starRating", "score",
        ))
        if rating is not None and rating > 2:
            continue  # keep only 1★/2★ when a rating is present
        slot["reviews"].append(str(text).strip())

    # Surface only products that actually have negatives to work on
    result = [p for p in by_asin.values() if p["reviews"]]
    return sorted(result, key=lambda p: len(p["reviews"]), reverse=True)


async def _apify_run_and_fetch(asins: list[str], marketplace: str, max_per: int,
                               progress_cb=None) -> list[dict]:
    """Start the actor asynchronously, poll the run to completion, then read the
    dataset. Avoids the run-sync 300s cap / gateway timeouts on long scrapes."""
    token = os.getenv("APIFY_TOKEN")
    if not token:
        raise HTTPException(503, "APIFY_TOKEN manquant — configurez la clé Apify côté serveur.")
    payload = _apify_build_payload(asins, marketplace, max_per)
    base = "https://api.apify.com/v2"
    async with httpx.AsyncClient(timeout=120) as client:
        # 1) Start the run
        try:
            r = await client.post(f"{base}/acts/{APIFY_REVIEWS_ACTOR}/runs?token={token}", json=payload)
        except httpx.HTTPError as e:
            raise HTTPException(502, f"Apify injoignable : {type(e).__name__}")
        if r.status_code in (401, 403):
            raise HTTPException(502, "APIFY_TOKEN invalide ou sans accès à cet acteur.")
        if r.status_code == 404:
            raise HTTPException(502, f"Acteur Apify introuvable : '{APIFY_REVIEWS_ACTOR}'. Vérifiez APIFY_REVIEWS_ACTOR.")
        if r.status_code == 400:
            raise HTTPException(502, f"Apify a refusé l'entrée (400) : {r.text[:200]}")
        if not r.is_success:
            raise HTTPException(502, f"Apify a renvoyé {r.status_code} : {r.text[:200]}")
        data = (r.json() or {}).get("data", {})
        run_id = data.get("id")
        ds_id = data.get("defaultDatasetId")
        status = data.get("status", "READY")
        if not run_id or not ds_id:
            raise HTTPException(502, "Réponse Apify inattendue au démarrage du run.")

        # 2) Poll the run (up to ~12 min — runs in the background, frontend polls the job)
        for _ in range(180):
            if status == "SUCCEEDED":
                break
            if status in ("FAILED", "ABORTED", "TIMED-OUT", "TIMED_OUT", "TIMING-OUT"):
                raise HTTPException(502, f"Le run Apify s'est terminé en échec ({status}).")
            await asyncio.sleep(4)
            try:
                pr = await client.get(f"{base}/actor-runs/{run_id}?token={token}")
                status = (pr.json() or {}).get("data", {}).get("status", status)
            except httpx.HTTPError:
                continue  # transient — keep polling
            if progress_cb:
                progress_cb(status)
        else:
            raise HTTPException(504, "Le scraping Apify a dépassé le délai. Réessayez avec moins d'ASIN.")

        # 3) Read the dataset
        di = await client.get(f"{base}/datasets/{ds_id}/items?token={token}&clean=true&format=json")
        if not di.is_success:
            raise HTTPException(502, f"Lecture du dataset Apify impossible ({di.status_code}).")
        try:
            items = di.json()
        except Exception:
            raise HTTPException(502, "Dataset Apify illisible (JSON attendu).")
    if not isinstance(items, list):
        items = items.get("items", []) if isinstance(items, dict) else []
    return _apify_parse_items(items)


async def _run_fetch_reviews_job(job_id: str, asins: list[str], marketplace: str, email: str):
    _jobs[job_id]["status"] = "running"
    try:
        def _cb(status):
            _jobs[job_id]["label"] = status
        products = await _apify_run_and_fetch(asins, marketplace, 50, _cb)
        total = sum(len(p["reviews"]) for p in products)
        _jobs[job_id].update({
            "status": "done", "progress": 1,
            "result": {"products": products, "count": total, "asins_found": len(products)},
        })
    except HTTPException as e:
        _jobs[job_id].update({"status": "failed", "error": str(e.detail)})
    except Exception as e:
        log.error("fetch_reviews_job failed", extra={"job_id": job_id, "error": str(e)})
        _jobs[job_id].update({"status": "failed", "error": f"{type(e).__name__}: {str(e)[:300]}"})


class ReviewsFetchBulkRequest(BaseModel):
    asins: list[str] = []
    marketplace: str = "amazon_fr"


@app.post("/api/reviews/fetch-bulk")
async def reviews_fetch_bulk(req: ReviewsFetchBulkRequest, request: Request):
    """Lance un job Apify (asynchrone) pour récupérer les avis 1★/2★ d'une liste
    d'ASIN. Renvoie un job_id à suivre via /api/jobs/{id}."""
    clean: list[str] = []
    for a in req.asins:
        a = (a or "").strip().upper()
        if re.match(r'^[A-Z0-9]{10}$', a):
            clean.append(a)
    clean = list(dict.fromkeys(clean))[:200]  # dédoublonne, plafonne à 200 ASIN/run
    if not clean:
        raise HTTPException(400, "Aucun ASIN valide (format attendu : 10 caractères, ex. B0XXXXXXXXX).")
    if not os.getenv("APIFY_TOKEN"):
        raise HTTPException(503, "APIFY_TOKEN manquant — configurez la clé Apify côté serveur.")

    email = getattr(request.state, "user_email", None)
    if email and not _rate_limit(f"{email}:reviews_fetch_bulk", limit=20, window=3600):
        raise HTTPException(429, "Trop de récupérations — réessayez dans une heure.")

    job_id = str(uuid4())
    _jobs[job_id] = {"status": "pending", "progress": 0, "total": 1, "created_at": time.time()}
    _cleanup_jobs()
    asyncio.create_task(_run_fetch_reviews_job(job_id, clean, req.marketplace, email or ""))
    return {"job_id": job_id}


@app.post("/api/reviews/parse-html")
async def reviews_parse_html(body: dict):
    """Extract review texts from raw HTML source pasted by the user (Ctrl+U fallback)."""
    import re as _re_rv
    html = body.get("html_content", "")
    if not html or len(html) < 200:
        raise HTTPException(400, "HTML trop court — collez bien l'intégralité de la source")

    seen: set = set()
    reviews: list[str] = []

    def _clean_add(raw: str):
        t = _re_rv.sub(r'<[^>]+>', ' ', raw)
        t = _re_rv.sub(r'\s+', ' ', t).strip()
        if t and len(t) > 20 and t.lower() not in seen:
            seen.add(t.lower())
            reviews.append(t)

    # Pattern 1: data-hook="review-body" — capture up to the OUTER closing </span></span>
    for m in _re_rv.finditer(
        r'data-hook=["\']review-body["\'][^>]*>(.*?)</span>\s*</span>',
        html, _re_rv.DOTALL
    ):
        _clean_add(m.group(1))

    # Pattern 2: class="review-text" (older Amazon layout)
    if not reviews:
        for m in _re_rv.finditer(
            r'class=["\'][^"\']*\breview-text\b[^"\']*["\'][^>]*>(.*?)</span>',
            html, _re_rv.DOTALL
        ):
            _clean_add(m.group(1))

    # Pattern 3: split-based last resort — scan raw text between data-hook markers
    if not reviews:
        parts = _re_rv.split(r'data-hook=["\']review-body["\']', html)
        for part in parts[1:]:
            gt = part.find('>')
            if gt < 0:
                continue
            chunk = part[gt + 1: gt + 6000]
            # Find approximate end at double-</span> boundary
            end = chunk.find('</span>\n')
            if end < 0:
                end = min(2500, len(chunk))
            _clean_add(chunk[:end])

    if not reviews:
        raise HTTPException(
            404,
            "Aucun avis trouvé dans ce code source. "
            "Assurez-vous d'avoir copié la source de la page /product-reviews/ (pas la page produit principale)."
        )

    return {"reviews": reviews, "count": len(reviews)}
